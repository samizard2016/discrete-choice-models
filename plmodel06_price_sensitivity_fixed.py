import pandas as pd
import numpy as np
from collections.abc import Iterable
import collections
collections.Iterable = Iterable  # Patch for pylogit Python 3.10 compatibility
import pylogit as pl
import statsmodels.api as sm
from collections import OrderedDict
import openpyxl
import os
import json

class DiscreteChoiceModel:
    def __init__(self, markets=None, use_pylogit=True, exclude_none=False, output_dir="cattle_feed_output"):
        self.markets = markets if markets else ['Maharashtra', 'Gujarat', 'Punjab', 'South']
        self.models = {}
        self.coef_dfs = {}
        self.use_pylogit = use_pylogit
        self.exclude_none = exclude_none
        self.brand_share = {}
        self.value_share = {}
        self.feature_importance = {}
        self.price_sensitivity = {}
        self.output_dir = output_dir
    
    def fit(self, df, target='Chosen', features=None):
        # Handle nan and data types
        df['CP'] = pd.to_numeric(df['CP'], errors='coerce').fillna(0.0)
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0.0)
        for col in ['FC', 'AH', 'VAS', 'Credit', 'herd_size', 'cattle_type']:
            if col in df.columns:
                df[col] = df[col].fillna('None').astype(str)
        
        # Dynamic feature selection
        mah_df = df[df['Market'] == 'Maharashtra']
        if features is None:
            features = ['Price', 'CP', 'FC', 'AH']
            if len(mah_df['VAS'].value_counts()) > 1 and mah_df['VAS'].isna().sum() / len(mah_df) < 0.2:
                features.append('VAS')
            if len(mah_df['Credit'].value_counts()) > 1 and mah_df['Credit'].isna().sum() / len(mah_df) < 0.2:
                features.append('Credit')
        
        # Exclude None choices
        if self.exclude_none:
            df = df[df['chosen_profile'] != 99].copy()
        
        # Calculate Brand Share
        chosen_df = df[df['Chosen'] == 1]
        self.brand_share = chosen_df.groupby(['Market', 'Brand']).size().reset_index(name='Choices')
        self.brand_share['Brand_Share'] = self.brand_share.groupby('Market')['Choices'].transform(lambda x: x / x.sum() if x.sum() > 0 else 0)
        
        # Calculate Value Share
        self.value_share = chosen_df.groupby(['Market', 'Brand'])['Price'].sum().reset_index(name='Total_Price')
        self.value_share['Value_Share'] = self.value_share.groupby('Market')['Total_Price'].transform(lambda x: x / x.sum() if x.sum() > 0 else 0)
        
        for market in self.markets:
            market_df = df[df['Market'] == market].copy()
            if market_df.empty:
                print(f"Warning: No data for {market}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                continue
            
            # Verify columns
            missing_cols = [col for col in features if col not in market_df.columns]
            if missing_cols:
                print(f"Warning: Missing columns {missing_cols} for {market}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                continue
            
            if self.use_pylogit:
                try:
                    # Pylogit setup
                    pylogit_df = market_df.copy()
                    pylogit_df['obs_id'] = pylogit_df['ChoiceID']
                    pylogit_df['alt_id'] = pylogit_df['Brand']
                    pylogit_df['choice'] = pylogit_df[target]
                    
                    # Specification with OrderedDict
                    spec = OrderedDict([
                        ('Price', 'all_same'),
                        ('CP', 'all_same'),
                        ('FC', 'all_different'),
                        ('AH', 'all_different')
                    ])
                    if 'VAS' in features:
                        spec['VAS'] = 'all_different'
                    if 'Credit' in features:
                        spec['Credit'] = 'all_different'
                    varnames = []
                    for var in features:
                        if var in ['Price', 'CP']:
                            varnames.append(var)
                        else:
                            unique_vals = market_df[var].unique()
                            varnames.extend([f"{var}_{val}" for val in unique_vals if val != 'None'])
                    name_dict = OrderedDict([(var, var) for var in varnames])
                    
                    # Fit pylogit
                    model = pl.create_choice_model(
                        data=pylogit_df,
                        alt_id_col='alt_id',
                        obs_id_col='obs_id',
                        choice_col='choice',
                        specification=spec,
                        model_type='MNL',
                        names=name_dict
                    )
                    init_vals = np.zeros(len(varnames))
                    bounds = [(-10, 10) for _ in varnames]
                    results = model.fit_mle(init_vals, bounds=bounds, print_res=False)
                    coef_df = pd.DataFrame({
                        'Coefficient': results.params,
                        'P-value': results.pvalues
                    }, index=varnames)
                    self.models[market] = results
                    self.coef_dfs[market] = coef_df
                    print(f"Pylogit model fitted for {market}")
                except Exception as e:
                    print(f"Pylogit failed for {market}: {e}. Switching to statsmodels.")
                    self.use_pylogit = False
            
            if not self.use_pylogit:
                try:
                    # Statsmodels setup
                    categorical_cols = [col for col in ['FC', 'AH', 'VAS', 'Credit'] if col in features and col in market_df.columns]
                    market_df_encoded = pd.get_dummies(market_df, columns=categorical_cols, drop_first=True)
                    feature_cols = [col for col in market_df_encoded.columns if col.startswith(('Price', 'CP') + tuple(categorical_cols))]
                    if not feature_cols:
                        print(f"Warning: No valid features after encoding for {market}")
                        self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                        self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                        self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                        continue
                    X = market_df_encoded[feature_cols].astype(float).dropna()
                    y = market_df_encoded.loc[X.index, target]
                    
                    # Add constant
                    X = sm.add_constant(X)
                    
                    # Fit statsmodels
                    model = sm.Logit(y, X).fit_regularized(method='l1', alpha=0.01, disp=0, maxiter=1000)
                    coef_df = pd.DataFrame({
                        'Coefficient': model.params,
                        'P-value': model.pvalues.fillna(1.0)
                    }, index=X.columns)
                    self.models[market] = model
                    self.coef_dfs[market] = coef_df
                    print(f"Statsmodels model fitted for {market}")
                except Exception as e:
                    print(f"Statsmodels failed for {market}: {e}")
                    self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                    self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                    self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
            
            # Calculate feature importance
            importance = self.get_feature_importance(market)
            self.feature_importance[market] = pd.DataFrame({
                'Feature': importance.index,
                'Importance': importance.values
            })
            
            # Simulate price sensitivity for Cargill
            self.price_sensitivity[market] = self.simulate_price_sensitivity(market_df, market, 'Cargill', features)
        
        return self
    
    def get_attribute_utilities(self, market):
        coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
        return coef_df['Coefficient'] if 'Coefficient' in coef_df.columns else pd.Series(dtype=float)
    
    def get_feature_importance(self, market):
        utilities = self.get_attribute_utilities(market)
        if utilities.empty:
            return pd.Series(dtype=float)
        importance = utilities.abs() / utilities.abs().sum() if utilities.abs().sum() > 0 else utilities
        return importance
    
    def simulate_price_sensitivity(self, market_df, market, brand='Cargill', features=None):
        if market not in self.models or isinstance(self.models[market], pd.DataFrame):
            print(f"Warning: No valid model for {market}. Skipping price sensitivity analysis.")
            return pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
        
        # Get coefficients
        coefs = self.get_attribute_utilities(market)
        if coefs.empty:
            print(f"Warning: No coefficients for {market}. Skipping price sensitivity analysis.")
            return pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
        
        # Prepare data for prediction
        market_df = market_df.copy()
        choice_sets = market_df.groupby('ChoiceID')
        results = []
        
        # Price change range: -10% to +10% in 1% increments
        price_changes = np.arange(-0.10, 0.11, 0.01)
        
        for change in price_changes:
            utilities = []
            
            for choice_id, group in choice_sets:
                # Validate choice set size
                if len(group) < 1:
                    print(f"Warning: Empty choice set {choice_id} in {market}. Skipping.")
                    continue
                
                group = group.copy()
                # Adjust Cargill's price
                cargill_mask = group['Brand'] == brand
                if cargill_mask.any():
                    base_price = group.loc[cargill_mask, 'Price'].iloc[0]
                    group.loc[cargill_mask, 'Price'] = base_price * (1 + change)
                
                # Compute utilities for each alternative
                group_utilities = []
                for _, row in group.iterrows():
                    utility = 0
                    for var in coefs.index:
                        if var == 'Price':
                            utility += coefs[var] * row['Price']
                        elif var == 'CP':
                            utility += coefs[var] * row['CP']
                        elif var.startswith(('FC_', 'AH_', 'VAS_', 'Credit_')):
                            var_name, var_value = var.split('_', 1)
                            if row[var_name] == var_value:
                                utility += coefs[var]
                    group_utilities.append(utility)
                
                # Compute choice probabilities (softmax)
                exp_utilities = np.exp(group_utilities)
                probs = exp_utilities / exp_utilities.sum() if exp_utilities.sum() > 0 else np.ones(len(group_utilities)) / len(group_utilities)
                
                # Assign probabilities to alternatives
                for i, (_, row) in enumerate(group.iterrows()):
                    utilities.append({
                        'ChoiceID': choice_id,
                        'Brand': row['Brand'],
                        'Probability': probs[i],
                        'Price': row['Price']
                    })
            
            # Aggregate probabilities and prices
            if utilities:
                util_df = pd.DataFrame(utilities)
                brand_share = util_df[util_df['Brand'] == brand]['Probability'].mean()
                value_sum = (util_df['Probability'] * util_df['Price']).sum()
                brand_value = (util_df[util_df['Brand'] == brand]['Probability'] * util_df[util_df['Brand'] == brand]['Price']).sum()
                value_share = brand_value / value_sum if value_sum > 0 else 0
                
                results.append({
                    'Price_Change_Percent': change * 100,
                    'Brand_Share': brand_share,
                    'Value_Share': value_share
                })
        
        return pd.DataFrame(results) if results else pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
    
    def add_model_summaries_to_workbook(self, workbook):
        for market in self.markets:
            # Create or update summary sheet
            if f"{market}_summary" in workbook.sheetnames:
                sheet = workbook[f"{market}_summary"]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sheet = workbook.create_sheet(title=f"{market}_summary")
            
            # Write coefficients and p-values
            coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
            sheet['A1'] = 'Feature'
            sheet['B1'] = 'Coefficient'
            sheet['C1'] = 'P-value'
            for r, (index, row) in enumerate(coef_df.iterrows(), start=2):
                sheet[f'A{r}'] = index
                sheet[f'B{r}'] = row.get('Coefficient', float('nan'))
                sheet[f'C{r}'] = row.get('P-value', float('nan'))
            
            # Append feature importance
            importance_df = self.feature_importance.get(market, pd.DataFrame(columns=['Feature', 'Importance']))
            if not importance_df.empty:
                start_row = len(coef_df) + 4
                sheet[f'A{start_row}'] = 'Feature Importance'
                sheet[f'A{start_row + 1}'] = 'Feature'
                sheet[f'B{start_row + 1}'] = 'Importance'
                for r, (index, row) in enumerate(importance_df.iterrows(), start=start_row + 2):
                    sheet[f'A{r}'] = row['Feature']
                    sheet[f'B{r}'] = row['Importance']
            
            # Create brand share sheet
            if f"{market}_brand_share" in workbook.sheetnames:
                brand_sheet = workbook[f"{market}_brand_share"]
                for row in brand_sheet.iter_rows(min_row=1, max_row=brand_sheet.max_row, min_col=1, max_col=brand_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                brand_sheet = workbook.create_sheet(title=f"{market}_brand_share")
            brand_share_market = self.brand_share[self.brand_share['Market'] == market][['Brand', 'Brand_Share']]
            brand_sheet['A1'] = 'Brand'
            brand_sheet['B1'] = 'Brand_Share'
            for r, (index, row) in enumerate(brand_share_market.iterrows(), start=2):
                brand_sheet[f'A{r}'] = row['Brand']
                brand_sheet[f'B{r}'] = row['Brand_Share']
            
            # Create value share sheet
            if f"{market}_value_share" in workbook.sheetnames:
                value_sheet = workbook[f"{market}_value_share"]
                for row in value_sheet.iter_rows(min_row=1, max_row=value_sheet.max_row, min_col=1, max_col=value_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                value_sheet = workbook.create_sheet(title=f"{market}_value_share")
            value_share_market = self.value_share[self.value_share['Market'] == market][['Brand', 'Value_Share']]
            value_sheet['A1'] = 'Brand'
            value_sheet['B1'] = 'Value_Share'
            for r, (index, row) in enumerate(value_share_market.iterrows(), start=2):
                value_sheet[f'A{r}'] = row['Brand']
                value_sheet[f'B{r}'] = row['Value_Share']
            
            # Create price sensitivity sheet
            if f"{market}_price_sensitivity" in workbook.sheetnames:
                sensitivity_sheet = workbook[f"{market}_price_sensitivity"]
                for row in sensitivity_sheet.iter_rows(min_row=1, max_row=sensitivity_sheet.max_row, min_col=1, max_col=sensitivity_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sensitivity_sheet = workbook.create_sheet(title=f"{market}_price_sensitivity")
            sensitivity_df = self.price_sensitivity.get(market, pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']))
            sensitivity_sheet['A1'] = 'Price_Change_Percent'
            sensitivity_sheet['B1'] = 'Brand_Share'
            sensitivity_sheet['C1'] = 'Value_Share'
            for r, (index, row) in enumerate(sensitivity_df.iterrows(), start=2):
                sensitivity_sheet[f'A{r}'] = row['Price_Change_Percent']
                sheet[f'B{r}'] = row['Brand_Share']
                sheet[f'C{r}'] = row['Value_Share']
            # Add note about chart
            sensitivity_sheet['A{0}'.format(len(sensitivity_df) + 4)] = 'Note: Line chart saved as HTML file at {0}/{1}_price_sensitivity.html'.format(self.output_dir, market)
    
    def generate_price_sensitivity_charts(self):
        os.makedirs(self.output_dir, exist_ok=True)
        for market in self.markets:
            sensitivity_df = self.price_sensitivity.get(market, pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']))
            if sensitivity_df.empty:
                print(f"Warning: No price sensitivity data for {market}. Skipping chart generation.")
                continue
            
            # Prepare data for Recharts
            chart_data = sensitivity_df[['Price_Change_Percent', 'Brand_Share', 'Value_Share']].to_dict(orient='records')
            chart_data_json = json.dumps(chart_data)
            
            # Generate HTML with Recharts
            html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>{market} Price Sensitivity Chart</title>
    <script src="https://cdn.jsdelivr.net/npm/react@18.2.0/umd/react.production.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/react-dom@18.2.0/umd/react-dom.production.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/recharts@2.12.7/dist/recharts.umd.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/@babel/standalone@7.24.7/babel.min.js"></script>
    <style>
        body {{ font-family: Arial, sans-serif; display: flex; justify-content: center; align-items: center; height: 100vh; margin: 0; background-color: #f0f0f0; }}
        #chart {{ width: 800px; height: 500px; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }}
    </style>
</head>
<body>
    <div id="chart"></div>
    <script type="text/babel">
        const {{ LineChart, Line, XAxis, YAxis, CartesianGrid, Tooltip, Legend }} = Recharts;
        const data = {chart_data_json};

        const App = () => (
            <LineChart width={800} height={500} data={{data}} margin={{ top: 20, right: 30, left: 20, bottom: 10 }}>
                <CartesianGrid strokeDasharray="3 3" />
                <XAxis
                    dataKey="Price_Change_Percent"
                    label={{ value: 'Price Change (%)', position: 'insideBottom', offset: -5 }}
                    domain={[-10, 10]}
                    ticks={[-10, -8, -6, -4, -2, 0, 2, 4, 6, 8, 10]}
                />
                <YAxis
                    label={{ value: 'Share', angle: -90, position: 'insideLeft' }}
                    domain={[0, 'auto']}
                />
                <Tooltip formatter={{(value) => (value * 100).toFixed(2) + '%'}} />
                <Legend verticalAlign="top" height={36} />
                <Line type="monotone" dataKey="Brand_Share" name="Brand Share" stroke="#8884d8" strokeWidth={2} />
                <Line type="monotone" dataKey="Value_Share" name="Value Share" stroke="#82ca9d" strokeWidth={2} />
            </LineChart>
        );

        ReactDOM.render(<App />, document.getElementById('chart'));
    </script>
</body>
</html>
"""
            # Save HTML file
            chart_path = os.path.join(self.output_dir, f"{market}_price_sensitivity.html")
            with open(chart_path, 'w') as f:
                f.write(html_content)
            print(f"Saved price sensitivity chart for {market} to {chart_path}")

if __name__ == "__main__":
    # Load data
    df = pd.read_csv("combined_choice_data_v2.csv")
    
    # Initialize workbook
    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    
    # Fit models and save summaries
    model = DiscreteChoiceModel(markets=['Maharashtra', 'Gujarat', 'Punjab', 'South'])
    model.fit(df)
    model.add_model_summaries_to_workbook(workbook)
    model.generate_price_sensitivity_charts()
    workbook.save('cargill_findings_v5.xlsx')