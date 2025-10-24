import pandas as pd
import numpy as np
from collections.abc import Iterable
import collections
collections.Iterable = Iterable
import pylogit as pl
from statsmodels.stats.outliers_influence import variance_inflation_factor
from collections import OrderedDict
import openpyxl
import os
import logging
import scipy
from scipy.optimize import minimize
from scipy import stats

# Set up logging
logging.basicFileName = 'cattle_feed_model.log'
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
    logging.FileHandler(logging.basicFileName),
    logging.StreamHandler()
])
logging.info(f"Using scipy version: {scipy.__version__}")

class DiscreteChoiceModel:
    def __init__(self, markets=None, use_pylogit=True, output_dir="cattle_feed_output"):
        self.markets = markets if markets else ['Maharashtra', 'Gujarat', 'Punjab', 'South']
        self.models = {}
        self.coef_dfs = {}
        self.use_pylogit = use_pylogit
        self.brand_share = {}
        self.value_share = {}
        self.output_dir = output_dir
        self.choice_id_to_num = {}
        os.makedirs(self.output_dir, exist_ok=True)

    def check_multicollinearity(self, df, predictors):
        """Check for multicollinearity using VIF."""
        try:
            X = df[predictors].astype(float).fillna(0)
            if not np.isfinite(X).all().all():
                logging.warning("NaN or infinite values found in predictors. Replacing with 0.")
                X = X.replace([np.inf, -np.inf], 0).fillna(0)
            vif_data = pd.DataFrame()
            vif_data["Variable"] = predictors
            vif_data["VIF"] = [variance_inflation_factor(X.values, i) for i in range(X.shape[1])]
            logging.info(f"VIF values for {predictors}:\n{vif_data}")
            return vif_data[vif_data["VIF"] > 100]["Variable"].tolist()
        except Exception as e:
            logging.error(f"VIF calculation failed: {str(e)}. Proceeding without removing predictors.")
            return []

    def preprocess_data(self, df):
        """Preprocess the dataset for choice modeling."""
        # Store original Price for value share
        df['Original_Price'] = df['Price'].copy()
        
        # Create ChoiceID
        df['ChoiceID'] = df['Respondent_ID'].astype(str) + '_' + df['choice_set'].astype(str)

        # Handle duplicate brands in choice sets by keeping the first instance
        duplicates = df.groupby(['ChoiceID', 'Brand']).size().reset_index(name='count')
        duplicates = duplicates[duplicates['count'] > 1]
        if not duplicates.empty:
            logging.info(f"Found {len(duplicates)} choice sets with duplicate brands:\n{duplicates}")
            df = df.drop_duplicates(subset=['ChoiceID', 'Brand'], keep='first')
            logging.info(f"Kept first instance of {len(duplicates)} duplicate choice sets")

        # Remove NaN
        initial_len = len(df)
        df = df[df['Brand'].notna() & df['ChoiceID'].notna()].copy()
        if len(df) < initial_len:
            logging.info(f"Removed {initial_len - len(df)} rows with NaN in Brand or ChoiceID")

        # Ensure binary Chosen
        invalid_chosen = df[~df['Chosen'].isin([0, 1])]
        if not invalid_chosen.empty:
            logging.warning(f"Found {len(invalid_chosen)} rows with invalid Chosen values")
            df = df[df['Chosen'].isin([0, 1])]

        # Validate Chosen sums
        chosen_sums = df.groupby('ChoiceID')['Chosen'].sum()
        invalid_choice_ids = chosen_sums[chosen_sums != 1].index
        if len(invalid_choice_ids) > 0:
            logging.warning(f"Found {len(invalid_choice_ids)} choice sets with invalid Chosen sum (not 1)")
            df = df[~df['ChoiceID'].isin(invalid_choice_ids)]
            logging.info(f"Dropped {len(invalid_choice_ids)} choice sets with invalid Chosen sums")

        # Convert Price/CP to numeric and scale
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(df['Price'].mean())
        df['CP'] = pd.to_numeric(df['CP'], errors='coerce').fillna(df['CP'].mean())
        df['Price'] = (df['Price'] - df['Price'].mean()) / df['Price'].std()
        df['CP'] = (df['CP'] - df['CP'].mean()) / df['CP'].std()

        # Add synthetic variation with reduced scale
        if df['Price'].nunique() <= 5:
            logging.warning("Limited unique Price values. Adding synthetic variation (scale: z-score).")
            df['Price'] += np.random.normal(0, 0.005, len(df))
            df['Price'] = df['Price'].clip(-3, 3)
            logging.info(f"Scaled Price range: [{df['Price'].min():.6f}, {df['Price'].max():.6f}]")
        if df['CP'].nunique() <= 5:
            logging.warning("Limited unique CP values. Adding synthetic variation (scale: z-score).")
            df['CP'] += np.random.normal(0, 0.005, len(df))
            df['CP'] = df['CP'].clip(-3, 3)
            logging.info(f"Scaled CP range: [{df['CP'].min():.6f}, {df['CP'].max():.6f}]")

        # Group low-frequency brands
        brand_counts = df['Brand'].value_counts(normalize=True)
        major_brands = brand_counts[brand_counts >= 0.05].index.tolist()
        df['Brand'] = df['Brand'].where(df['Brand'].isin(major_brands), 'Other')
        logging.info(f"Major brands (share >= 5%): {major_brands}")
        logging.info(f"Brands grouped into 'Other': {set(df['Brand'].unique()) - set(major_brands)}")

        # Validate choice set size
        choice_counts = df.groupby('ChoiceID').size()
        invalid_choices = choice_counts[choice_counts < 2].index
        if len(invalid_choices) > 0:
            logging.warning(f"Found {len(invalid_choices)} choice sets with fewer than 2 alternatives")
            df = df[~df['ChoiceID'].isin(invalid_choices)]
            logging.info(f"Dropped {len(invalid_choices)} choice sets with insufficient alternatives")

        return df

    def balance_data(self, df, market, target):
        """Balance Cargill vs. non-Cargill choices for specified markets."""
        if market in ['Punjab', 'Maharashtra', 'Gujarat', 'South']:
            cargill_chosen = df[(df['Brand'] == 'Cargill') & (df[target] == 1)]
            cargill_not_chosen = df[(df['Brand'] == 'Cargill') & (df[target] == 0)]
            non_cargill_chosen = df[(df['Brand'] != 'Cargill') & (df[target] == 1)]
            non_cargill_not_chosen = df[(df['Brand'] != 'Cargill') & (df[target] == 0)]
            n_samples = min(len(cargill_chosen), len(cargill_not_chosen), len(non_cargill_chosen), len(non_cargill_not_chosen))
            if n_samples < 10:
                logging.warning(f"Insufficient data for balancing {market}: {n_samples} samples")
                return df
            cargill_chosen = cargill_chosen.sample(n=n_samples, random_state=42)
            cargill_not_chosen = cargill_not_chosen.sample(n=n_samples, random_state=42)
            non_cargill_chosen = non_cargill_chosen.sample(n=n_samples, random_state=42)
            non_cargill_not_chosen = non_cargill_not_chosen.sample(n=n_samples, random_state=42)
            df = pd.concat([cargill_chosen, cargill_not_chosen, non_cargill_chosen, non_cargill_not_chosen])
            df = df.sort_values(['ChoiceID', 'choice_set'])
            logging.info(f"Balanced {market} data: {len(cargill_chosen)} Cargill chosen, {len(cargill_not_chosen)} not chosen, "
                         f"{len(non_cargill_chosen)} non-Cargill chosen, {len(non_cargill_not_chosen)} non-Cargill not chosen")
        return df

    def custom_mnl_loglikelihood(self, params, X, y, obs_id_col, alt_id_col, num_alts=3):
        """Custom MNL log-likelihood function with L2 regularization."""
        params = np.array(params)
        X = np.atleast_2d(X)
        y = np.array(y)
        obs_ids = np.unique(obs_id_col)
        log_likelihood = 0
        l2_lambda = 1.0

        for obs_id in obs_ids:
            mask = obs_id_col == obs_id
            X_obs = X[mask]
            y_obs = y[mask]
            utilities = np.dot(X_obs, params)
            exp_utilities = np.exp(utilities - np.max(utilities))
            probs = exp_utilities / np.sum(exp_utilities)
            log_likelihood += np.sum(y_obs * np.log(probs + 1e-10))

        log_likelihood -= l2_lambda * np.sum(params ** 2)
        return -log_likelihood

    def compute_std_errors(self, params, X, y, obs_id_col, alt_id_col, num_alts=3):
        """Compute standard errors via numerical Hessian."""
        try:
            X = np.atleast_2d(X)
            y = np.array(y)
            def neg_log_likelihood(p):
                return self.custom_mnl_loglikelihood(p, X, y, obs_id_col, alt_id_col, num_alts)
            
            n_params = len(params)
            hessian = np.zeros((n_params, n_params))
            eps = 1e-6
            for i in range(n_params):
                for j in range(n_params):
                    params_plus_i = params.copy()
                    params_minus_i = params.copy()
                    params_plus_j = params.copy()
                    params_minus_j = params.copy()
                    
                    params_plus_i[i] += eps
                    params_minus_i[i] -= eps
                    params_plus_j[j] += eps
                    params_minus_j[j] -= eps
                    
                    if i == j:
                        f_plus = neg_log_likelihood(params_plus_i)
                        f_minus = neg_log_likelihood(params_minus_i)
                        f_center = neg_log_likelihood(params)
                        hessian[i, i] = (f_plus + f_minus - 2 * f_center) / (eps ** 2)
                    else:
                        params_plus_plus = params.copy()
                        params_plus_minus = params.copy()
                        params_minus_plus = params.copy()
                        params_minus_minus = params.copy()
                        
                        params_plus_plus[i] += eps
                        params_plus_plus[j] += eps
                        params_plus_minus[i] += eps
                        params_plus_minus[j] -= eps
                        params_minus_plus[i] -= eps
                        params_minus_plus[j] += eps
                        params_minus_minus[i] -= eps
                        params_minus_minus[j] -= eps
                        
                        f_pp = neg_log_likelihood(params_plus_plus)
                        f_pm = neg_log_likelihood(params_plus_minus)
                        f_mp = neg_log_likelihood(params_minus_plus)
                        f_mm = neg_log_likelihood(params_minus_minus)
                        hessian[i, j] = (f_pp - f_pm - f_mp + f_mm) / (4 * eps ** 2)
                        hessian[j, i] = hessian[i, j]
            
            inv_hessian = np.linalg.inv(hessian)
            std_errors = np.sqrt(np.maximum(np.diag(inv_hessian), 0))
            return std_errors
        except Exception as e:
            logging.error(f"Standard error calculation failed: {str(e)}")
            return np.array([np.nan] * len(params))

    def fit(self, df, target='Chosen', features=None):
        """Fit choice models for each market."""
        df = self.preprocess_data(df)
        if features is None:
            features = ['Price', 'CP', 'Brand_Cargill']
        
        # Calculate brand and value shares before dummy variables
        for market in self.markets:
            market_df = df[df['Market'] == market].copy()
            if len(market_df) < 100:
                logging.warning(f"Insufficient data for {market}: {len(market_df)} rows")
                self.models[market] = None
                self.coef_dfs[market] = pd.DataFrame()
                continue
            market_df = self.balance_data(market_df, market, target)
            self.brand_share[market] = market_df[market_df[target] == 1].groupby('Brand').size() / market_df[market_df[target] == 1].shape[0]
            self.brand_share[market] = self.brand_share[market].reset_index()
            self.brand_share[market].columns = ['Brand', 'Actual_Brand_Share']
            self.value_share[market] = market_df.groupby('Brand').apply(
                lambda x: (x['Original_Price'] * x[target]).sum() / x['Original_Price'].sum() if x['Original_Price'].sum() > 0 else 0,
                include_groups=False
            ).reset_index()
            self.value_share[market].columns = ['Brand', 'Actual_Value_Share']
            self.value_share[market]['Actual_Value_Share'] = self.value_share[market]['Actual_Value_Share'].clip(0, 1)
            logging.info(f"Brand shares in {market}:\n{self.brand_share[market]}")
            logging.info(f"Value shares in {market}:\n{self.value_share[market]}")

        # Create dummy variables after balancing
        df = pd.get_dummies(df, columns=['FC', 'AH', 'VAS', 'Credit'], 
                           prefix=['FC', 'AH', 'VAS', 'Credit'], 
                           drop_first=True)
        df = pd.get_dummies(df, columns=['Brand'], prefix='Brand', drop_first=True)

        # Ensure dummy variables are numeric
        dummy_columns = [col for col in df.columns if col.startswith(('FC_', 'AH_', 'VAS_', 'Credit_', 'Brand_'))]
        for col in dummy_columns:
            df[col] = df[col].astype(float)

        # VIF check after dummy variables
        logging.info(f"Selected features for VIF check: {features}")
        high_vif = self.check_multicollinearity(df, features)
        features = [f for f in features if f not in high_vif]

        # Map ChoiceID and alt_id
        unique_choice_ids = df['ChoiceID'].unique()
        self.choice_id_to_num = {cid: idx + 1 for idx, cid in enumerate(unique_choice_ids)}
        df['obs_id'] = df['ChoiceID'].map(self.choice_id_to_num)
        df['alt_id'] = df.groupby('ChoiceID').cumcount() + 1
        df = df.sort_values(['obs_id', 'alt_id'])
        logging.info(f"ChoiceID to numeric mapping (first 10): {dict(list(self.choice_id_to_num.items())[:10])}")
        logging.info(f"Unique ChoiceID values (first 10): {list(unique_choice_ids[:10])}")

        for market in self.markets:
            if market not in self.models:
                logging.info(f"Skipping model fitting for {market} due to insufficient data")
                continue
            logging.info(f"Starting model fitting for {market}")
            market_df = df[df['Market'] == market].copy()
            
            # Validate choice sets
            choice_counts = market_df.groupby('obs_id').size()
            valid_obs = choice_counts[choice_counts >= 3].index
            market_df = market_df[market_df['obs_id'].isin(valid_obs)]
            logging.info(f"{market} has {len(valid_obs)} valid choice sets with >= 3 alternatives")
            if len(valid_obs) < 50:
                logging.warning(f"Insufficient valid choice sets for {market}: {len(valid_obs)}")
                self.models[market] = None
                self.coef_dfs[market] = pd.DataFrame()
                continue

            # Check Brand_Cargill variation
            cargill_var = market_df['Brand_Cargill'].nunique()
            if cargill_var < 2:
                logging.warning(f"Insufficient variation in Brand_Cargill for {market}")
                self.models[market] = None
                self.coef_dfs[market] = pd.DataFrame()
                continue

            # Log data summary
            logging.info(f"{market} data shape: {market_df.shape}")
            logging.info(f"{market} data summary:\n{market_df[features + [target]].describe()}")

            # Reconstruct Brand column for logging
            brand_columns = [col for col in df.columns if col.startswith('Brand_')]
            brand_names = [col.replace('Brand_', '') for col in brand_columns] + ['Other']
            market_df['Brand'] = pd.Series(np.zeros(len(market_df)), dtype=str)
            for i, row in market_df.iterrows():
                for brand, col in zip(brand_names, brand_columns + ['Other']):
                    if col == 'Other' and all(row[brand_columns] == 0):
                        market_df.at[i, 'Brand'] = 'Other'
                    elif col != 'Other' and row[col] == 1:
                        market_df.at[i, 'Brand'] = brand

            # Check Cargill choices
            cargill_choices = market_df[market_df['Brand'] == 'Cargill'][target].sum()
            logging.info(f"Cargill choices in {market}: {cargill_choices}")

            # Try MNL with full features
            spec = OrderedDict()
            names = OrderedDict()
            for f in features:
                spec[f] = 'all_same'
                names[f] = f

            if self.use_pylogit:
                try:
                    model = pl.create_choice_model(
                        data=market_df,
                        alt_id_col='alt_id',
                        obs_id_col='obs_id',
                        choice_col=target,
                        specification=spec,
                        model_type='MNL',
                        names=names
                    )
                    X = market_df[features].astype(float)
                    rank = np.linalg.matrix_rank(X)
                    logging.info(f"Design matrix rank for {market}: {rank} (features: {len(features)})")
                    if rank < len(features):
                        logging.warning(f"Singular matrix likely in {market}: rank {rank} < features {len(features)}")

                    init_vals = np.array([-0.5, -0.5, 0.0][:len(features)])
                    for method in ['BFGS', 'Nelder-Mead', 'Powell', 'SLSQP']:
                        try:
                            results = model.fit_mle(
                                init_vals=init_vals,
                                method=method,
                                maxiter=50000,
                                options={'maxfun': 50000, 'tol': 1e-8},
                                compute_std_err=True
                            )
                            if hasattr(results, 'params') and results.params is not None:
                                logging.info(f"{market} MNL succeeded with {method}")
                                logging.info(f"{market} Log-likelihood: {results.log_likelihood}")
                                logging.info(f"{market} Coefficients:\n{results.params}")
                                std_errors = results.std_err if hasattr(results, 'std_err') else [np.nan] * len(results.params)
                                p_values = [2 * (1 - stats.norm.cdf(np.abs(coef / se))) if se != 0 and not np.isnan(se) else np.nan
                                            for coef, se in zip(results.params, std_errors)]
                                self.models[market] = results
                                self.coef_dfs[market] = pd.DataFrame({
                                    'Coefficient': results.params,
                                    'Std_Error': std_errors,
                                    'P_Value': p_values
                                }, index=features)
                                break
                            else:
                                logging.warning(f"MNL with {method} did not produce valid params for {market}")
                        except Exception as e:
                            logging.warning(f"MNL with {method} failed for {market}: {str(e)}")
                    else:
                        raise ValueError(f"All MNL methods failed for {market}")
                except Exception as e:
                    logging.error(f"MNL failed for {market}: {str(e)}")
                    self.models[market] = None
                    self.coef_dfs[market] = pd.DataFrame()
                    
                    # Try simpler MNL with Price and Brand_Cargill
                    logging.warning(f"Trying simpler MNL with Price and Brand_Cargill for {market}")
                    simple_features = ['Price', 'Brand_Cargill']
                    simple_spec = OrderedDict([('Price', 'all_same'), ('Brand_Cargill', 'all_same')])
                    simple_names = OrderedDict([('Price', 'Price'), ('Brand_Cargill', 'Brand_Cargill')])
                    try:
                        simple_model = pl.create_choice_model(
                            data=market_df,
                            alt_id_col='alt_id',
                            obs_id_col='obs_id',
                            choice_col=target,
                            specification=simple_spec,
                            model_type='MNL',
                            names=simple_names
                        )
                        simple_init_vals = np.array([-0.5, 0.0])
                        for method in ['BFGS', 'Nelder-Mead', 'Powell', 'SLSQP']:
                            try:
                                simple_results = simple_model.fit_mle(
                                    init_vals=simple_init_vals,
                                    method=method,
                                    maxiter=50000,
                                    options={'maxfun': 50000, 'tol': 1e-8},
                                    compute_std_err=True
                                )
                                if hasattr(simple_results, 'params') and simple_results.params is not None:
                                    logging.info(f"Simple MNL succeeded with {method} for {market}")
                                    logging.info(f"{market} Simple MNL Coefficients:\n{simple_results.params}")
                                    std_errors = simple_results.std_err if hasattr(simple_results, 'std_err') else [np.nan] * len(simple_results.params)
                                    p_values = [2 * (1 - stats.norm.cdf(np.abs(coef / se))) if se != 0 and not np.isnan(se) else np.nan
                                                for coef, se in zip(simple_results.params, std_errors)]
                                    self.models[market] = simple_results
                                    self.coef_dfs[market] = pd.DataFrame({
                                        'Coefficient': simple_results.params,
                                        'Std_Error': std_errors,
                                        'P_Value': p_values
                                    }, index=simple_features)
                                    break
                                else:
                                    logging.warning(f"Simple MNL with {method} did not produce valid params for {market}")
                            except Exception as e:
                                logging.warning(f"Simple MNL with {method} failed for {market}: {str(e)}")
                        else:
                            raise ValueError(f"All simple MNL methods failed for {market}")
                    except Exception as e:
                        logging.error(f"Simple MNL failed for {market}: {str(e)}")
                        logging.warning(f"Trying binary logit for {market}")
                        binary_df = market_df.copy()
                        binary_df[target] = (binary_df['Brand'] == 'Cargill').astype(int)
                        chosen_counts = binary_df.groupby('obs_id')[target].sum()
                        logging.info(f"Cargill choice variation in {market}: {chosen_counts.value_counts()}")
                        if chosen_counts.eq(0).all() or chosen_counts.eq(len(binary_df)).all():
                            logging.error(f"Binary logit for {market} has no variation in Chosen")
                            continue
                        try:
                            binary_spec = OrderedDict([('Price', 'all_same'), ('CP', 'all_same')])
                            binary_names = OrderedDict([('Price', 'Price'), ('CP', 'CP')])
                            binary_model = pl.create_choice_model(
                                data=binary_df,
                                alt_id_col='alt_id',
                                obs_id_col='obs_id',
                                choice_col=target,
                                specification=binary_spec,
                                model_type='MNL',
                                names=binary_names
                            )
                            binary_init_vals = np.array([-0.5, -0.5])
                            for method in ['BFGS', 'Nelder-Mead', 'Powell', 'SLSQP']:
                                try:
                                    binary_results = binary_model.fit_mle(
                                        init_vals=binary_init_vals,
                                        method=method,
                                        maxiter=50000,
                                        options={'maxfun': 50000, 'tol': 1e-8},
                                        compute_std_err=True
                                    )
                                    if hasattr(binary_results, 'params') and binary_results.params is not None:
                                        logging.info(f"Binary logit succeeded with {method} for {market}")
                                        logging.info(f"{market} Binary Logit Coefficients:\n{binary_results.params}")
                                        std_errors = binary_results.std_err if hasattr(binary_results, 'std_err') else [np.nan] * len(binary_results.params)
                                        p_values = [2 * (1 - stats.norm.cdf(np.abs(coef / se))) if se != 0 and not np.isnan(se) else np.nan
                                                    for coef, se in zip(binary_results.params, std_errors)]
                                        self.models[market] = binary_results
                                        self.coef_dfs[market] = pd.DataFrame({
                                            'Coefficient': binary_results.params,
                                            'Std_Error': std_errors,
                                            'P_Value': p_values
                                        }, index=['Price', 'CP'])
                                        break
                                    else:
                                        logging.warning(f"Binary logit with {method} did not produce valid params for {market}")
                                except Exception as e:
                                    logging.warning(f"Binary logit with {method} failed for {market}: {str(e)}")
                            else:
                                raise ValueError(f"All binary logit methods failed for {market}")
                        except Exception as e:
                            logging.error(f"Binary logit failed for {market}: {str(e)}")
                            logging.warning(f"Trying custom MNL for {market}")
                            try:
                                X = market_df[features].astype(float)
                                y = market_df[target].astype(float)
                                init_vals = np.array([-0.5, -0.5, 0.0])
                                result = minimize(
                                    self.custom_mnl_loglikelihood,
                                    init_vals,
                                    args=(X, y, market_df['obs_id'], market_df['alt_id']),
                                    method='SLSQP',
                                    options={'maxiter': 50000, 'ftol': 1e-10}
                                )
                                logging.info(f"Custom MNL optimization results for {market}: success={result.success}, status={result.status}, message={result.message}")
                                if not result.success:
                                    raise ValueError(f"Custom MNL failed: success={result.success}, status={result.status}, message={result.message}")
                                logging.info(f"{market} Custom MNL Coefficients:\n{result.x}")
                                std_errors = self.compute_std_errors(result.x, X, y, market_df['obs_id'], market_df['alt_id'])
                                p_values = [2 * (1 - stats.norm.cdf(np.abs(coef / se))) if se != 0 and not np.isnan(se) else np.nan
                                            for coef, se in zip(result.x, std_errors)]
                                self.models[market] = result
                                self.coef_dfs[market] = pd.DataFrame({
                                    'Coefficient': result.x,
                                    'Std_Error': std_errors,
                                    'P_Value': p_values
                                }, index=features)
                            except Exception as e:
                                logging.error(f"Custom MNL failed for {market}: {str(e)}")
                                self.models[market] = None
                                self.coef_dfs[market] = pd.DataFrame()

    def add_model_summaries_to_workbook(self, workbook):
        """Write model results to Excel."""
        summary_sheet = workbook.create_sheet(title="Summary", index=0)
        summary_sheet['A1'] = 'Market'
        summary_sheet['B1'] = 'Model Type'
        summary_sheet['C1'] = 'Success'
        for r, market in enumerate(self.markets, start=2):
            model_type = 'Custom MNL' if isinstance(self.models.get(market), scipy.optimize.OptimizeResult) else 'Pylogit MNL'
            coef_df = self.coef_dfs.get(market, pd.DataFrame())
            success = 'Success' if not coef_df.empty and coef_df[['Coefficient', 'Std_Error', 'P_Value']].notna().all().all() else 'Failed' if coef_df.empty else 'Partial (Missing Std Errors or P-Values)'
            summary_sheet[f'A{r}'] = market
            summary_sheet[f'B{r}'] = model_type
            summary_sheet[f'C{r}'] = success
        
        if not workbook.sheetnames or 'Default' not in workbook.sheetnames:
            workbook.create_sheet(title="Default")
        
        for market in self.markets:
            coef_df = self.coef_dfs.get(market, pd.DataFrame())
            logging.info(f"Writing {len(coef_df)} coefficients for {market}")
            
            if f"{market}_coefficients" in workbook.sheetnames:
                coef_sheet = workbook[f"{market}_coefficients"]
                for row in coef_sheet.iter_rows(min_row=1, max_row=coef_sheet.max_row, min_col=1, max_col=coef_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                coef_sheet = workbook.create_sheet(title=f"{market}_coefficients")
            
            if not coef_df.empty:
                coef_sheet['A1'] = 'Parameter'
                coef_sheet['B1'] = 'Coefficient'
                coef_sheet['C1'] = 'Std_Error'
                coef_sheet['D1'] = 'P_Value'
                for r, (index, row) in enumerate(coef_df.iterrows(), start=2):
                    coef_sheet[f'A{r}'] = index
                    coef_sheet[f'B{r}'] = row['Coefficient']
                    coef_sheet[f'C{r}'] = row['Std_Error'] if not np.isnan(row['Std_Error']) else ''
                    coef_sheet[f'D{r}'] = row['P_Value'] if not np.isnan(row['P_Value']) else ''
            
            for col in ['A', 'B', 'C', 'D']:
                column_letter = col
                max_length = 0
                column = coef_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                coef_sheet.column_dimensions[column_letter].width = adjusted_width

            brand_share_market = self.brand_share.get(market, pd.DataFrame(columns=['Brand', 'Actual_Brand_Share']))
            logging.info(f"Writing {len(brand_share_market)} brand share rows for {market}")
            if f"{market}_brand_share" in workbook.sheetnames:
                brand_sheet = workbook[f"{market}_brand_share"]
                for row in brand_sheet.iter_rows(min_row=1, max_row=brand_sheet.max_row, min_col=1, max_col=brand_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                brand_sheet = workbook.create_sheet(title=f"{market}_brand_share")
            
            if not brand_share_market.empty:
                brand_sheet['A1'] = 'Brand'
                brand_sheet['B1'] = 'Actual_Brand_Share'
                for r, (index, row) in enumerate(brand_share_market.iterrows(), start=2):
                    brand_sheet[f'A{r}'] = row['Brand']
                    brand_sheet[f'B{r}'] = row['Actual_Brand_Share']
            
            for col in ['A', 'B']:
                column_letter = col
                max_length = 0
                column = brand_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                brand_sheet.column_dimensions[column_letter].width = adjusted_width

            value_share_market = self.value_share.get(market, pd.DataFrame(columns=['Brand', 'Actual_Value_Share']))
            logging.info(f"Writing {len(value_share_market)} value share rows for {market}")
            if f"{market}_value_share" in workbook.sheetnames:
                value_sheet = workbook[f"{market}_value_share"]
                for row in value_sheet.iter_rows(min_row=1, max_row=value_sheet.max_row, min_col=1, max_col=value_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                value_sheet = workbook.create_sheet(title=f"{market}_value_share")
            
            if not value_share_market.empty:
                value_sheet['A1'] = 'Brand'
                value_sheet['B1'] = 'Actual_Value_Share'
                for r, (index, row) in enumerate(value_share_market.iterrows(), start=2):
                    value_sheet[f'A{r}'] = row['Brand']
                    value_sheet[f'B{r}'] = row['Actual_Value_Share']
            
            for col in ['A', 'B']:
                column_letter = col
                max_length = 0
                column = value_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                value_sheet.column_dimensions[column_letter].width = adjusted_width

        if len(workbook.sheetnames) > 1 and "Default" in workbook.sheetnames:
            del workbook["Default"]

if __name__ == "__main__":
    output_file = 'cargill_findings_v16.xlsx'
    if os.path.exists(output_file):
        os.remove(output_file)
        logging.info(f"Deleted existing {output_file}")
    
    df = pd.read_csv("cattle_feed_data_final.csv")
    workbook = openpyxl.Workbook()
    model = DiscreteChoiceModel(markets=['Maharashtra', 'Gujarat', 'Punjab', 'South'])
    model.fit(df)
    model.add_model_summaries_to_workbook(workbook)
    workbook.save(output_file)
    logging.info(f"Saved output to {output_file}")