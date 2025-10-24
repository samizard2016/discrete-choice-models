import pandas as pd
import numpy as np
from collections.abc import Iterable
import collections
collections.Iterable = Iterable  # Patch for pylogit Python 3.10 compatibility
from collections import OrderedDict
import pylogit as pl
import openpyxl
import os
import json
from openpyxl.utils import get_column_letter

class DiscreteChoiceModel:
    def __init__(self, markets=None, use_pylogit=True, exclude_none=False, output_dir="cattle_feed_output"):
        self.markets = markets if markets else ['Maharashtra', 'Gujarat', 'Punjab', 'South']
        self.models = {}
        self.coef_dfs = {}
        self.use_pylogit = use_pylogit
        self.exclude_none = exclude_none
        self.brand_share = {}
        self.value_share = {}
        self.predicted_brand_share = {}
        self.predicted_value_share = {}
        self.feature_importance = {}
        self.price_sensitivity = {}
        self.output_dir = output_dir
        self.brand_to_id = {}
        self.id_to_brand = {}
        self.choice_id_to_num = {}
    
    def fit(self, df, target='Chosen', features=None):
        # Data validation: Check for duplicate brands in choice sets
        duplicates = df.groupby(['ChoiceID', 'Brand']).size().reset_index(name='count')
        duplicates = duplicates[duplicates['count'] > 1]
        if not duplicates.empty:
            print(f"Found {len(duplicates)} choice sets with duplicate brands:\n{duplicates}")
            duplicate_choice_ids = duplicates['ChoiceID'].unique()
            df = df[~df['ChoiceID'].isin(duplicate_choice_ids)].copy()
            print(f"Removed {len(duplicate_choice_ids)} choice sets with duplicate brands")

        # Check for negative Price or CP
        if (df['Price'] < 0).any() or (df['CP'] < 0).any():
            print("Warning: Negative values found in Price or CP. Ensuring positive values.")
            df['Price'] = df['Price'].clip(lower=0.0)
            df['CP'] = df['CP'].clip(lower=0.0)

        # Remove rows with NaN in Brand or ChoiceID
        initial_len = len(df)
        df = df[df['Brand'].notna() & df['ChoiceID'].notna()].copy()
        if len(df) < initial_len:
            print(f"Removed {initial_len - len(df)} rows with NaN in Brand or ChoiceID")
        
        # Map ChoiceID to numeric obs_id
        unique_choice_ids = df['ChoiceID'].unique()
        self.choice_id_to_num = {cid: idx + 1 for idx, cid in enumerate(unique_choice_ids)}
        df['obs_id_num'] = df['ChoiceID'].map(self.choice_id_to_num)
        print(f"ChoiceID to numeric mapping (first 10): {dict(list(self.choice_id_to_num.items())[:10])}")
        
        # Rescale Price and CP
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce')
        df['CP'] = pd.to_numeric(df['CP'], errors='coerce')
        for col in ['Price', 'CP']:
            if df[col].isna().any():
                print(f"Warning: {col} contains NaN after coercion. Filling with 0.0")
                df[col] = df[col].fillna(0.0)
        # Rescale to positive ranges
        df['Price'] = (df['Price'] - df['Price'].min()) / (df['Price'].max() - df['Price'].min() + 1e-6) * 9.0 + 1.0
        df['CP'] = (df['CP'] - df['CP'].min()) / (df['CP'].max() - df['CP'].min() + 1e-6) * 200
        for col in ['Price', 'CP']:
            print(f"{col} rescaled range: [{df[col].min():.6f}, {df[col].max():.6f}]")
        
        # Debug: Check categorical columns
        cat_cols = ['FC', 'AH', 'VAS', 'Credit']
        for col in cat_cols:
            if col in df.columns:
                print(f"{col} unique values: {df[col].unique()[:20]}")
        
        # One-hot encode categorical variables with fixed categories
        cat_categories = {
            'FC': ['<1%', '1-3%', '>3%'],
            'AH': ['Loss', 'Maintain', 'Gain'],
            'VAS': ['None', 'Basic', 'Through lbl_social media', 'Advanced'],
            'Credit': ['None', '<30d', '>30d']
        }
        for col in cat_cols:
            if col in df.columns:
                df[col] = df[col].fillna('None').astype(str)
                df[col] = pd.Categorical(df[col], categories=cat_categories[col])
                dummy_df = pd.get_dummies(df[col], prefix=col, drop_first=True)
                for dummy_col in dummy_df.columns:
                    dummy_df[dummy_col] = dummy_df[dummy_col].astype(float)
                print(f"Created {len(dummy_df.columns)} dummy variables for {col}: {dummy_df.columns.tolist()}")
                df = pd.concat([df, dummy_df], axis=1)
        
        # Map Brand to numeric alt_id
        unique_brands = df['Brand'].unique()
        self.brand_to_id = {brand: idx + 1 for idx, brand in enumerate(unique_brands)}
        self.id_to_brand = {idx: brand for brand, idx in self.brand_to_id.items()}
        df['alt_id'] = df['Brand'].map(self.brand_to_id)
        print(f"Unique Brand values: {unique_brands}")
        print(f"Brand to ID mapping: {self.brand_to_id}")
        print(f"Unique ChoiceID values (first 10): {df['ChoiceID'].unique()[:10]}")
        print(f"Unique Chosen values: {df['Chosen'].unique()}")
        
        # Validate choice sets
        choice_sets = df.groupby('obs_id_num')
        choice_set_sizes = choice_sets.size()
        print(f"Choice set sizes:\n{choice_set_sizes.value_counts()}")
        invalid_sets = choice_sets.filter(lambda x: x['Chosen'].sum() != 1)
        if not invalid_sets.empty:
            print(f"Warning: {len(invalid_sets['obs_id_num'].unique())} choice sets have invalid Chosen values (sum != 1)")
            print(f"Invalid choice sets sample (first 15):\n{invalid_sets[['obs_id_num', 'Brand', 'Chosen']].head(15)}")
            valid_obs_ids = choice_sets.filter(lambda x: x['Chosen'].sum() == 1)['obs_id_num'].unique()
            print(f"Keeping {len(valid_obs_ids)} valid choice sets")
            df = df[df['obs_id_num'].isin(valid_obs_ids)].copy()
        
        # Debug: Check brand and choice distribution
        for market in self.markets:
            market_df = df[df['Market'] == market]
            print(f"{market} brand distribution:\n{market_df['Brand'].value_counts()}")
            print(f"{market} choice distribution:\n{market_df[market_df['Chosen'] == 1]['Brand'].value_counts()}")
        
        # Feature selection including categorical variables
        features = ['Price', 'CP', 'FC_1-3%', 'FC_>3%', 'AH_Maintain', 'AH_Gain', 
                    'VAS_Basic', 'VAS_Through lbl_social media', 'VAS_Advanced', 
                    'Credit_<30d', 'Credit_>30d']
        print(f"Using features: {features}")
        
        # Calculate Actual Brand Share
        chosen_df = df[df['Chosen'] == 1]
        self.brand_share = chosen_df.groupby(['Market', 'Brand']).size().reset_index(name='Choices')
        self.brand_share['Actual_Brand_Share'] = self.brand_share.groupby('Market')['Choices'].transform(lambda x: x / x.sum() if x.sum() > 0 else 0)
        
        # Calculate Actual Value Share
        self.value_share = chosen_df.groupby(['Market', 'Brand'])['Price'].sum().reset_index(name='Total_Price')
        self.value_share['Actual_Value_Share'] = self.value_share.groupby('Market')['Total_Price'].transform(lambda x: x / x.sum() if x.sum() > 0 else 0)
        
        for market in self.markets:
            market_df = df[df['Market'] == market].copy()
            if market_df.empty:
                print(f"Warning: No data for {market}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                self.predicted_brand_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share'])
                self.predicted_value_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Value_Share'])
                continue
            
            # Verify columns
            missing_cols = [col for col in features if col not in market_df.columns]
            if missing_cols:
                print(f"Warning: Missing columns {missing_cols} for {market}. Adjusting features.")
                features_market = [col for col in features if col in market_df.columns]
            else:
                features_market = features.copy()
            
            try:
                # Pylogit setup
                pylogit_df = market_df.copy()
                pylogit_df['obs_id'] = pylogit_df['obs_id_num']
                pylogit_df['alt_id'] = pylogit_df['Brand'].map(self.brand_to_id)
                pylogit_df['choice'] = pylogit_df[target]
                
                # Debug: Check data types and values
                print(f"{market} pylogit_df dtypes:\n{pylogit_df[features_market + ['obs_id', 'alt_id', 'choice']].dtypes}")
                print(f"{market} unique obs_id values: {pylogit_df['obs_id'].unique()[:10]}")
                print(f"{market} unique alt_id values: {pylogit_df['alt_id'].unique()}")
                for col in features_market:
                    print(f"{market} unique {col} values: {pylogit_df[col].unique()[:10]}")
                
                # Check feature correlation
                corr_matrix = pylogit_df[features_market].corr()
                print(f"{market} feature correlation matrix:\n{corr_matrix}")
                
                # Specification with brand-specific coefficients
                unique_brands = pylogit_df['Brand'].unique()
                spec = OrderedDict([(var, [[brand] for brand in unique_brands]) for var in features_market])
                name_dict = OrderedDict([(var, [f"{var}_{brand}" for brand in unique_brands]) for var in features_market])
                
                # Fit pylogit with stronger regularization
                model = pl.create_choice_model(
                    data=pylogit_df,
                    alt_id_col='alt_id',
                    obs_id_col='obs_id',
                    choice_col='choice',
                    specification=spec,
                    model_type='MNL',
                    names=name_dict
                )
                init_vals = np.array([-1.0 if 'Price' in var else 0.5 for var in name_dict.keys()])
                bounds = [(-100, 100) for _ in name_dict.keys()]
                for method in ['L-BFGS-B', 'SLSQP', 'BFGS']:
                    try:
                        print(f"Trying optimizer {method} for {market}")
                        results = model.fit_mle(
                            init_vals,
                            bounds=bounds,
                            print_res=True,
                            method=method,
                            maxiter=2000,
                            ftol=1e-10,
                            ridge=1e-3
                        )
                        if results is not None:
                            break
                        print(f"Optimizer {method} returned None for {market}")
                    except Exception as e:
                        print(f"Optimizer {method} failed for {market}: {e}")
                if results is None:
                    raise ValueError("Model fitting returned None")
                coef_df = pd.DataFrame({
                    'Coefficient': results.params,
                    'P-value': results.pvalues
                }, index=name_dict.keys())
                self.models[market] = results
                self.coef_dfs[market] = coef_df
                print(f"Pylogit model fitted for {market}: {len(coef_df)} coefficients")
            except Exception as e:
                print(f"Pylogit failed for {market}: {e}")
                import traceback
                print(f"Stack trace for {market}:\n{traceback.format_exc()}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                self.predicted_brand_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share'])
                self.predicted_value_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Value_Share'])
                continue
            
            # Calculate feature importance
            importance = self.get_feature_importance(market)
            self.feature_importance[market] = pd.DataFrame({
                'Feature': importance.index,
                'Importance': importance.values
            })
            
            # Simulate price sensitivity
            self.price_sensitivity[market], self.predicted_brand_share[market], self.predicted_value_share[market] = self.simulate_price_sensitivity(market_df, market, 'Cargill', features_market)
        
        return self
    
    def get_attribute_utilities(self, market):
        coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
        return coef_df['Coefficient'] if 'Coefficient' in coef_df.columns else pd.Series(dtype=float)
    
    def get_feature_importance(self, market):
        utilities = self.get_attribute_utilities(market)
        if utilities.empty:
            return pd.Series(dtype=float)
        importance = utilities.abs() / utilities.abs().sum() if utilities.abs().sum() > 0 else utilities
        return importance
    
    def simulate_price_sensitivity(self, market_df, market, brand='Cargill', features=None):
        if market not in self.models or isinstance(self.models[market], pd.DataFrame):
            print(f"Warning: No valid model for {market}. Skipping price sensitivity analysis.")
            return (pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
        
        coefs = self.get_attribute_utilities(market)
        if coefs.empty:
            print(f"Warning: No coefficients for {market}. Skipping price sensitivity analysis.")
            return (pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
        
        choice_sets = market_df.groupby('obs_id_num')
        choice_set_sizes = choice_sets.size()
        print(f"{market} choice set sizes:\n{choice_set_sizes.value_counts()}")
        print(f"Number of choice sets in {market}: {len(choice_sets)}")
        print(f"Cargill present in {market}: {brand in market_df['Brand'].values}")
        
        # Extract Price coefficients for each brand
        price_coefs = {brand: coefs[f'Price_{brand}'] for brand in market_df['Brand'].unique() if f'Price_{brand}' in coefs.index}
        print(f"Price coefficients for {market}: {price_coefs}")
        
        # Price change range
        price_changes = [round(x, 2) for x in np.arange(-0.10, 0.11, 0.01)]
        results = []
        predicted_brand_share = None
        predicted_value_share = None
        
        baseline_prices = market_df.groupby('Brand')['Price'].mean().to_dict()
        
        for change in price_changes:
            utilities = []
            for choice_id, group in choice_sets:
                if len(group) < 1:
                    print(f"Warning: Empty choice set {choice_id} in {market}. Skipping.")
                    continue
                group = group.copy()
                cargill_mask = group['Brand'] == brand
                if cargill_mask.any():
                    base_price = group.loc[cargill_mask, 'Price'].iloc[0]
                    if base_price <= 0:
                        print(f"Warning: Non-positive base price for Cargill in choice set {choice_id}, {market}. Setting to 1.0.")
                        base_price = 1.0
                    group.loc[cargill_mask, 'Price'] = base_price * (1 + change)
                
                group_utilities = []
                for _, row in group.iterrows():
                    utility = 0
                    for var in features:
                        for brand in market_df['Brand'].unique():
                            coef_name = f"{var}_{row['Brand']}"
                            if coef_name in coefs.index:
                                if var == 'Price':
                                    utility += coefs[coef_name] * row['Price']
                                elif var in row.index:
                                    utility += coefs[coef_name] * row.get(var, 0.0)
                    group_utilities.append(utility)
                
                exp_utilities = np.exp(np.array(group_utilities, dtype=np.float64))
                sum_exp = exp_utilities.sum()
                probs = exp_utilities / sum_exp if sum_exp > 0 else np.ones(len(group_utilities)) / len(group_utilities)
                probs = np.round(probs, 10)
                prob_sum = probs.sum()
                if not np.isclose(prob_sum, 1.0, atol=1e-8):
                    probs = probs / prob_sum if prob_sum > 0 else probs
                
                for i, (_, row) in enumerate(group.iterrows()):
                    utilities.append({
                        'ChoiceID': choice_id,
                        'Brand': row['Brand'],
                        'Probability': probs[i],
                        'Price': row['Price']
                    })
            
            if utilities:
                util_df = pd.DataFrame(utilities)
                choice_set_weights = choice_set_sizes / choice_set_sizes.sum()
                choice_set_counts = util_df.groupby('ChoiceID').size()
                weights = util_df['ChoiceID'].map(choice_set_weights / choice_set_counts).fillna(0)
                
                brand_share = (util_df[util_df['Brand'] == brand]['Probability'] * weights[util_df['Brand'] == brand]).sum() if brand in util_df['Brand'].values else 0.0
                
                value_sum = 0
                brand_value = 0
                for b in util_df['Brand'].unique():
                    brand_mask = util_df['Brand'] == b
                    brand_prob = (util_df[brand_mask]['Probability'] * weights[brand_mask]).sum()
                    brand_price = util_df[brand_mask]['Price'].mean()
                    value_sum += brand_prob * brand_price
                    if b == brand:
                        brand_value = brand_prob * brand_price
                value_share = brand_value / value_sum if value_sum > 0 else 0.0
                
                results.append({
                    'Price_Change_Percent': change * 100,
                    'Brand_Share': brand_share,
                    'Value_Share': value_share
                })
                
                if change == 0:
                    predicted_brand_share = util_df.groupby('Brand').apply(
                        lambda x: (x['Probability'] * weights[x.index]).sum(),
                        include_groups=False
                    ).reset_index(name='Predicted_Brand_Share')
                    predicted_brand_share['Predicted_Brand_Share'] = predicted_brand_share['Predicted_Brand_Share'].round(10)
                    brand_share_sum = predicted_brand_share['Predicted_Brand_Share'].sum()
                    print(f"Sum of Predicted_Brand_Share for {market}: {brand_share_sum:.6f}")
                    if not np.isclose(brand_share_sum, 1.0, atol=1e-8):
                        print(f"Warning: Predicted_Brand_Share sum for {market} is {brand_share_sum:.6f}, adjusting to sum to 1.0")
                        predicted_brand_share['Predicted_Brand_Share'] = predicted_brand_share['Predicted_Brand_Share'] / brand_share_sum if brand_share_sum > 0 else 0.0
                    
                    value_sum_all = 0
                    predicted_value_share = []
                    for b in util_df['Brand'].unique():
                        brand_mask = util_df['Brand'] == b
                        brand_prob = (util_df[brand_mask]['Probability'] * weights[brand_mask]).sum()
                        brand_price = util_df[brand_mask]['Price'].mean()
                        value_sum_all += brand_prob * brand_price
                        predicted_value_share.append({'Brand': b, 'Predicted_Value_Share': brand_prob * brand_price})
                    predicted_value_share = pd.DataFrame(predicted_value_share)
                    predicted_value_share['Predicted_Value_Share'] = predicted_value_share['Predicted_Value_Share'] / value_sum_all if value_sum_all > 0 else 0.0
                    predicted_value_share['Predicted_Value_Share'] = predicted_value_share['Predicted_Value_Share'].round(10)
                    value_share_sum = predicted_value_share['Predicted_Value_Share'].sum()
                    print(f"Sum of Predicted_Value_Share for {market}: {value_share_sum:.6f}")
                    if not np.isclose(value_share_sum, 1.0, atol=1e-8):
                        print(f"Warning: Predicted_Value_Share sum for {market} is {value_share_sum:.6f}, adjusting to sum to 1.0")
                        predicted_value_share['Predicted_Value_Share'] = predicted_value_share['Predicted_Value_Share'] / value_sum_all if value_sum_all > 0 else 0.0
                    
                    cargill_predicted = predicted_brand_share[predicted_brand_share['Brand'] == brand]['Predicted_Brand_Share'].iloc[0] if brand in predicted_brand_share['Brand'].values else 0.0
                    cargill_price_sensitivity = brand_share if brand in util_df['Brand'].values else 0.0
                    if not np.isclose(cargill_predicted, cargill_price_sensitivity, atol=1e-6):
                        print(f"Warning: Cargill Predicted_Brand_Share ({cargill_predicted:.6f}) does not match Brand_Share at 0% ({cargill_price_sensitivity:.6f}) in {market}")
        
        result_df = pd.DataFrame(results)
        if not result_df.empty:
            brand_share_diff = result_df['Brand_Share'].diff().dropna()
            price_coef = price_coefs.get(brand, 0)
            if price_coef < 0 and not all(brand_share_diff <= 0):
                print(f"Warning: Brand_Share in {market} does not consistently decrease with price increase (price coef={price_coef:.6f})")
            elif price_coef > 0 and not all(brand_share_diff >= 0):
                print(f"Warning: Brand_Share in {market} does not consistently increase with price increase (price_coef={price_coef:.6f})")
            value_share_diff = result_df['Value_Share'].diff().dropna()
            if price_coef < 0 and not all(value_share_diff <= 0):
                print(f"Warning: Value_Share in {market} does not consistently decrease with price increase (price coef={price_coef:.6f})")
        
        return (result_df, predicted_brand_share, predicted_value_share)
    
    def add_model_summaries_to_workbook(self, workbook):
        for market in self.markets:
            if f"{market}_summary" in workbook.sheetnames:
                sheet = workbook[f"{market}_summary"]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sheet = workbook.create_sheet(title=f"{market}_summary")
            
            coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
            print(f"Writing {len(coef_df)} coefficients for {market}")
            sheet['A1'] = 'Feature'
            sheet['B1'] = 'Coefficient'
            sheet['C1'] = 'P-value'
            for r, (index, row) in enumerate(coef_df.iterrows(), start=2):
                sheet[f'A{r}'] = index
                sheet[f'B{r}'] = row.get('Coefficient', float('nan'))
                sheet[f'C{r}'] = row.get('P-value', float('nan'))
            
            importance_df = self.feature_importance.get(market, pd.DataFrame(columns=['Feature', 'Importance']))
            if not importance_df.empty:
                start_row = len(coef_df) + 6
                print(f"Writing {len(importance_df)} feature importance rows for {market} at row {start_row}")
                sheet[f'A{start_row}'] = 'Feature Importance'
                sheet[f'A{start_row + 1}'] = 'Feature'
                sheet[f'B{start_row + 1}'] = 'Importance'
                for r, (index, row) in enumerate(importance_df.iterrows(), start=start_row + 2):
                    sheet[f'A{r}'] = row['Feature']
                    sheet[f'B{r}'] = row['Importance']
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            if f"{market}_brand_share" in workbook.sheetnames:
                brand_sheet = workbook[f"{market}_brand_share"]
                for row in brand_sheet.iter_rows(min_row=1, max_row=brand_sheet.max_row, min_col=1, max_col=brand_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                brand_sheet = workbook.create_sheet(title=f"{market}_brand_share")
            brand_share_market = self.brand_share[self.brand_share['Market'] == market][['Brand', 'Actual_Brand_Share']]
            predicted_brand_share = self.predicted_brand_share.get(market, pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']))
            brand_share_market = brand_share_market.merge(predicted_brand_share, on='Brand', how='outer').fillna(0.0).infer_objects(copy=False)
            print(f"Writing {len(brand_share_market)} brand share rows for {market}")
            brand_sheet['A1'] = 'Brand'
            brand_sheet['B1'] = 'Actual_Brand_Share'
            brand_sheet['C1'] = 'Predicted_Brand_Share'
            for r, (index, row) in enumerate(brand_share_market.iterrows(), start=2):
                brand_sheet[f'A{r}'] = row['Brand']
                brand_sheet[f'B{r}'] = row['Actual_Brand_Share']
                brand_sheet[f'C{r}'] = row['Predicted_Brand_Share']
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = brand_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                brand_sheet.column_dimensions[column_letter].width = adjusted_width
            
            if f"{market}_value_share" in workbook.sheetnames:
                value_sheet = workbook[f"{market}_value_share"]
                for row in value_sheet.iter_rows(min_row=1, max_row=value_sheet.max_row, min_col=1, max_col=value_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                value_sheet = workbook.create_sheet(title=f"{market}_value_share")
            value_share_market = self.value_share[self.value_share['Market'] == market][['Brand', 'Actual_Value_Share']]
            predicted_value_share = self.predicted_value_share.get(market, pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
            value_share_market = value_share_market.merge(predicted_value_share, on='Brand', how='outer').fillna(0.0).infer_objects(copy=False)
            print(f"Writing {len(value_share_market)} value share rows for {market}")
            value_sheet['A1'] = 'Brand'
            value_sheet['B1'] = 'Actual_Value_Share'
            value_sheet['C1'] = 'Predicted_Value_Share'
            for r, (index, row) in enumerate(value_share_market.iterrows(), start=2):
                value_sheet[f'A{r}'] = row['Brand']
                value_sheet[f'B{r}'] = row['Actual_Value_Share']
                value_sheet[f'C{r}'] = row['Predicted_Value_Share']
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = value_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                value_sheet.column_dimensions[column_letter].width = adjusted_width
            
            if f"{market}_price_sensitivity" in workbook.sheetnames:
                sensitivity_sheet = workbook[f"{market}_price_sensitivity"]
                for row in sensitivity_sheet.iter_rows(min_row=1, max_row=sensitivity_sheet.max_row, min_col=1, max_col=sensitivity_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sensitivity_sheet = workbook.create_sheet(title=f"{market}_price_sensitivity")
            sensitivity_df = self.price_sensitivity.get(market, pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']))
            print(f"Writing {len(sensitivity_df)} price sensitivity rows for {market}")
            sensitivity_sheet['A1'] = 'Price_Change_Percent'
            sensitivity_sheet['B1'] = 'Brand_Share'
            sensitivity_sheet['C1'] = 'Value_Share'
            for r, (index, row) in enumerate(sensitivity_df.iterrows(), start=2):
                sensitivity_sheet[f'A{r}'] = row['Price_Change_Percent']
                sensitivity_sheet[f'B{r}'] = row['Brand_Share']
                sensitivity_sheet[f'C{r}'] = row['Value_Share']
            sensitivity_sheet['A{0}'.format(len(sensitivity_df) + 4)] = 'Note: Line chart can be created in Excel using the data above.'
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = sensitivity_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sensitivity_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def generate_price_sensitivity_charts(self):
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            print(f"Output directory {self.output_dir} created or exists")
        except Exception as e:
            print(f"Error creating output directory {self.output_dir}: {e}")
            return
        
        for market in self.markets:
            sensitivity_df = self.price_sensitivity.get(market, pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']))
            if sensitivity_df.empty:
                print(f"Warning: No price sensitivity data for {market}. Skipping chart generation.")
                continue
            
            try:
                chart_data = sensitivity_df[['Price_Change_Percent', 'Brand_Share', 'Value_Share']].to_dict(orient='records')
                chart_data_json = json.dumps(chart_data)
            except Exception as e:
                print(f"Error preparing chart data for {market}: {e}")
                continue
            
            html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>{market} Price Sensitivity Chart</title>
    <script src="https://cdn.jsdelivr.net/npm/react@18.2.0/umd/react.production.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/react-dom@18.2.0/umd/react-dom.production.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/recharts@2.12.7/dist/recharts.umd.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/@babel/standalone@7.24.7/babel.min.js"></script>
    <style>
        body {{ font-family: Arial, sans-serif; display: flex; justify-content: center; align-items: center; height: 100vh; margin: 0; background-color: #f0f0f0; }}
        #chart {{ width: 800px; height: 500px; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }}
    </style>
</head>
<body>
    <div id="chart"></div>
    <script type="text/babel">
        const {{ LineChart, Line, XAxis, YAxis, CartesianGrid, Tooltip, Legend }} = Recharts;
        const data = {chart_data_json};

        const App = () => (
            <LineChart width={{800}} height={{500}} data={{data}} margin={{ top: 20, right: 30, left: 20, bottom: 10 }}>
                <CartesianGrid strokeDasharray="3 3" />
                <XAxis
                    dataKey="Price_Change_Percent"
                    label={{ value: 'Price Change (%)', position: 'insideBottom', offset: -5 }}
                    domain={{[-10, 10]}}
                    ticks={{[-10, -8, -6, -4, -2, 0, 2, 4, 6, 8, 10]}}
                />
                <YAxis
                    label={{ value: 'Share', angle: -90, position: 'insideLeft' }}
                    domain={{[0, 'auto']}}
                />
                <Tooltip formatter={{(value) => (value * 100).toFixed(2) + '%'}} />
                <Legend verticalAlign="top" height={{36}} />
                <Line type="monotone" dataKey="Brand_Share" name="Predicted Brand Share" stroke="#8884d8" strokeWidth={{2}} />
                <Line type="monotone" dataKey="Value_Share" name="Predicted Value Share" stroke="#82ca9d" strokeWidth={{2}} />
            </LineChart>
        );

        ReactDOM.render(<App />, document.getElementById('chart'));
    </script>
</body>
</html>
"""
            chart_path = os.path.join(self.output_dir, f"{market}_price_sensitivity.html")
            try:
                with open(chart_path, 'w') as f:
                    f.write(html_content)
                print(f"Saved price sensitivity chart for {market} to {chart_path}")
            except Exception as e:
                print(f"Error saving chart for {market} to {chart_path}: {e}")

if __name__ == "__main__":
    df = pd.read_csv("combined_choice_data_v2.csv")
    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    model = DiscreteChoiceModel(markets=['Maharashtra', 'Gujarat', 'Punjab', 'South'], use_pylogit=True)
    model.fit(df)
    model.add_model_summaries_to_workbook(workbook)
    model.generate_price_sensitivity_charts()
    workbook.save('cargill_findings_v21.xlsx')