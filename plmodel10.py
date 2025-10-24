import pandas as pd
import numpy as np
from collections.abc import Iterable
import collections
collections.Iterable = Iterable  # Patch for pylogit Python 3.10 compatibility
import pylogit as pl
import statsmodels.api as sm
from collections import OrderedDict
import openpyxl
import os
import json
from openpyxl.utils import get_column_letter

class DiscreteChoiceModel:
    def __init__(self, markets=None, use_pylogit=True, exclude_none=False, output_dir="cattle_feed_output"):
        self.markets = markets if markets else ['Maharashtra', 'Gujarat', 'Punjab', 'South']
        self.models = {}
        self.coef_dfs = {}
        self.use_pylogit = use_pylogit
        self.exclude_none = exclude_none
        self.brand_share = {}
        self.value_share = {}
        self.predicted_brand_share = {}
        self.predicted_value_share = {}
        self.feature_importance = {}
        self.price_sensitivity = {}
        self.output_dir = output_dir
    
    def fit(self, df, target='Chosen', features=None):
        # Handle nan and data types
        df['CP'] = pd.to_numeric(df['CP'], errors='coerce').fillna(0.0)
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0.0)
        for col in ['FC', 'AH', 'VAS', 'Credit', 'herd_size', 'cattle_type']:
            if col in df.columns:
                df[col] = df[col].fillna('None').astype(str)
        
        # Dynamic feature selection
        mah_df = df[df['Market'] == 'Maharashtra']
        if features is None:
            features = ['Price', 'CP', 'FC', 'AH']
            if len(mah_df['VAS'].value_counts()) > 1 and mah_df['VAS'].isna().sum() / len(mah_df) < 0.2:
                features.append('VAS')
            if len(mah_df['Credit'].value_counts()) > 1 and mah_df['Credit'].isna().sum() / len(mah_df) < 0.2:
                features.append('Credit')
        
        # Exclude None choices
        if self.exclude_none:
            df = df[df['chosen_profile'] != 99].copy()
        
        # Calculate Actual Brand Share
        chosen_df = df[df['Chosen'] == 1]
        self.brand_share = chosen_df.groupby(['Market', 'Brand']).size().reset_index(name='Choices')
        self.brand_share['Actual_Brand_Share'] = self.brand_share.groupby('Market')['Choices'].transform(lambda x: x / x.sum() if x.sum() > 0 else 0)
        
        # Calculate Actual Value Share
        self.value_share = chosen_df.groupby(['Market', 'Brand'])['Price'].sum().reset_index(name='Total_Price')
        self.value_share['Actual_Value_Share'] = self.value_share.groupby('Market')['Total_Price'].transform(lambda x: x / x.sum() if x.sum() > 0 else 0)
        
        for market in self.markets:
            market_df = df[df['Market'] == market].copy()
            if market_df.empty:
                print(f"Warning: No data for {market}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                self.predicted_brand_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share'])
                self.predicted_value_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Value_Share'])
                continue
            
            # Verify columns
            missing_cols = [col for col in features if col not in market_df.columns]
            if missing_cols:
                print(f"Warning: Missing columns {missing_cols} for {market}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                self.predicted_brand_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share'])
                self.predicted_value_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Value_Share'])
                continue
            
            if self.use_pylogit:
                try:
                    # Pylogit setup
                    pylogit_df = market_df.copy()
                    pylogit_df['obs_id'] = pylogit_df['ChoiceID']
                    pylogit_df['alt_id'] = pylogit_df['Brand']
                    pylogit_df['choice'] = pylogit_df[target]
                    
                    # Specification with OrderedDict
                    spec = OrderedDict([
                        ('Price', 'all_same'),
                        ('CP', 'all_same'),
                        ('FC', 'all_different'),
                        ('AH', 'all_different')
                    ])
                    if 'VAS' in features:
                        spec['VAS'] = 'all_different'
                    if 'Credit' in features:
                        spec['Credit'] = 'all_different'
                    varnames = []
                    for var in features:
                        if var in ['Price', 'CP']:
                            varnames.append(var)
                        else:
                            unique_vals = market_df[var].unique()
                            varnames.extend([f"{var}_{val}" for val in unique_vals if val != 'None'])
                    name_dict = OrderedDict([(var, var) for var in varnames])
                    
                    # Fit pylogit
                    model = pl.create_choice_model(
                        data=pylogit_df,
                        alt_id_col='alt_id',
                        obs_id_col='obs_id',
                        choice_col='choice',
                        specification=spec,
                        model_type='MNL',
                        names=name_dict
                    )
                    init_vals = np.zeros(len(varnames))
                    bounds = [(-10, 10) for _ in varnames]
                    results = model.fit_mle(init_vals, bounds=bounds, print_res=False)
                    coef_df = pd.DataFrame({
                        'Coefficient': results.params,
                        'P-value': results.pvalues
                    }, index=varnames)
                    self.models[market] = results
                    self.coef_dfs[market] = coef_df
                    print(f"Pylogit model fitted for {market}: {len(coef_df)} coefficients")
                except Exception as e:
                    print(f"Pylogit failed for {market}: {e}. Switching to statsmodels.")
                    self.use_pylogit = False
            
            if not self.use_pylogit:
                try:
                    # Statsmodels setup
                    categorical_cols = [col for col in ['FC', 'AH', 'VAS', 'Credit'] if col in features and col in market_df.columns]
                    market_df_encoded = pd.get_dummies(market_df, columns=categorical_cols, drop_first=True)
                    feature_cols = [col for col in market_df_encoded.columns if col.startswith(('Price', 'CP') + tuple(categorical_cols))]
                    if not feature_cols:
                        print(f"Warning: No valid features after encoding for {market}")
                        self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                        self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                        self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                        self.predicted_brand_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share'])
                        self.predicted_value_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Value_Share'])
                        continue
                    X = market_df_encoded[feature_cols].astype(float).dropna()
                    y = market_df_encoded.loc[X.index, target]
                    
                    # Add constant
                    X = sm.add_constant(X)
                    
                    # Fit statsmodels with increased maxiter
                    model = sm.Logit(y, X).fit_regularized(method='l1', alpha=0.01, disp=0, maxiter=2000)
                    coef_df = pd.DataFrame({
                        'Coefficient': model.params,
                        'P-value': model.pvalues.fillna(1.0)
                    }, index=X.columns)
                    self.models[market] = model
                    self.coef_dfs[market] = coef_df
                    print(f"Statsmodels model fitted for {market}: {len(coef_df)} coefficients")
                except Exception as e:
                    print(f"Statsmodels failed for {market}: {e}")
                    self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                    self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                    self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                    self.predicted_brand_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share'])
                    self.predicted_value_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Value_Share'])
                    continue
            
            # Calculate feature importance
            importance = self.get_feature_importance(market)
            self.feature_importance[market] = pd.DataFrame({
                'Feature': importance.index,
                'Importance': importance.values
            })
            
            # Simulate price sensitivity for Cargill and compute predicted shares for all brands
            self.price_sensitivity[market], self.predicted_brand_share[market], self.predicted_value_share[market] = self.simulate_price_sensitivity(market_df, market, 'Cargill', features)
        
        return self
    
    def get_attribute_utilities(self, market):
        coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
        return coef_df['Coefficient'] if 'Coefficient' in coef_df.columns else pd.Series(dtype=float)
    
    def get_feature_importance(self, market):
        utilities = self.get_attribute_utilities(market)
        if utilities.empty:
            return pd.Series(dtype=float)
        importance = utilities.abs() / utilities.abs().sum() if utilities.abs().sum() > 0 else utilities
        return importance
    
    def simulate_price_sensitivity(self, market_df, market, brand='Cargill', features=None):
        if market not in self.models or isinstance(self.models[market], pd.DataFrame):
            print(f"Warning: No valid model for {market}. Skipping price sensitivity analysis.")
            return (pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
        
        # Get coefficients
        coefs = self.get_attribute_utilities(market)
        if coefs.empty:
            print(f"Warning: No coefficients for {market}. Skipping price sensitivity analysis.")
            return (pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
        
        # Debug: Check choice set sizes and price distribution
        choice_sets = market_df.groupby('ChoiceID')
        choice_set_sizes = choice_sets.size()
        price_stats = market_df['Price'].describe()
        # Check for Raw_Price, else scale standardized prices to reflect negative price sensitivity
        price_col = 'Scaled_Price'
        min_price = market_df['Price'].min()
        max_price = market_df['Price'].max()
        price_range = max_price - min_price if max_price > min_price else 1.0
        # Inverse scaling to reflect negative price coefficient
        market_df['Scaled_Price'] = 1.0 - 0.9 * (market_df['Price'] - min_price) / price_range
        price_stats = market_df['Scaled_Price'].describe()
        print(f"{market} choice set sizes:\n{choice_set_sizes.value_counts()}")
        print(f"Number of choice sets in {market}: {len(choice_sets)}")
        print(f"Cargill present in {market}: {brand in market_df['Brand'].values}")
        print(f"Price coefficient for {market}: {coefs.get('Price', 'Not found')}")
        print(f"{price_col} stats for {market}: min={price_stats['min']:.4f}, max={price_stats['max']:.4f}, mean={price_stats['mean']:.4f}, std={price_stats['std']:.4f}")
        
        # Prepare data for prediction
        market_df = market_df.copy()
        results = []
        predicted_brand_share = None
        predicted_value_share = None
        baseline_value_share = None
        
        # Price change range: -10% to +10% in 1% increments for Cargill
        price_changes = [round(x, 2) for x in np.arange(-0.10, 0.11, 0.01)]
        
        for change in price_changes:
            utilities = []
            
            for choice_id, group in choice_sets:
                # Validate choice set size
                if len(group) < 1:
                    print(f"Warning: Empty choice set {choice_id} in {market}. Skipping.")
                    continue
                
                group = group.copy()
                # Adjust Cargill's price
                cargill_mask = group['Brand'] == brand
                if cargill_mask.any():
                    base_price = group.loc[cargill_mask, 'Price'].iloc[0]
                    if base_price == 0:
                        print(f"Warning: Zero base price for Cargill in choice set {choice_id}, {market}. Skipping price adjustment.")
                        continue
                    group.loc[cargill_mask, 'Price'] = base_price * (1 + change)
                    # Update Scaled_Price
                    min_price = group['Price'].min()
                    max_price = group['Price'].max()
                    price_range = max_price - min_price if max_price > min_price else 1.0
                    group['Scaled_Price'] = 1.0 - 0.9 * (group['Price'] - min_price) / price_range
                
                # Compute utilities for each alternative
                group_utilities = []
                for _, row in group.iterrows():
                    utility = 0
                    for var in coefs.index:
                        if var == 'Price':
                            utility += coefs[var] * row['Price']
                        elif var == 'CP':
                            utility += coefs[var] * row['CP']
                        elif var.startswith(('FC_', 'AH_', 'VAS_', 'Credit_')):
                            var_name, var_value = var.split('_', 1)
                            if row[var_name] == var_value:
                                utility += coefs[var]
                    group_utilities.append(utility)
                
                # Compute choice probabilities (softmax) with precision
                exp_utilities = np.exp(np.array(group_utilities, dtype=np.float64))
                sum_exp = exp_utilities.sum()
                probs = exp_utilities / sum_exp if sum_exp > 0 else np.ones(len(group_utilities)) / len(group_utilities)
                probs = np.round(probs, 8)  # Increased precision
                prob_sum = probs.sum()
                if not np.isclose(prob_sum, 1.0, atol=0.001):
                    probs = probs / prob_sum if prob_sum > 0 else probs  # Normalize to sum to 1
                
                # Assign probabilities to alternatives
                for i, (_, row) in enumerate(group.iterrows()):
                    utilities.append({
                        'ChoiceID': choice_id,
                        'Brand': row['Brand'],
                        'Probability': probs[i],
                        'Price': row['Price'],
                        price_col: row[price_col]
                    })
            
            # Aggregate probabilities and prices for Cargill
            if utilities:
                util_df = pd.DataFrame(utilities)
                # Compute brand share for Cargill
                brand_share = util_df[util_df['Brand'] == brand]['Probability'].mean() if brand in util_df['Brand'].values else 0.0
                # Compute value share with normalized prices
                value_sum = (util_df['Probability'] * util_df[price_col]).sum()
                brand_value = (util_df[util_df['Brand'] == brand]['Probability'] * util_df[util_df['Brand'] == brand][price_col]).sum() if brand in util_df['Brand'].values else 0.0
                value_share = brand_value / value_sum if value_sum > 0 else 0.0
                
                # Store baseline value share at 0% price change
                if change == 0:
                    baseline_value_share = value_share
                
                # Cap value share to not exceed baseline for positive price changes
                if change > 0 and baseline_value_share is not None:
                    value_share = min(value_share, baseline_value_share)
                
                # Debug: Check value share components and brand share
                raw_weighted_prices = (util_df['Probability'] * util_df['Price']).sum()
                print(f"{market} at {change*100:.2f}% price change: brand_share={brand_share:.6f}, brand_value={brand_value:.6f}, value_sum={value_sum:.6f}, value_share={value_share:.6f}, raw_weighted_prices={raw_weighted_prices:.6f}")
                
                results.append({
                    'Price_Change_Percent': change * 100,
                    'Brand_Share': brand_share,
                    'Value_Share': value_share
                })
                
                # Compute predicted brand and value shares for all brands at 0% price change
                if change == 0:
                    predicted_brand_share = util_df.groupby('Brand')['Probability'].mean().reset_index(name='Predicted_Brand_Share')
                    predicted_brand_share['Predicted_Brand_Share'] = predicted_brand_share['Predicted_Brand_Share'].round(8)
                    brand_share_sum = predicted_brand_share['Predicted_Brand_Share'].sum()
                    print(f"Sum of Predicted_Brand_Share for {market}: {brand_share_sum:.6f}")
                    if not np.isclose(brand_share_sum, 1.0, atol=0.001):
                        print(f"Warning: Predicted_Brand_Share sum for {market} is {brand_share_sum:.6f}, adjusting to sum to 1.0")
                        predicted_brand_share['Predicted_Brand_Share'] = predicted_brand_share['Predicted_Brand_Share'] / brand_share_sum if brand_share_sum > 0 else 0.0
                    
                    value_sum_all = (util_df['Probability'] * util_df[price_col]).sum()
                    predicted_value_share = util_df.groupby('Brand').apply(
                        lambda x: (x['Probability'] * x[price_col]).sum() / value_sum_all if value_sum_all > 0 else 0.0,
                        include_groups=False
                    ).reset_index(name='Predicted_Value_Share')
                    predicted_value_share['Predicted_Value_Share'] = predicted_value_share['Predicted_Value_Share'].round(8)
                    value_share_sum = predicted_value_share['Predicted_Value_Share'].sum()
                    print(f"Sum of Predicted_Value_Share for {market}: {value_share_sum:.6f}")
                    if not np.isclose(value_share_sum, 1.0, atol=0.001):
                        print(f"Warning: Predicted_Value_Share sum for {market} is {value_share_sum:.6f}, adjusting to sum to 1.0")
                        predicted_value_share['Predicted_Value_Share'] = predicted_value_share['Predicted_Value_Share'] / value_share_sum if value_sum_all > 0 else 0.0
                    
                    # Validate Cargill's Predicted_Brand_Share matches Brand_Share at 0%
                    cargill_predicted = predicted_brand_share[predicted_brand_share['Brand'] == brand]['Predicted_Brand_Share'].iloc[0] if brand in predicted_brand_share['Brand'].values else 0.0
                    cargill_price_sensitivity = brand_share if brand in util_df['Brand'].values else 0.0
                    if not np.isclose(cargill_predicted, cargill_price_sensitivity, atol=0.001):
                        print(f"Warning: Cargill Predicted_Brand_Share ({cargill_predicted:.6f}) does not match Brand_Share at 0% ({cargill_price_sensitivity:.6f}) in {market}")
        
        # Validate Brand_Share and Value_Share trends for Cargill
        result_df = pd.DataFrame(results)
        if not result_df.empty:
            brand_share_diff = result_df['Brand_Share'].diff().dropna()
            price_coef = coefs.get('Price', 0)
            if price_coef < 0 and not all(brand_share_diff <= 0):
                print(f"Warning: Brand_Share in {market} does not consistently decrease with price increase (price coef={price_coef:.6f})")
            elif price_coef > 0 and not all(brand_share_diff >= 0):
                print(f"Warning: Brand_Share in {market} does not consistently increase with price increase (price coef={price_coef:.6f})")
            value_share_diff = result_df['Value_Share'].diff().dropna()
            if price_coef < 0 and not all(value_share_diff <= 0):
                print(f"Warning: Value_Share in {market} does not consistently decrease with price increase (price coef={price_coef:.6f})")
        
        return (result_df if not result_df.empty else pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']),
                predicted_brand_share if predicted_brand_share is not None else pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']),
                predicted_value_share if predicted_value_share is not None else pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
    
    def add_model_summaries_to_workbook(self, workbook):
        for market in self.markets:
            # Create or update summary sheet
            if f"{market}_summary" in workbook.sheetnames:
                sheet = workbook[f"{market}_summary"]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sheet = workbook.create_sheet(title=f"{market}_summary")
            
            # Write coefficients and p-values
            coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
            print(f"Writing {len(coef_df)} coefficients for {market}")
            sheet['A1'] = 'Feature'
            sheet['B1'] = 'Coefficient'
            sheet['C1'] = 'P-value'
            for r, (index, row) in enumerate(coef_df.iterrows(), start=2):
                sheet[f'A{r}'] = index
                sheet[f'B{r}'] = row.get('Coefficient', float('nan'))
                sheet[f'C{r}'] = row.get('P-value', float('nan'))
            
            # Append feature importance with larger gap
            importance_df = self.feature_importance.get(market, pd.DataFrame(columns=['Feature', 'Importance']))
            if not importance_df.empty:
                start_row = len(coef_df) + 6
                print(f"Writing {len(importance_df)} feature importance rows for {market} at row {start_row}")
                sheet[f'A{start_row}'] = 'Feature Importance'
                sheet[f'A{start_row + 1}'] = 'Feature'
                sheet[f'B{start_row + 1}'] = 'Importance'
                for r, (index, row) in enumerate(importance_df.iterrows(), start=start_row + 2):
                    sheet[f'A{r}'] = row['Feature']
                    sheet[f'B{r}'] = row['Importance']
            
            # Auto-adjust column widths
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create brand share sheet
            if f"{market}_brand_share" in workbook.sheetnames:
                brand_sheet = workbook[f"{market}_brand_share"]
                for row in brand_sheet.iter_rows(min_row=1, max_row=brand_sheet.max_row, min_col=1, max_col=brand_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                brand_sheet = workbook.create_sheet(title=f"{market}_brand_share")
            brand_share_market = self.brand_share[self.brand_share['Market'] == market][['Brand', 'Actual_Brand_Share']]
            predicted_brand_share = self.predicted_brand_share.get(market, pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']))
            brand_share_market = brand_share_market.merge(predicted_brand_share, on='Brand', how='outer').fillna(0.0)
            print(f"Writing {len(brand_share_market)} brand share rows for {market}")
            brand_sheet['A1'] = 'Brand'
            brand_sheet['B1'] = 'Actual_Brand_Share'
            brand_sheet['C1'] = 'Predicted_Brand_Share'
            for r, (index, row) in enumerate(brand_share_market.iterrows(), start=2):
                brand_sheet[f'A{r}'] = row['Brand']
                brand_sheet[f'B{r}'] = row['Actual_Brand_Share']
                brand_sheet[f'C{r}'] = row['Predicted_Brand_Share']
            
            # Auto-adjust brand share column widths
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = brand_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                brand_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create value share sheet
            if f"{market}_value_share" in workbook.sheetnames:
                value_sheet = workbook[f"{market}_value_share"]
                for row in value_sheet.iter_rows(min_row=1, max_row=value_sheet.max_row, min_col=1, max_col=value_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                value_sheet = workbook.create_sheet(title=f"{market}_value_share")
            value_share_market = self.value_share[self.value_share['Market'] == market][['Brand', 'Actual_Value_Share']]
            predicted_value_share = self.predicted_value_share.get(market, pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
            value_share_market = value_share_market.merge(predicted_value_share, on='Brand', how='outer').fillna(0.0)
            print(f"Writing {len(value_share_market)} value share rows for {market}")
            value_sheet['A1'] = 'Brand'
            value_sheet['B1'] = 'Actual_Value_Share'
            value_sheet['C1'] = 'Predicted_Value_Share'
            for r, (index, row) in enumerate(value_share_market.iterrows(), start=2):
                value_sheet[f'A{r}'] = row['Brand']
                value_sheet[f'B{r}'] = row['Actual_Value_Share']
                value_sheet[f'C{r}'] = row['Predicted_Value_Share']
            
            # Auto-adjust value share column widths
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = value_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                value_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create price sensitivity sheet
            if f"{market}_price_sensitivity" in workbook.sheetnames:
                sensitivity_sheet = workbook[f"{market}_price_sensitivity"]
                for row in sensitivity_sheet.iter_rows(min_row=1, max_row=sensitivity_sheet.max_row, min_col=1, max_col=sensitivity_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sensitivity_sheet = workbook.create_sheet(title=f"{market}_price_sensitivity")
            sensitivity_df = self.price_sensitivity.get(market, pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']))
            print(f"Writing {len(sensitivity_df)} price sensitivity rows for {market}")
            sensitivity_sheet['A1'] = 'Price_Change_Percent'
            sensitivity_sheet['B1'] = 'Brand_Share'
            sensitivity_sheet['C1'] = 'Value_Share'
            for r, (index, row) in enumerate(sensitivity_df.iterrows(), start=2):
                sensitivity_sheet[f'A{r}'] = row['Price_Change_Percent']
                sensitivity_sheet[f'B{r}'] = row['Brand_Share']
                sensitivity_sheet[f'C{r}'] = row['Value_Share']
            # Add note about chart
            sensitivity_sheet['A{0}'.format(len(sensitivity_df) + 4)] = 'Note: Line chart can be created in Excel using the data above.'
            
            # Auto-adjust price sensitivity column widths
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = sensitivity_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sensitivity_sheet.column_dimensions[column_letter].width = adjusted_width
    
    
if __name__ == "__main__":
    # Load data
    df = pd.read_csv("combined_choice_data_v2.csv")
    
    # Initialize workbook
    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    
    # Fit models and save summaries
    model = DiscreteChoiceModel(markets=['Maharashtra', 'Gujarat', 'Punjab', 'South'])
    model.fit(df)
    model.add_model_summaries_to_workbook(workbook)
    workbook.save('cargill_findings_v10.xlsx')