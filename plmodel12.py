import pandas as pd
import numpy as np
from collections.abc import Iterable
import collections
collections.Iterable = Iterable
from collections import OrderedDict
import pylogit as pl
import openpyxl
import os
import json

class DiscreteChoiceModel:
    def __init__(self, markets=None, use_pylogit=True, output_dir="cattle_feed_output"):
        self.markets = markets if markets else ['Maharashtra', 'Gujarat', 'Punjab', 'South']
        self.models = {}
        self.coef_dfs = {}
        self.use_pylogit = use_pylogit
        self.brand_share = {}
        self.value_share = {}
        self.predicted_brand_share = {}
        self.predicted_value_share = {}
        self.feature_importance = {}
        self.price_sensitivity = {}
        self.output_dir = output_dir
        self.brand_to_id = {}
        self.id_to_brand = {}
        self.choice_id_to_num = {}
    
    def fit(self, df, target='Chosen', features=None):
        # Remove duplicates
        duplicates = df.groupby(['ChoiceID', 'Brand']).size().reset_index(name='count')
        duplicates = duplicates[duplicates['count'] > 1]
        if not duplicates.empty:
            print(f"Found {len(duplicates)} choice sets with duplicate brands:\n{duplicates}")
            df = df[~df['ChoiceID'].isin(duplicates['ChoiceID'].unique())].copy()
            print(f"Removed {len(duplicates)} choice sets with duplicate brands")

        # Shift negative Price/CP
        if (df['Price'] < 0).any() or (df['CP'] < 0).any():
            print("Warning: Negative values found in Price or CP. Shifting to realistic ranges.")
            df['Price'] = df['Price'] - df['Price'].min() + 100.0
            df['CP'] = df['CP'] - df['CP'].min() + 10.0
            print(f"Adjusted Price range: [{df['Price'].min():.6f}, {df['Price'].max():.6f}]")
            print(f"Adjusted CP range: [{df['CP'].min():.6f}, {df['CP'].max():.6f}]")

        # Remove NaN
        initial_len = len(df)
        df = df[df['Brand'].notna() & df['ChoiceID'].notna()].copy()
        if len(df) < initial_len:
            print(f"Removed {initial_len - len(df)} rows with NaN in Brand or ChoiceID")
        
        # Map ChoiceID to obs_id
        unique_choice_ids = df['ChoiceID'].unique()
        self.choice_id_to_num = {cid: idx + 1 for idx, cid in enumerate(unique_choice_ids)}
        df['obs_id_num'] = df['ChoiceID'].map(self.choice_id_to_num)
        print(f"ChoiceID to numeric mapping (first 10): {dict(list(self.choice_id_to_num.items())[:10])}")
        
        # Add synthetic variation to Price/CP
        for col in ['Price', 'CP']:
            unique_vals = df[col].unique()
            if len(unique_vals) <= 3:
                print(f"Warning: {col} has only {len(unique_vals)} unique values. Adding synthetic variation.")
                noise = np.random.normal(0, 0.01 * (df[col].max() - df[col].min()), len(df))
                df[col] = df[col] + noise
                df[col] = df[col].clip(df[col].min(), df[col].max())
        
        # Rescale Price/CP
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(df['Price'].mean())
        df['CP'] = pd.to_numeric(df['CP'], errors='coerce').fillna(df['CP'].mean())
        df['Price'] = (df['Price'] - df['Price'].min()) / (df['Price'].max() - df['Price'].min() + 1e-6) * 900 + 100
        df['CP'] = (df['CP'] - df['CP'].min()) / (df['CP'].max() - df['CP'].min() + 1e-6) * 20 + 10
        print(f"Price rescaled range: [{df['Price'].min():.6f}, {df['Price'].max():.6f}]")
        print(f"CP rescaled range: [{df['CP'].min():.6f}, {df['CP'].max():.6f}]")
        
        # Map Brand to alt_id
        unique_brands = df['Brand'].unique()
        self.brand_to_id = {brand: idx + 1 for idx, brand in enumerate(unique_brands)}
        self.id_to_brand = {idx: brand for brand, idx in self.brand_to_id.items()}
        df['alt_id'] = df['Brand'].map(self.brand_to_id)
        print(f"Unique Brand values: {unique_brands}")
        print(f"Brand to ID mapping: {self.brand_to_id}")
        print(f"Unique ChoiceID values (first 10): {df['ChoiceID'].unique()[:10]}")
        print(f"Unique Chosen values: {df['Chosen'].unique()}")
        
        # Exclude invalid choice sets
        choice_sets = df.groupby('obs_id_num')
        choice_set_sizes = choice_sets.size()
        print(f"Choice set sizes:\n{choice_set_sizes.value_counts()}")
        invalid_sets = choice_sets.filter(lambda x: x['Chosen'].sum() != 1)
        if not invalid_sets.empty:
            print(f"Warning: {len(invalid_sets['obs_id_num'].unique())} choice sets have invalid Chosen values")
            valid_obs_ids = choice_sets.filter(lambda x: x['Chosen'].sum() == 1)['obs_id_num'].unique()
            print(f"Keeping {len(valid_obs_ids)} valid choice sets")
            df = df[df['obs_id_num'].isin(valid_obs_ids)].copy()
        
        # Debug distributions
        for market in self.markets:
            market_df = df[df['Market'] == market]
            print(f"{market} brand distribution:\n{market_df['Brand'].value_counts()}")
            print(f"{market} choice distribution:\n{market_df[market_df['Chosen'] == 1]['Brand'].value_counts()}")
        
        # Feature selection
        features = ['Price']  # Simplified to Price only
        print(f"Using features: {features}")
        
        # Calculate Actual Brand/Value Shares
        chosen_df = df[df['Chosen'] == 1]
        self.brand_share = chosen_df.groupby(['Market', 'Brand']).size().reset_index(name='Choices')
        self.brand_share['Actual_Brand_Share'] = self.brand_share.groupby('Market')['Choices'].transform(lambda x: x / x.sum())
        self.value_share = chosen_df.groupby(['Market', 'Brand'])['Price'].sum().reset_index(name='Total_Price')
        self.value_share['Actual_Value_Share'] = self.value_share.groupby('Market')['Total_Price'].transform(lambda x: x / x.sum())
        
        for market in self.markets:
            market_df = df[df['Market'] == market].copy()
            if market_df.empty:
                print(f"Warning: No data for {market}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                self.predicted_brand_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share'])
                self.predicted_value_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Value_Share'])
                continue
            
            # Try MNL first
            try:
                pylogit_df = market_df.copy()
                pylogit_df['obs_id'] = pylogit_df['obs_id_num']
                pylogit_df['alt_id'] = pylogit_df['Brand'].map(self.brand_to_id)
                pylogit_df['choice'] = pylogit_df[target]
                
                print(f"{market} pylogit_df dtypes:\n{pylogit_df[features + ['obs_id', 'alt_id', 'choice']].dtypes}")
                print(f"{market} unique obs_id values: {pylogit_df['obs_id'].unique()[:10]}")
                print(f"{market} unique alt_id values: {pylogit_df['alt_id'].unique()}")
                for col in features:
                    print(f"{market} unique {col} values: {len(pylogit_df[col].unique())} unique values")
                print(f"{market} feature correlation matrix:\n{pylogit_df[features].corr()}")
                
                spec = OrderedDict([(var, 'all_same') for var in features])
                name_dict = OrderedDict([(var, var) for var in features])
                
                model = pl.create_choice_model(
                    data=pylogit_df,
                    alt_id_col='alt_id',
                    obs_id_col='obs_id',
                    choice_col='choice',
                    specification=spec,
                    model_type='MNL',
                    names=name_dict
                )
                init_vals = np.array([-0.1])  # Price only
                bounds = [(-1000, 1000)]
                results = model.fit_mle(
                    init_vals,
                    bounds=bounds,
                    print_res=True,
                    method='SLSQP',
                    maxiter=10000,
                    ftol=1e-14,
                    ridge=1e-2
                )
                if results is None:
                    raise ValueError("MNL fitting returned None")
                coef_df = pd.DataFrame({
                    'Coefficient': results.params,
                    'P-value': results.pvalues
                }, index=name_dict.values())
                self.models[market] = results
                self.coef_dfs[market] = coef_df
                print(f"MNL model fitted for {market}: {len(coef_df)} coefficients")
            except Exception as e:
                print(f"MNL failed for {market}: {e}")
                import traceback
                print(f"Stack trace for {market} MNL:\n{traceback.format_exc()}")
                
                # Fallback to binary logit (Cargill vs. others)
                try:
                    print(f"Trying binary logit for {market} (Cargill vs. others)")
                    pylogit_df['choice'] = (pylogit_df['Brand'] == 'Cargill').astype(int)
                    model = pl.create_choice_model(
                        data=pylogit_df,
                        alt_id_col='alt_id',
                        obs_id_col='obs_id',
                        choice_col='choice',
                        specification=spec,
                        model_type='Logit',
                        names=name_dict
                    )
                    results = model.fit_mle(
                        init_vals,
                        bounds=bounds,
                        print_res=True,
                        method='SLSQP',
                        maxiter=10000,
                        ftol=1e-14,
                        ridge=1e-2
                    )
                    if results is None:
                        raise ValueError("Binary logit fitting returned None")
                    coef_df = pd.DataFrame({
                        'Coefficient': results.params,
                        'P-value': results.pvalues
                    }, index=name_dict.values())
                    self.models[market] = results
                    self.coef_dfs[market] = coef_df
                    print(f"Binary logit model fitted for {market}: {len(coef_df)} coefficients")
                except Exception as e:
                    print(f"Binary logit failed for {market}: {e}")
                    print(f"Stack trace for {market} binary logit:\n{traceback.format_exc()}")
                    self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                    self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                    self.price_sensitivity[market] = pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share'])
                    self.predicted_brand_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share'])
                    self.predicted_value_share[market] = pd.DataFrame(columns=['Brand', 'Predicted_Value_Share'])
                    continue
            
            # Calculate feature importance
            importance = self.get_feature_importance(market)
            self.feature_importance[market] = pd.DataFrame({
                'Feature': importance.index,
                'Importance': importance.values
            })
            
            # Simulate price sensitivity
            self.price_sensitivity[market], self.predicted_brand_share[market], self.predicted_value_share[market] = self.simulate_price_sensitivity(market_df, market, 'Cargill', features)
        
        return self
    
    def get_attribute_utilities(self, market):
        coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
        return coef_df['Coefficient'] if 'Coefficient' in coef_df.columns else pd.Series(dtype=float)
    
    def get_feature_importance(self, market):
        utilities = self.get_attribute_utilities(market)
        if utilities.empty:
            return pd.Series(dtype=float)
        importance = utilities.abs() / utilities.abs().sum() if utilities.abs().sum() > 0 else utilities
        return importance
    
    def simulate_price_sensitivity(self, market_df, market, brand='Cargill', features=None):
        if market not in self.models or isinstance(self.models[market], pd.DataFrame):
            print(f"Warning: No valid model for {market}. Skipping price sensitivity analysis.")
            return (pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
        
        coefs = self.get_attribute_utilities(market)
        if coefs.empty:
            print(f"Warning: No coefficients for {market}. Skipping price sensitivity analysis.")
            return (pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']),
                    pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
        
        choice_sets = market_df.groupby('obs_id_num')
        choice_set_sizes = choice_sets.size()
        print(f"{market} choice set sizes:\n{choice_set_sizes.value_counts()}")
        print(f"Number of choice sets in {market}: {len(choice_sets)}")
        print(f"Cargill present in {market}: {brand in market_df['Brand'].values}")
        
        price_changes = [round(x, 2) for x in np.arange(-0.10, 0.11, 0.01)]
        results = []
        predicted_brand_share = None
        predicted_value_share = None
        
        baseline_prices = market_df.groupby('Brand')['Price'].mean().to_dict()
        print(f"Baseline prices for {market}: {baseline_prices}")
        
        for change in price_changes:
            utilities = []
            for choice_id, group in choice_sets:
                group = group.copy()
                cargill_mask = group['Brand'] == brand
                if cargill_mask.any():
                    base_price = group.loc[cargill_mask, 'Price'].iloc[0]
                    if base_price <= 0:
                        print(f"Warning: Non-positive base price for Cargill in choice set {choice_id}, {market}. Setting to 100.0.")
                        base_price = 100.0
                    group.loc[cargill_mask, 'Price'] = base_price * (1 + change)
                
                group_utilities = []
                for _, row in group.iterrows():
                    utility = 0
                    for var in features:
                        if var in coefs.index and var in row.index:
                            utility += coefs[var] * row[var]
                    group_utilities.append(utility)
                
                print(f"Choice set {choice_id} utilities: {group_utilities}")
                exp_utilities = np.exp(np.minimum(group_utilities, 500))
                sum_exp = exp_utilities.sum()
                probs = exp_utilities / sum_exp if sum_exp > 0 else np.ones(len(group_utilities)) / len(group_utilities)
                probs = np.round(probs, 10)
                prob_sum = probs.sum()
                if not np.isclose(prob_sum, 1.0, atol=1e-8):
                    probs = probs / prob_sum if prob_sum > 0 else probs
                print(f"Choice set {choice_id} probabilities: {probs.tolist()}")
                
                for i, (_, row) in enumerate(group.iterrows()):
                    utilities.append({
                        'ChoiceID': choice_id,
                        'Brand': row['Brand'],
                        'Probability': probs[i],
                        'Price': row['Price']
                    })
            
            if utilities:
                util_df = pd.DataFrame(utilities)
                choice_set_weights = choice_set_sizes / choice_set_sizes.sum()
                weights = util_df['ChoiceID'].map(choice_set_weights).fillna(0)
                
                brand_share = (util_df[util_df['Brand'] == brand]['Probability'] * weights[util_df['Brand'] == brand]).sum() if brand in util_df['Brand'].values else 0.0
                
                value_sum = 0
                brand_value = 0
                for b in util_df['Brand'].unique():
                    brand_mask = util_df['Brand'] == b
                    brand_prob = (util_df[brand_mask]['Probability'] * weights[brand_mask]).sum()
                    brand_price = util_df[brand_mask]['Price'].mean()
                    value_sum += brand_prob * brand_price
                    if b == brand:
                        brand_value = brand_prob * brand_price
                value_share = brand_value / value_sum if value_sum > 0 else 0.0
                
                results.append({
                    'Price_Change_Percent': change * 100,
                    'Brand_Share': brand_share,
                    'Value_Share': value_share
                })
                
                if change == 0:
                    predicted_brand_share = util_df.groupby('Brand').apply(
                        lambda x: (x['Probability'] * weights[x.index]).sum(),
                        include_groups=False
                    ).reset_index(name='Predicted_Brand_Share')
                    predicted_brand_share['Predicted_Brand_Share'] = predicted_brand_share['Predicted_Brand_Share'].round(10)
                    brand_share_sum = predicted_brand_share['Predicted_Brand_Share'].sum()
                    print(f"Sum of Predicted_Brand_Share for {market}: {brand_share_sum:.6f}")
                    if not np.isclose(brand_share_sum, 1.0, atol=1e-8):
                        print(f"Warning: Predicted_Brand_Share sum for {market} is {brand_share_sum:.6f}, adjusting to sum to 1.0")
                        predicted_brand_share['Predicted_Brand_Share'] = predicted_brand_share['Predicted_Brand_Share'] / brand_share_sum if brand_share_sum > 0 else 0.0
                    
                    value_sum_all = 0
                    predicted_value_share = []
                    for b in util_df['Brand'].unique():
                        brand_mask = util_df['Brand'] == b
                        brand_prob = (util_df[brand_mask]['Probability'] * weights[brand_mask]).sum()
                        brand_price = util_df[brand_mask]['Price'].mean()
                        value_sum_all += brand_prob * brand_price
                        predicted_value_share.append({'Brand': b, 'Predicted_Value_Share': brand_prob * brand_price})
                    predicted_value_share = pd.DataFrame(predicted_value_share)
                    predicted_value_share['Predicted_Value_Share'] = predicted_value_share['Predicted_Value_Share'] / value_sum_all if value_sum_all > 0 else 0.0
                    predicted_value_share['Predicted_Value_Share'] = predicted_value_share['Predicted_Value_Share'].round(10)
                    value_share_sum = predicted_value_share['Predicted_Value_Share'].sum()
                    print(f"Sum of Predicted_Value_Share for {market}: {value_share_sum:.6f}")
                    if not np.isclose(value_share_sum, 1.0, atol=1e-8):
                        print(f"Warning: Predicted_Value_Share sum for {market} is {value_share_sum:.6f}, adjusting to sum to 1.0")
                        predicted_value_share['Predicted_Value_Share'] = predicted_value_share['Predicted_Value_Share'] / value_sum_all if value_sum_all > 0 else 0.0
                    
                    cargill_predicted = predicted_brand_share[predicted_brand_share['Brand'] == brand]['Predicted_Brand_Share'].iloc[0] if brand in predicted_brand_share['Brand'].values else 0.0
                    cargill_price_sensitivity = brand_share if brand in util_df['Brand'].values else 0.0
                    if not np.isclose(cargill_predicted, cargill_price_sensitivity, atol=1e-6):
                        print(f"Warning: Cargill Predicted_Brand_Share ({cargill_predicted:.6f}) does not match Brand_Share at 0% ({cargill_price_sensitivity:.6f}) in {market}")
        
        result_df = pd.DataFrame(results)
        if not result_df.empty:
            price_coef = coefs.get('Price', 0)
            brand_share_diff = result_df['Brand_Share'].diff().dropna()
            if price_coef < 0 and not all(brand_share_diff <= 0):
                print(f"Warning: Brand_Share in {market} does not consistently decrease with price increase (price coef={price_coef:.6f})")
            elif price_coef > 0 and not all(brand_share_diff >= 0):
                print(f"Warning: Brand_Share in {market} does not consistently increase with price increase (price coef={price_coef:.6f})")
            value_share_diff = result_df['Value_Share'].diff().dropna()
            if price_coef < 0 and not all(value_share_diff <= 0):
                print(f"Warning: Value_Share in {market} does not consistently decrease with price increase (price_coef={price_coef:.6f})")
        
        return (result_df, predicted_brand_share, predicted_value_share)
    
    def add_model_summaries_to_workbook(self, workbook):
        for market in self.markets:
            if f"{market}_summary" in workbook.sheetnames:
                sheet = workbook[f"{market}_summary"]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sheet = workbook.create_sheet(title=f"{market}_summary")
            
            coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
            print(f"Writing {len(coef_df)} coefficients for {market}")
            sheet['A1'] = 'Feature'
            sheet['B1'] = 'Coefficient'
            sheet['C1'] = 'P-value'
            for r, (index, row) in enumerate(coef_df.iterrows(), start=2):
                sheet[f'A{r}'] = index
                sheet[f'B{r}'] = row.get('Coefficient', float('nan'))
                sheet[f'C{r}'] = row.get('P-value', float('nan'))
            
            importance_df = self.feature_importance.get(market, pd.DataFrame(columns=['Feature', 'Importance']))
            if not importance_df.empty:
                start_row = len(coef_df) + 6
                print(f"Writing {len(importance_df)} feature importance rows for {market} at row {start_row}")
                sheet[f'A{start_row}'] = 'Feature Importance'
                sheet[f'A{start_row + 1}'] = 'Feature'
                sheet[f'B{start_row + 1}'] = 'Importance'
                for r, (index, row) in enumerate(importance_df.iterrows(), start=start_row + 2):
                    sheet[f'A{r}'] = row['Feature']
                    sheet[f'B{r}'] = row['Importance']
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            if f"{market}_brand_share" in workbook.sheetnames:
                brand_sheet = workbook[f"{market}_brand_share"]
                for row in brand_sheet.iter_rows(min_row=1, max_row=brand_sheet.max_row, min_col=1, max_col=brand_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                brand_sheet = workbook.create_sheet(title=f"{market}_brand_share")
            brand_share_market = self.brand_share[self.brand_share['Market'] == market][['Brand', 'Actual_Brand_Share']]
            predicted_brand_share = self.predicted_brand_share.get(market, pd.DataFrame(columns=['Brand', 'Predicted_Brand_Share']))
            brand_share_market = brand_share_market.merge(predicted_brand_share, on='Brand', how='outer').fillna(0.0).infer_objects(copy=False)
            print(f"Writing {len(brand_share_market)} brand share rows for {market}")
            brand_sheet['A1'] = 'Brand'
            brand_sheet['B1'] = 'Actual_Brand_Share'
            brand_sheet['C1'] = 'Predicted_Brand_Share'
            for r, (index, row) in enumerate(brand_share_market.iterrows(), start=2):
                brand_sheet[f'A{r}'] = row['Brand']
                brand_sheet[f'B{r}'] = row['Actual_Brand_Share']
                brand_sheet[f'C{r}'] = row['Predicted_Brand_Share']
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            if f"{market}_value_share" in workbook.sheetnames:
                value_sheet = workbook[f"{market}_value_share"]
                for row in value_sheet.iter_rows(min_row=1, max_row=value_sheet.max_row, min_col=1, max_col=value_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                value_sheet = workbook.create_sheet(title=f"{market}_value_share")
            value_share_market = self.value_share[self.value_share['Market'] == market][['Brand', 'Actual_Value_Share']]
            predicted_value_share = self.predicted_value_share.get(market, pd.DataFrame(columns=['Brand', 'Predicted_Value_Share']))
            value_share_market = value_share_market.merge(predicted_value_share, on='Brand', how='outer').fillna(0.0).infer_objects(copy=False)
            print(f"Writing {len(value_share_market)} value share rows for {market}")
            value_sheet['A1'] = 'Brand'
            value_sheet['B1'] = 'Actual_Value_Share'
            value_sheet['C1'] = 'Predicted_Value_Share'
            for r, (index, row) in enumerate(value_share_market.iterrows(), start=2):
                value_sheet[f'A{r}'] = row['Brand']
                value_sheet[f'B{r}'] = row['Actual_Value_Share']
                value_sheet[f'C{r}'] = row['Predicted_Value_Share']
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = value_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                value_sheet.column_dimensions[column_letter].width = adjusted_width
            
            if f"{market}_price_sensitivity" in workbook.sheetnames:
                sensitivity_sheet = workbook[f"{market}_price_sensitivity"]
                for row in sensitivity_sheet.iter_rows(min_row=1, max_row=sensitivity_sheet.max_row, min_col=1, max_col=sensitivity_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sensitivity_sheet = workbook.create_sheet(title=f"{market}_price_sensitivity")
            sensitivity_df = self.price_sensitivity.get(market, pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']))
            print(f"Writing {len(sensitivity_df)} price sensitivity rows for {market}")
            sensitivity_sheet['A1'] = 'Price_Change_Percent'
            sensitivity_sheet['B1'] = 'Brand_Share'
            sensitivity_sheet['C1'] = 'Value_Share'
            for r, (index, row) in enumerate(sensitivity_df.iterrows(), start=2):
                sensitivity_sheet[f'A{r}'] = row['Price_Change_Percent']
                sensitivity_sheet[f'B{r}'] = row['Brand_Share']
                sensitivity_sheet[f'C{r}'] = row['Value_Share']
            sensitivity_sheet['A{0}'.format(len(sensitivity_df) + 4)] = 'Note: Line chart can be created in Excel using the data above.'
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = sensitivity_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sensitivity_sheet.column_dimensions[column_letter].width = adjusted_width    
 

if __name__ == "__main__":
    df = pd.read_csv("combined_choice_data_v2.csv")
    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    model = DiscreteChoiceModel(markets=['Maharashtra', 'Gujarat', 'Punjab', 'South'], use_pylogit=True)
    model.fit(df)
    model.add_model_summaries_to_workbook(workbook)
    workbook.save('cargill_findings_v24.xlsx')