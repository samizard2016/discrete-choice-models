import logging
import os
import pandas as pd
import numpy as np
import random
from itertools import combinations
import statsmodels.api as sm
import matplotlib.pyplot as plt
import seaborn as sns
import shap
from statsmodels.stats.outliers_influence import variance_inflation_factor
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.line import LineProperties
import warnings
import textwrap
import pickle

warnings.filterwarnings('ignore', category=FutureWarning)

class EnhancedBayesianChoiceAnalyzer:
    def __init__(self, profiles=None, choices=None, groups=None, rename_dict=None):
        # Clear all existing handlers to prevent conflicts
        logging.getLogger().handlers.clear()

        # Use a non-OneDrive directory for logs
        log_file = r'C:\logs\kdcm_simulator.log'
        os.makedirs(r'C:\logs', exist_ok=True)
        # Alternative: OneDrive path
        # log_file = r'C:\Users\PaulSa\OneDrive - Kantar\Documents\works 2025\discrete_choice_modeling\schneider\revised model\kdcm_simulator.log'

        # Create named logger
        self.logger = logging.getLogger("kdcm_logger")
        self.logger.setLevel(logging.DEBUG)
        self.logger.propagate = False  # Disable propagation to avoid duplicates

        # Add file handler with no buffering
        file_handler = logging.FileHandler(log_file, mode='a', delay=False)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(logging.Formatter(
            "%(asctime)s %(levelname)-8s [%(filename)s:%(lineno)d] %(message)s",
            datefmt="%d-%m-%Y %H:%M:%S"
        ))

        # Add console handler for debugging
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.DEBUG)
        console_handler.setFormatter(logging.Formatter(
            "%(asctime)s %(levelname)-8s [%(filename)s:%(lineno)d] %(message)s",
            datefmt="%d-%m-%Y %H:%M:%S"
        ))

        # Add handlers only if not already present
        if not self.logger.handlers:
            self.logger.addHandler(file_handler)
            self.logger.addHandler(console_handler)

        self.logger.info("A new session of KDCM Simulator started at 06:03 PM IST, July 03, 2025")
        for handler in self.logger.handlers:
            if isinstance(handler, logging.FileHandler):
                handler.flush()

        self.logger.debug(f"Root logger handlers: {logging.getLogger().handlers}")
        self.logger.debug(f"kdcm_logger handlers: {self.logger.handlers}")

        # Initialize DiscreteChoiceAnalyzer attributes
        self.profiles = profiles.copy() if profiles is not None else None
        self.choices = choices.copy() if choices is not None else None
        self.groups = groups.copy() if groups is not None else None
        self.rename_dict = rename_dict or {}
        self.model = None
        self.utilities = None
        self.feature_importance = None
        self.choice_data = None
        self.price_mean = None
        self.price_std = None

    def save_dataframe_to_excel(self, df, output_file, sheet_name):
        """
        Save a DataFrame to an Excel file with a specified sheet name.
        
        Parameters:
        - df: pandas DataFrame to save
        - output_file: Path to the output Excel file
        - sheet_name: Name of the sheet to save the DataFrame to
        """
        try:
            self.logger.info(f"Saving DataFrame to {output_file} in sheet '{sheet_name}' at 06:03 PM IST, July 03, 2025")
            df.to_excel(output_file, sheet_name=sheet_name, index=True, engine='openpyxl')
            self.logger.info(f"DataFrame saved successfully to {output_file} in sheet '{sheet_name}'")
            
            # Flush logs
            for handler in self.logger.handlers:
                if isinstance(handler, logging.FileHandler):
                    handler.flush()
        except Exception as e:
            self.logger.error(f"Error saving DataFrame to {output_file}: {str(e)}")
            raise

    def create_market_share_line_chart(self, input_file="price_scenario_shares.xlsx", output_file="price_shares_with_chart.xlsx", data_sheet="Market_Shares"):
        """
        Create a line chart with all price columns as series in a 'plot' sheet, with reduced line thickness.
        """
        try:
            self.logger.info("Starting line chart creation for market shares at 06:03 PM IST, July 03, 2025")

            # Load the workbook
            workbook = openpyxl.load_workbook(input_file)
            data_worksheet = workbook[data_sheet]

            # Create a new "plot" sheet
            if "plot" in workbook.sheetnames:
                workbook.remove(workbook["plot"])
            plot_worksheet = workbook.create_sheet("plot")

            # Get all column headers (skip 'Profile' in column A)
            headers = [cell.value for cell in data_worksheet[1] if cell.value]
            if "Profile" not in headers:
                raise ValueError("Expected 'Profile' column in the data")
            price_columns = [i + 1 for i, header in enumerate(headers) if header != "Profile"]

            # Define categories (Profile column)
            categories = Reference(data_worksheet, min_col=1, min_row=2, max_row=data_worksheet.max_row)

            # Create line chart
            chart = LineChart()
            chart.title = "Simulated Price Shares by Profile"
            chart.style = 10
            chart.x_axis.title = "Profile"
            chart.y_axis.title = "Price Share (%)"

            # Add each price column as a series
            for col in price_columns:
                data = Reference(data_worksheet, min_col=col, min_row=1, max_row=data_worksheet.max_row)
                chart.add_data(data, titles_from_data=True)

            # Set line thickness for all series
            for series in chart.series:
                series.graphicalProperties.line = LineProperties(w=0.75 * 12700)  # 0.75 points

            chart.set_categories(categories)

            # Add chart to the "plot" sheet
            plot_worksheet.add_chart(chart, "A1")

            # Save the workbook
            workbook.save(output_file)
            self.logger.info(f"Line chart with {len(price_columns)} series (thinner lines) created in 'plot' sheet and saved to {output_file}")

            # Flush logs
            for handler in self.logger.handlers:
                if isinstance(handler, logging.FileHandler):
                    handler.flush()

        except Exception as e:
            self.logger.error(f"Error creating market share line chart: {str(e)}")
            raise

    def evaluate_price_scenario_lp(self, price_scenarios, profile_labels=None, plot=True):
        """
        Evaluate market shares for price scenarios, ensuring all profiles are included.
        """
        if self.model is None:
            raise ValueError("Model is not fitted yet. Call fit_model() first.")
        
        self.logger.info("Starting price scenario evaluation at 06:03 PM IST, July 03, 2025")
        
        # Use full profiles without filtering
        profiles = self.profiles.copy()
        profiles['Price'] = pd.to_numeric(profiles['Price'], errors='coerce')
        
        # Replace typos in AF column
        profiles['AF'] = profiles['AF'].replace({
            'Current Measuremnt and time stamped fault records on mobile app': 
            'Current Measurement and time stamped fault records on mobile app',
            'Scalable connectivity at breaker level- Basic Modbus (Breaker Status, contorl, terminal temp alarm)': 
            'Scalable connectivity at breaker level- Basic Modbus (Breaker Status, control, terminal temp alarm)'
        })
        
        # Encode profiles
        profiles_encoded = pd.get_dummies(
            profiles,
            columns=['SP', 'AF'],
            drop_first=True,
            dtype=float
        )
        
        # Rename columns
        rename_dict = {
            'SP_High performance Ics=Icu=Icw 66kA for 1sec': 'Size_Perf_High',
            'AF_Higher Electrical life from 6,000 to 7,500 operations without maintenance': 'Adv_Feat_ElecLife',
            'AF_Visible Health indication (Breaker Status, trip cause Indication - OL,SC,GF)': 'Adv_Feat_Health',
            'AF_Scalable connectivity at breaker level- Basic Modbus (Breaker Status, control, terminal temp alarm)': 'Adv_Feat_ModbusBasic',
            'AF_Current Measurement and time stamped fault records on mobile app': 'Adv_Feat_Current',
            'AF_Scalable connectivity at breaker level- Modbus Ethernet': 'Adv_Feat_ModbusEth',
            'AF_Operator Safety - Arc Flash reduction during maintenance': 'Adv_Feat_Safety'
        }
        profiles_encoded.rename(columns=rename_dict, inplace=True)
        
        # Ensure all required columns are present
        X_cols = [col for col in self.utilities.index if col != 'const' and not col.endswith('_Panel')]
        for col in X_cols:
            if col not in profiles_encoded.columns:
                profiles_encoded[col] = 0.0
        
        # Center price as in prepare_data
        profiles_encoded['Price'] = -(profiles['Price'].values / 1000 - self.price_mean / 1000)
        
        # Validate input price scenarios
        if isinstance(price_scenarios, dict):
            scenarios_df = pd.DataFrame(price_scenarios, dtype=float)
        else:
            scenarios_df = price_scenarios.astype(float)
        
        # Ensure scenarios_df index matches profiles_encoded index
        expected_indices = profiles_encoded.index
        if not scenarios_df.index.equals(expected_indices):
            self.logger.warning(f"Price scenarios index ({scenarios_df.index.tolist()}) does not match profiles index ({expected_indices.tolist()}). Reindexing.")
            scenarios_df = scenarios_df.reindex(expected_indices, fill_value=profiles['Price'].mean())
        
        shares = pd.DataFrame(index=profiles_encoded.index, columns=scenarios_df.columns)
        
        # Calculate shares for each scenario
        for scenario in scenarios_df.columns:
            temp_profiles = profiles_encoded.copy()
            temp_profiles['Price'] = -scenarios_df[scenario].values / 1000  # Update all prices
            X = temp_profiles[X_cols]
            X = sm.add_constant(X, has_constant='add')
            utilities = self.utilities[['const'] + X_cols]
            V = X @ utilities
            exp_V = np.exp(V)
            sum_exp_V = exp_V.sum()
            probabilities = exp_V / sum_exp_V
            shares[scenario] = probabilities
        
        # Apply profile labels
        if profile_labels is None:
            profile_labels = [f'Profile {i+1}' for i in range(len(profiles_encoded))]
        elif isinstance(profile_labels, dict):
            profile_labels = [profile_labels.get(i, f'Profile {i+1}') for i in profiles_encoded.index]
        elif isinstance(profile_labels, list):
            if len(profile_labels) != len(profiles_encoded.index):
                raise ValueError(f"profile_labels list length ({len(profile_labels)}) must match number of profiles ({len(profiles_encoded)})")
        
        shares.index = profile_labels
        
        # Save shares to Excel
        shares_output = shares * 100  # Convert to percentages
        shares_output.reset_index(inplace=True)
        shares_output.rename(columns={'index': 'Profile'}, inplace=True)
        self.save_dataframe_to_excel(shares_output, "price_scenario_shares.xlsx", sheet_name="Market_Shares")
        
        # Create line chart
        self.create_market_share_line_chart(input_file="price_scenario_shares.xlsx", output_file="price_shares_with_chart.xlsx", data_sheet="Market_Shares")
        
        # Plot using matplotlib if requested
        if plot:
            plot_data = shares.reset_index().melt(id_vars='index', var_name='Scenario', value_name='Share')
            plot_data['Profile'] = plot_data['index'].map({i: label for i, label in enumerate(profile_labels)})
            
            plt.figure(figsize=(12, 6))
            sns.lineplot(data=plot_data, x='Profile', y='Share', hue='Scenario', style='Scenario', markers=True, dashes=False)
            plt.title('Profile Shares Across Price Scenarios')
            plt.xlabel('Profile')
            plt.ylabel('Choice Probability (Share, %)')
            plt.xticks(rotation=45, ha='right')
            plt.legend(title='Scenario', bbox_to_anchor=(1.05, 1), loc='upper left')
            plt.tight_layout()
            plt.savefig('profile_shares_line.png')
            plt.close()
        
        self.logger.info(f"Price scenario evaluation completed. Shares calculated for {len(shares)} profiles.")
        return shares * 100

    # Include other DiscreteChoiceAnalyzer methods
    def validate_data(self):
        choice_data = self.choice_data.copy()
        choice_data['Price'] = (choice_data['Price'] + self.price_mean / 1000) * 1000
        price_choice = choice_data.groupby('chosen')['Price'].mean()
        self.logger.info(f"Mean Price by Chosen Status:\n {price_choice}")
        if price_choice[1] > price_choice[0]:
            self.logger.warning("Warning: Chosen profiles have higher mean price, possible data issue.")

        choices = self.choices.copy()
        def parse_profiles(x):
            try:
                if isinstance(x, str):
                    return [int(v.strip()) for v in x.strip('[]').split(',')]
                return x
            except:
                return []
        choices['parsed_profiles'] = choices['profiles_presented'].apply(parse_profiles)
        invalid = choices[choices.apply(lambda x: x['chosen_profile'] not in x['parsed_profiles'], axis=1)]
        self.logger.info(f"Invalid chosen profiles not in profiles_presented: {len(invalid)}")
        if len(invalid) > 0:
            self.logger.warning("Warning: Some chosen profiles are not in profiles_presented, indicating data errors.")
            self.logger.warning(f"Invalid choices sample:\n{invalid[['chosen_profile', 'profiles_presented']].head()}")

        choice_sets = self.choices.copy()
        correlations_modbus = []
        for _, row in choice_sets.iterrows():
            profiles = parse_profiles(row['profiles_presented'])
            if not all(p in self.profiles.index for p in profiles):
                continue
            set_data = self.profiles.loc[profiles].copy()
            set_data['Price'] = set_data['Price'] / 1000
            set_encoded = pd.get_dummies(set_data, columns=['SP', 'AF'], drop_first=True, dtype=float)
            set_encoded.rename(columns=self.rename_dict, inplace=True)
            if 'Adv_Feat_ModbusBasic' in set_encoded.columns and 'Price' in set_encoded.columns:
                corr_modbus = set_encoded[['Price', 'Adv_Feat_ModbusBasic']].corr().iloc[0, 1]
                if not np.isnan(corr_modbus):
                    correlations_modbus.append(corr_modbus)
        if correlations_modbus:
            self.logger.info(f"Mean Price-Adv_Feat_ModbusBasic correlation: {np.mean(correlations_modbus):.4f}")
        else:
            self.logger.info("No Price-Adv_Feat_ModbusBasic correlations computed.")

        respondent_choices = choice_data.groupby('respondent_id')['chosen'].sum()
        self.logger.info(f"Respondent choice counts:\n{respondent_choices.value_counts()}")
        if (respondent_choices == 0).any():
            self.logger.warning("Warning: Some respondents never chose a profile (possible inattention).")
        choice_set_counts = choice_data.groupby('respondent_id')['choice_set'].nunique()
        self.logger.info(f"Choice sets per respondent:\n{choice_set_counts.value_counts()}")
        if choice_set_counts.max() < 16:
            self.logger.warning(f"Warning: Maximum choice sets per respondent ({choice_set_counts.max()}) is below expected (16).")

        return price_choice

    def prepare_data(self):
        profiles = self.profiles.copy()
        self.logger.info(f"Uncentered Price range:\n{profiles['Price'].describe()}")
        profiles['Price'] = pd.to_numeric(profiles['Price'], errors='coerce')
        profiles['Price'] = profiles['Price'] / 1000
        self.price_mean = profiles['Price'].mean() * 1000
        self.price_std = profiles['Price'].std() * 1000
        
        profiles['AF'] = profiles['AF'].replace({
            'Current Measuremnt and time stamped fault records on mobile app':
            'Current Measurement and time stamped fault records on mobile app',
            'Scalable connectivity at breaker level- Basic Modbus (Breaker Status, contorl, terminal temp alarm)':
            'Scalable connectivity at breaker level- Basic Modbus (Breaker Status, control, terminal temp alarm)'
        })
        
        self.logger.info(f"Unique SP values: {profiles['SP'].unique()}")
        self.logger.info(f"Unique AF values: {sorted(profiles['AF'].unique())}")
        self.logger.info(f"Dropped AF level (first alphabetically): {sorted(profiles['AF'].unique())[0]}")
        
        profiles = profiles[profiles['AF'] != 'Terminal Temperature threshold monitoring']
        profiles_encoded = pd.get_dummies(
            profiles,
            columns=['SP', 'AF'],
            drop_first=True,
            dtype=float
        )
        
        self.logger.info(f"profiles_encoded columns before renaming: {profiles_encoded.columns.tolist()}")
        
        profiles_encoded.rename(columns=self.rename_dict, inplace=True)
        
        self.logger.info(f"profiles_encoded columns after renaming: {profiles_encoded.columns.tolist()}")
        self.logger.info(f"profiles_encoded shape: {profiles_encoded.shape}")
        
        valid_profiles = profiles_encoded.index.tolist()
        choices_filtered = self.choices.copy()
        initial_rows = len(choices_filtered)

        def parse_profiles(x):
            try:
                if isinstance(x, str):
                    return [int(v.strip()) for v in x.strip('[]').split(',')]
                return x
            except:
                return []
        choices_filtered['parsed_profiles'] = choices_filtered['profiles_presented'].apply(parse_profiles)
        choices_filtered = choices_filtered[choices_filtered['chosen_profile'].isin(valid_profiles)]
        choices_filtered = choices_filtered[choices_filtered['parsed_profiles'].apply(lambda x: all(p in valid_profiles for p in x))]
        self.logger.info(f"Dropping {initial_rows - len(choices_filtered)} choices with invalid profiles (chosen or presented): {set(choices_filtered['chosen_profile'].unique()) - set(valid_profiles)}")
        
        def filter_profiles_presented(profiles, valid_profiles):
            valid = [p for p in profiles if p in valid_profiles]
            return str(valid) if len(valid) >= 2 else None
        
        choices_filtered['profiles_presented'] = choices_filtered['parsed_profiles'].apply(
            lambda x: filter_profiles_presented(x, valid_profiles)
        )
        invalid_sets = choices_filtered[choices_filtered['profiles_presented'].isnull()]
        self.logger.info(f"Dropping {len(invalid_sets)} choice sets with fewer than 2 valid profiles")
        choices_filtered = choices_filtered[choices_filtered['profiles_presented'].notnull()]
        final_rows = len(choices_filtered)
        self.logger.info(f"Choices rows: {initial_rows} initial, {final_rows} after filtering ({initial_rows - final_rows} dropped)")
        
        if not all(choices_filtered['respondent_id'].isin(self.groups['respondent_id'])):
            missing_ids = choices_filtered[~choices_filtered['respondent_id'].isin(self.groups['respondent_id'])]['respondent_id'].unique()
            raise ValueError(f"respondent_id values in choices not found in groups: {missing_ids}")
        
        self.logger.info("Validation done. Attempting to create choice set data ...")
        
        choice_sets = []
        for choice_set in choices_filtered['choice_set'].unique():
            respondents = choices_filtered[choices_filtered['choice_set'] == choice_set]
            for _, respondent in respondents.iterrows():
                profiles_presented = parse_profiles(respondent['profiles_presented'])
                if not isinstance(profiles_presented, list):
                    continue
                for idx in profiles_presented:
                    if idx not in profiles_encoded.index:
                        continue
                    profile = profiles_encoded.loc[idx]
                    group_row = self.groups[self.groups['respondent_id'] == respondent['respondent_id']]
                    if group_row.empty:
                        raise ValueError(f"No group found for respondent_id {respondent['respondent_id']}")
                    group = 'Panel builder' if group_row['group'].iloc[0] == 'Panel builder' else 'Others'
                    row = {
                        **profile.to_dict(),
                        'respondent_id': int(respondent['respondent_id']),
                        'choice_set': int(choice_set),
                        'group': str(group),
                        'chosen': int(1 if idx == respondent['chosen_profile'] else 0)
                    }
                    choice_sets.append(row)
        
        self.choice_data = pd.DataFrame(choice_sets)
        
        for col in self.choice_data.columns:
            if self.choice_data[col].dtype == bool:
                self.choice_data[col] = self.choice_data[col].astype(float)
        
        unnamed_cols = [col for col in self.choice_data.columns if 'Unnamed' in col]
        if unnamed_cols:
            self.logger.info(f"Dropping Unnamed columns from choice_data: {unnamed_cols}")
            self.choice_data = self.choice_data.drop(columns=unnamed_cols)
        
        self.choice_data['Price'] = -(self.choice_data['Price'] - self.choice_data['Price'].mean())
        self.logger.info(f"Price range (centered, thousands):\n{self.choice_data['Price'].describe()}")
        
        self.logger.info(f"Group value counts:\n{self.choice_data['group'].value_counts()}")
        self.logger.info(f"Sample of choice_data:\n{self.choice_data[['respondent_id', 'choice_set', 'group', 'chosen']].head()}")
        self.logger.info(f"Final choice_data columns: {self.choice_data.columns.tolist()}")
        self.logger.info(f"Final choice_data dtypes:\n{self.choice_data.dtypes}")
        
        numeric_cols = [col for col in self.choice_data.columns if col not in ['respondent_id', 'choice_set', 'group', 'chosen']]
        if self.choice_data[numeric_cols].isna().any().any():
            raise ValueError("Missing values in numeric columns")
        for col in numeric_cols:
            if not pd.api.types.is_numeric_dtype(self.choice_data[col]):
                raise ValueError(f"Non-numeric column {col}: {self.choice_data[col].dtype}")
        
        self.validate_data()
        return self.choice_data

    def fit_model(self, group=None, use_interactions=False):
        X_cols = [
            'Price', 'Size_Perf_High', 'Adv_Feat_ElecLife', 'Adv_Feat_Current',
            'Adv_Feat_Safety', 'Adv_Feat_Health'
        ]
        
        if not all(col in self.choice_data.columns for col in X_cols):
            missing_cols = [col for col in X_cols if col not in self.choice_data.columns]
            raise ValueError(f"Missing columns in choice_data: {missing_cols}")
        
        data = self.choice_data if group is None else self.choice_data[self.choice_data['group'] == group]
        if group is None and use_interactions:
            data = self.choice_data.copy()
            data['Panel_builder'] = (data['group'] == 'Panel builder').astype(float)
            X = data[X_cols].copy()
            X['Price_Panel'] = X['Price'] * data['Panel_builder']
            X_cols.append('Price_Panel')
        else:
            X = data[X_cols]
        
        y = data['chosen']
        
        self.logger.info(f"Fitting model for {'all groups' if group is None else group}, interactions={use_interactions}")
        self.logger.info(f"X_cols: {X_cols}")
        self.logger.info(f"X shape: {X.shape}")
        self.logger.info(f"Correlation matrix:\n{X.corr()}")
        
        vif_data = pd.DataFrame()
        vif_data['Feature'] = X_cols
        vif_data['VIF'] = [variance_inflation_factor(X.values, i) for i in range(X.shape[1])]
        self.logger.info(f"VIF:\n{vif_data}")
        
        non_numeric_cols = [col for col in X.columns if not pd.api.types.is_numeric_dtype(X[col])]
        if non_numeric_cols:
            raise ValueError(f"Non-numeric columns in X: {non_numeric_cols}")
        if not pd.api.types.is_numeric_dtype(y):
            raise ValueError(f"Non-numeric y: {y.dtype}")
        
        X = sm.add_constant(X, has_constant='add')
        try:
            self.model = sm.Logit(y, X).fit(disp=0, maxiter=2000, method='bfgs')
            if not self.model.mle_retvals['converged']:
                self.logger.warning("bfgs failed to converge, trying newton...")
                self.model = sm.Logit(y, X).fit(disp=0, maxiter=2000, method='newton')
        except Exception as e:
            raise RuntimeError(f"Logit model fitting failed: {str(e)}")
        
        self.logger.info(f"Model Converged: {self.model.mle_retvals['converged']}")
        self.logger.info(f"Model Summary:\n{self.model.summary()}")
        
        self.utilities = self.model.params
        return self.utilities

    def calculate_feature_importance(self):
        if self.model is None:
            raise ValueError("Model is not fitted yet. Call fit_model() first.")
        
        X_cols = [col for col in self.utilities.index if col != 'const' and not col.endswith('_Panel')]
        if not X_cols:
            raise ValueError("No valid columns for feature importance after excluding interaction terms.")
        X = self.choice_data[X_cols]
        
        coef = self.model.params[X_cols].values
        intercept = self.model.params['const']
        explainer = shap.LinearExplainer((coef, intercept), shap.maskers.Independent(X))
        shap_values = explainer.shap_values(X)
        
        importance = np.abs(shap_values).mean(axis=0)
        total = importance.sum()
        self.feature_importance = {col: imp / total for col, imp in zip(X_cols, importance)}
        
        shap.summary_plot(shap_values, X, feature_names=X_cols, plot_type="bar", show=False)
        plt.savefig('shap_importance.png')
        plt.close()
        
        return self.feature_importance

    def save_model(self, file_path):
        if self.model is None or self.utilities is None:
            raise ValueError("Model not trained. Call fit_model() first.")
        
        with open(f"{file_path}_model.pkl", "wb") as f:
            pickle.dump({'model': self.model, 'utilities': self.utilities}, f)
        
        with open(f"{file_path}_data.pkl", "wb") as f:
            pickle.dump({
                'price_mean': self.price_mean,
                'price_std': self.price_std,
                'profiles': self.profiles,
                'choice_data': self.choice_data
            }, f)
        
        self.logger.info(f"Model and data saved to {file_path}_model.pkl and {file_path}_data.pkl")

    def restore_model(self, file_path):
        try:
            with open(f"{file_path}_model.pkl", "rb") as f:
                model_data = pickle.load(f)
                self.model = model_data['model']
                self.utilities = model_data['utilities']
        except FileNotFoundError:
            raise FileNotFoundError(f"Model file {file_path}_model.pkl not found")
        
        try:
            with open(f"{file_path}_data.pkl", "rb") as f:
                data = pickle.load(f)
                self.price_mean = data['price_mean']
                self.price_std = data['price_std']
                self.profiles = data['profiles']
                self.choice_data = data['choice_data']
        except FileNotFoundError:
            raise FileNotFoundError(f"Data file {file_path}_data.pkl not found")
        
        self.logger.info(f"Model and data restored from {file_path}_model.pkl and {file_path}_data.pkl")
        return self

if __name__ == "__main__":
    # Setup logging
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s [%(filename)s:%(lineno)d] %(message)s", datefmt="%d-%m-%Y %H:%M:%S")
    
    # Load data
    profiles = pd.read_excel('profiles.xlsx', index_col=0)
    choices = pd.read_excel("CBC_Data_Final_09Jun25.xlsx")
    groups = pd.read_excel("A2_9Jun25.xlsx")
    
    choices['respondent_id'] = pd.to_numeric(choices['respondent_id'], errors='coerce')
    choices['choice_set'] = pd.to_numeric(choices['choice_set'], errors='coerce')
    choices['chosen_profile'] = pd.to_numeric(choices['chosen_profile'], errors='coerce')
    
    for df in [profiles, choices, groups]:
        unnamed_cols = [col for col in df.columns if 'Unnamed' in col]
        if unnamed_cols:
            logging.getLogger("kdcm_logger").info(f"Dropping Unnamed columns from {df}: {unnamed_cols}")
            df.drop(columns=unnamed_cols, inplace=True)
    
    logging.getLogger("kdcm_logger").info(f"Profiles dtypes:\n{profiles.dtypes}")
    logging.getLogger("kdcm_logger").info(f"Choices dtypes:\n{choices.dtypes}")
    logging.getLogger("kdcm_logger").info(f"Groups dtypes:\n{groups.dtypes}")
    
    try:
        # Initialize analyzer
        analyzer = EnhancedBayesianChoiceAnalyzer(profiles, choices, groups, rename_dict={
            'SP_High performance Ics=Icu=Icw 66kA for 1sec': 'Size_Perf_High',
            'AF_Higher Electrical life from 6,000 to 7,500 operations without maintenance': 'Adv_Feat_ElecLife',
            'AF_Visible Health indication (Breaker Status, trip cause Indication - OL,SC,GF)': 'Adv_Feat_Health',
            'AF_Scalable connectivity at breaker level- Basic Modbus (Breaker Status, control, terminal temp alarm)': 'Adv_Feat_ModbusBasic',
            'AF_Current Measurement and time stamped fault records on mobile app': 'Adv_Feat_Current',
            'AF_Scalable connectivity at breaker level- Modbus Ethernet': 'Adv_Feat_ModbusEth',
            'AF_Operator Safety - Arc Flash reduction during maintenance': 'Adv_Feat_Safety'
        })
        
        # Prepare data and fit model
        choice_data = analyzer.prepare_data()
        utilities = analyzer.fit_model(use_interactions=True)
        importance = analyzer.calculate_feature_importance()
        analyzer.logger.info(f"Utilities:\n{utilities}")
        analyzer.logger.info(f"Feature Importance (SHAP):\n{importance}")
        
        # Save model
        analyzer.save_model("schneider_choice_model")
        
        # Define price scenarios
        price_scenarios = {
            'Baseline': profiles['Price'].to_dict(),
            '20% Increase': (profiles['Price'] * 1.2).to_dict(),
            '20% Decrease': (profiles['Price'] * 0.8).to_dict(),
            'Custom': {i: profiles['Price'].mean() * 1.1 for i in profiles.index}
        }
        analyzer.logger.info(f"Price Scenarios:\n{price_scenarios}")
        
        # Save scenarios to Excel
        df_scenarios = pd.DataFrame({
            'Profile': [f"Profile {i+1}" for i in range(16)],
            'Baseline': profiles['Price'].values,
            '20% Increase': (profiles['Price'] * 1.2).values,
            '20% Decrease': (profiles['Price'] * 0.8).values,
            'Custom': profiles['Price'].mean() * 1.1
        })
        analyzer.save_dataframe_to_excel(df_scenarios, "Scenarios.xlsx", sheet_name="Price_Scenarios")
        
        # Evaluate price scenarios
        shares = analyzer.evaluate_price_scenario_lp(price_scenarios, plot=True)
        analyzer.logger.info(f"Profile Shares (%):\n{shares.round(2)}")
        
        # Fit group-specific models
        analyzer.logger.info("\nFitting group-specific models...")
        for group in choice_data['group'].unique():
            analyzer.logger.info(f"\nFitting model for {group}...")
            utilities = analyzer.fit_model(group=group, use_interactions=False)
            analyzer.logger.info(f"Utilities for {group}:\n{utilities}")
            shares = analyzer.evaluate_price_scenario_lp(price_scenarios, plot=True)
            analyzer.logger.info(f"Profile Shares for {group} (%):\n{shares.round(2)}")
        
        analyzer.logger.info("Done")
    except Exception as e:
        analyzer.logger.error(f"Error occurred: {str(e)}")