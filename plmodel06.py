import pandas as pd
import numpy as np
from collections.abc import Iterable
import collections
collections.Iterable = Iterable  # Patch for pylogit Python 3.10 compatibility
import pylogit as pl
import statsmodels.api as sm
from collections import OrderedDict
import openpyxl

class DiscreteChoiceModel:
    def __init__(self, markets=None, use_pylogit=True, exclude_none=False):
        self.markets = markets if markets else ['Maharashtra', 'Gujarat', 'Punjab', 'South']
        self.models = {}
        self.coef_dfs = {}
        self.use_pylogit = use_pylogit
        self.exclude_none = exclude_none
        self.brand_share = {}
        self.value_share = {}
        self.feature_importance = {}
    
    def fit(self, df, target='Chosen', features=None):
        # Handle nan and data types
        df['CP'] = pd.to_numeric(df['CP'], errors='coerce').fillna(0.0)
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0.0)
        for col in ['FC', 'AH', 'VAS', 'Credit', 'herd_size', 'cattle_type']:
            if col in df.columns:
                df[col] = df[col].fillna('None').astype(str)
        
        # Dynamic feature selection
        mah_df = df[df['Market'] == 'Maharashtra']
        if features is None:
            features = ['Price', 'CP', 'FC', 'AH']
            if len(mah_df['VAS'].value_counts()) > 1 and mah_df['VAS'].isna().sum() / len(mah_df) < 0.2:
                features.append('VAS')
            if len(mah_df['Credit'].value_counts()) > 1 and mah_df['Credit'].isna().sum() / len(mah_df) < 0.2:
                features.append('Credit')
        
        # Exclude None choices
        if self.exclude_none:
            df = df[df['chosen_profile'] != 99].copy()
        
        # Calculate Brand Share
        chosen_df = df[df['Chosen'] == 1]
        self.brand_share = chosen_df.groupby(['Market', 'Brand']).size().reset_index(name='Choices')
        self.brand_share['Brand_Share'] = self.brand_share.groupby('Market')['Choices'].transform(lambda x: x / x.sum() if x.sum() > 0 else 0)
        
        # Calculate Value Share
        self.value_share = chosen_df.groupby(['Market', 'Brand'])['Price'].sum().reset_index(name='Total_Price')
        self.value_share['Value_Share'] = self.value_share.groupby('Market')['Total_Price'].transform(lambda x: x / x.sum() if x.sum() > 0 else 0)
        
        for market in self.markets:
            market_df = df[df['Market'] == market].copy()
            if market_df.empty:
                print(f"Warning: No data for {market}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                continue
            
            # Verify columns
            missing_cols = [col for col in features if col not in market_df.columns]
            if missing_cols:
                print(f"Warning: Missing columns {missing_cols} for {market}")
                self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                continue
            
            if self.use_pylogit:
                try:
                    # Pylogit setup
                    pylogit_df = market_df.copy()
                    pylogit_df['obs_id'] = pylogit_df['ChoiceID']
                    pylogit_df['alt_id'] = pylogit_df['Brand']
                    pylogit_df['choice'] = pylogit_df[target]
                    
                    # Specification with OrderedDict
                    spec = OrderedDict([
                        ('Price', 'all_same'),
                        ('CP', 'all_same'),
                        ('FC', 'all_different'),
                        ('AH', 'all_different')
                    ])
                    if 'VAS' in features:
                        spec['VAS'] = 'all_different'
                    if 'Credit' in features:
                        spec['Credit'] = 'all_different'
                    varnames = []
                    for var in features:
                        if var in ['Price', 'CP']:
                            varnames.append(var)
                        else:
                            unique_vals = market_df[var].unique()
                            varnames.extend([f"{var}_{val}" for val in unique_vals if val != 'None'])
                    name_dict = OrderedDict([(var, var) for var in varnames])
                    
                    # Fit pylogit
                    model = pl.create_choice_model(
                        data=pylogit_df,
                        alt_id_col='alt_id',
                        obs_id_col='obs_id',
                        choice_col='choice',
                        specification=spec,
                        model_type='MNL',
                        names=name_dict
                    )
                    init_vals = np.zeros(len(varnames))
                    bounds = [(-10, 10) for _ in varnames]
                    results = model.fit_mle(init_vals, bounds=bounds, print_res=False)
                    coef_df = pd.DataFrame({
                        'Coefficient': results.params,
                        'P-value': results.pvalues
                    }, index=varnames)
                    self.models[market] = results
                    self.coef_dfs[market] = coef_df
                    print(f"Pylogit model fitted for {market}")
                except Exception as e:
                    print(f"Pylogit failed for {market}: {e}. Switching to statsmodels.")
                    self.use_pylogit = False
            
            if not self.use_pylogit:
                try:
                    # Statsmodels setup
                    categorical_cols = [col for col in ['FC', 'AH', 'VAS', 'Credit'] if col in features and col in market_df.columns]
                    market_df_encoded = pd.get_dummies(market_df, columns=categorical_cols, drop_first=True)
                    feature_cols = [col for col in market_df_encoded.columns if col.startswith(('Price', 'CP') + tuple(categorical_cols))]
                    if not feature_cols:
                        print(f"Warning: No valid features after encoding for {market}")
                        self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                        self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
                        continue
                    X = market_df_encoded[feature_cols].astype(float).dropna()
                    y = market_df_encoded.loc[X.index, target]
                    
                    # Add constant
                    X = sm.add_constant(X)
                    
                    # Fit statsmodels
                    model = sm.Logit(y, X).fit_regularized(method='l1', alpha=0.01, disp=0, maxiter=1000)
                    coef_df = pd.DataFrame({
                        'Coefficient': model.params,
                        'P-value': model.pvalues.fillna(1.0)
                    }, index=X.columns)
                    self.models[market] = model
                    self.coef_dfs[market] = coef_df
                    print(f"Statsmodels model fitted for {market}")
                except Exception as e:
                    print(f"Statsmodels failed for {market}: {e}")
                    self.coef_dfs[market] = pd.DataFrame(columns=['Coefficient', 'P-value'])
                    self.feature_importance[market] = pd.DataFrame(columns=['Feature', 'Importance'])
            
            # Calculate feature importance
            importance = self.get_feature_importance(market)
            self.feature_importance[market] = pd.DataFrame({
                'Feature': importance.index,
                'Importance': importance.values
            })
        
        return self
    
    def get_attribute_utilities(self, market):
        coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
        return coef_df['Coefficient'] if 'Coefficient' in coef_df.columns else pd.Series(dtype=float)
    
    def get_feature_importance(self, market):
        utilities = self.get_attribute_utilities(market)
        if utilities.empty:
            return pd.Series(dtype=float)
        importance = utilities.abs() / utilities.abs().sum() if utilities.abs().sum() > 0 else utilities
        return importance
    
    def add_model_summaries_to_workbook(self, workbook):
        for market in self.markets:
            # Create or update summary sheet
            if f"{market}_summary" in workbook.sheetnames:
                sheet = workbook[f"{market}_summary"]
                # Clear existing content
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sheet = workbook.create_sheet(title=f"{market}_summary")
            
            # Write coefficients and p-values
            coef_df = self.coef_dfs.get(market, pd.DataFrame(columns=['Coefficient', 'P-value']))
            sheet['A1'] = 'Feature'
            sheet['B1'] = 'Coefficient'
            sheet['C1'] = 'P-value'
            for r, (index, row) in enumerate(coef_df.iterrows(), start=2):
                sheet[f'A{r}'] = index
                sheet[f'B{r}'] = row.get('Coefficient', float('nan'))
                sheet[f'C{r}'] = row.get('P-value', float('nan'))
            
            # Append feature importance
            importance_df = self.feature_importance.get(market, pd.DataFrame(columns=['Feature', 'Importance']))
            if not importance_df.empty:
                start_row = len(coef_df) + 4
                sheet[f'A{start_row}'] = 'Feature Importance'
                sheet[f'A{start_row + 1}'] = 'Feature'
                sheet[f'B{start_row + 1}'] = 'Importance'
                for r, (index, row) in enumerate(importance_df.iterrows(), start=start_row + 2):
                    sheet[f'A{r}'] = row['Feature']
                    sheet[f'B{r}'] = row['Importance']
            
            # Create brand share sheet
            if f"{market}_brand_share" in workbook.sheetnames:
                brand_sheet = workbook[f"{market}_brand_share"]
                for row in brand_sheet.iter_rows(min_row=1, max_row=brand_sheet.max_row, min_col=1, max_col=brand_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                brand_sheet = workbook.create_sheet(title=f"{market}_brand_share")
            brand_share_market = self.brand_share[self.brand_share['Market'] == market][['Brand', 'Brand_Share']]
            brand_sheet['A1'] = 'Brand'
            brand_sheet['B1'] = 'Brand_Share'
            for r, (index, row) in enumerate(brand_share_market.iterrows(), start=2):
                brand_sheet[f'A{r}'] = row['Brand']
                brand_sheet[f'B{r}'] = row['Brand_Share']
            
            # Create value share sheet
            if f"{market}_value_share" in workbook.sheetnames:
                value_sheet = workbook[f"{market}_value_share"]
                for row in value_sheet.iter_rows(min_row=1, max_row=value_sheet.max_row, min_col=1, max_col=value_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                value_sheet = workbook.create_sheet(title=f"{market}_value_share")
            value_share_market = self.value_share[self.value_share['Market'] == market][['Brand', 'Value_Share']]
            value_sheet['A1'] = 'Brand'
            value_sheet['B1'] = 'Value_Share'
            for r, (index, row) in enumerate(value_share_market.iterrows(), start=2):
                value_sheet[f'A{r}'] = row['Brand']
                value_sheet[f'B{r}'] = row['Value_Share']

if __name__ == "__main__":
    # Load data
    df = pd.read_csv("combined_choice_data_v2.csv")
    
    # Initialize workbook
    workbook = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    
    # Fit models and save summaries
    model = DiscreteChoiceModel(markets=['Maharashtra', 'Gujarat', 'Punjab', 'South'])
    model.fit(df)
    model.add_model_summaries_to_workbook(workbook)
    workbook.save('cargill_findings_v5.xlsx')