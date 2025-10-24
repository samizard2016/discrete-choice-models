import pandas as pd
import numpy as np
from collections.abc import Iterable
import collections
collections.Iterable = Iterable
import pylogit as pl
from statsmodels.stats.outliers_influence import variance_inflation_factor
from collections import OrderedDict
import openpyxl
import os
import logging
import uuid

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class DiscreteChoiceModel:
    def __init__(self, markets=None, use_pylogit=True, exclude_none=False, output_dir="cattle_feed_output"):
        self.markets = markets if markets else ['Maharashtra', 'Gujarat', 'Punjab', 'South']
        self.models = {}
        self.coef_dfs = {}
        self.use_pylogit = use_pylogit
        self.exclude_none = exclude_none
        self.brand_share = {}
        self.value_share = {}
        self.predicted_brand_share = {}
        self.predicted_value_share = {}
        self.feature_importance = {}
        self.price_sensitivity = {}
        self.output_dir = output_dir
        self.brand_to_id = {}
        self.id_to_brand = {}
        self.choice_id_to_num = {}
        os.makedirs(self.output_dir, exist_ok=True)

    def check_multicollinearity(self, df, predictors):
        """Check for multicollinearity using VIF."""
        X = df[predictors].fillna(0)
        vif_data = pd.DataFrame()
        vif_data["Variable"] = predictors
        vif_data["VIF"] = [variance_inflation_factor(X.values, i) for i in range(X.shape[1])]
        logging.info(f"VIF values for {predictors}:\n{vif_data}")
        high_vif = vif_data[vif_data["VIF"] > 10]["Variable"].tolist()
        if high_vif:
            logging.warning(f"High VIF detected for: {high_vif}")
        return high_vif

    def preprocess_data(self, df):
        """Preprocess the dataset for choice modeling using raw Price (INR/50kg)."""
        # Create ChoiceID
        df['ChoiceID'] = df['Respondent_ID'].astype(str) + '_' + df['choice_set'].astype(str)

        # Remove duplicate brands in choice sets
        duplicates = df.groupby(['ChoiceID', 'Brand']).size().reset_index(name='count')
        duplicates = duplicates[duplicates['count'] > 1]
        if not duplicates.empty:
            logging.info(f"Found {len(duplicates)} choice sets with duplicate brands:\n{duplicates}")
            df = df[~df['ChoiceID'].isin(duplicates['ChoiceID'].unique())].copy()
            logging.info(f"Removed {len(duplicates)} choice sets with duplicate brands")

        # Ensure positive Price/CP
        if (df['Price'] < 0).any() or (df['CP'] < 0).any():
            logging.warning("Negative values found in Price or CP. Clipping to positive range.")
            df['Price'] = df['Price'].clip(lower=1000.0)  # Minimum realistic price (INR/50kg)
            df['CP'] = df['CP'].clip(lower=0.1)  # Minimum realistic CP (proportion)
            logging.info(f"Adjusted Price range: [{df['Price'].min():.6f}, {df['Price'].max():.6f}]")
            logging.info(f"Adjusted CP range: [{df['CP'].min():.6f}, {df['CP'].max():.6f}]")

        # Remove NaN
        initial_len = len(df)
        df = df[df['Brand'].notna() & df['ChoiceID'].notna()].copy()
        if len(df) < initial_len:
            logging.info(f"Removed {initial_len - len(df)} rows with NaN in Brand or ChoiceID")

        # Ensure binary Chosen
        invalid_chosen = df[~df['Chosen'].isin([0, 1])]
        if not invalid_chosen.empty:
            logging.warning(f"Found {len(invalid_chosen)} choice sets with invalid Chosen values")
            df = df[df['Chosen'].isin([0, 1])]

        # Convert Price/CP to numeric
        df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(df['Price'].mean())
        df['CP'] = pd.to_numeric(df['CP'], errors='coerce').fillna(df['CP'].mean())

        # Log unique Brand values before creating dummies
        logging.info(f"Unique Brand values: {list(df['Brand'].unique())}")

        # Add synthetic variation if needed (realistic scale)
        if df['Price'].nunique() <= 5:
            logging.warning("Limited unique Price values. Adding synthetic variation (scale: INR/50kg).")
            df['Price'] += np.random.normal(0, 5.0, len(df))  # ±5 INR for realistic variation
            logging.info(f"Price range: [{df['Price'].min():.6f}, {df['Price'].max():.6f}]")
        if df['CP'].nunique() <= 5:
            logging.warning("Limited unique CP values. Adding synthetic variation.")
            df['CP'] += np.random.normal(0, 0.02, len(df))  # ±0.02 for CP
            logging.info(f"CP range: [{df['CP'].min():.6f}, {df['CP'].max():.6f}]")

        # Create dummy variables for categorical attributes
        df = pd.get_dummies(df, columns=['FC', 'AH', 'VAS', 'Credit'], 
                           prefix=['FC', 'AH', 'VAS', 'Credit'], 
                           drop_first=True)  # Drop FC_<1%, AH_Loss, VAS_None, Credit_None
        df = pd.get_dummies(df, columns=['Brand'], prefix='Brand', drop_first=True)  # Drop Brand_Cargill

        # Map ChoiceID and alt_id
        unique_choice_ids = df['ChoiceID'].unique()
        self.choice_id_to_num = {cid: idx + 1 for idx, cid in enumerate(unique_choice_ids)}
        df['obs_id'] = df['ChoiceID'].map(self.choice_id_to_num)
        df['alt_id'] = df.groupby('ChoiceID').cumcount() + 1
        logging.info(f"ChoiceID to numeric mapping (first 10): {dict(list(self.choice_id_to_num.items())[:10])}")
        logging.info(f"Unique ChoiceID values (first 10): {list(unique_choice_ids[:10])}")

        return df

    def fit(self, df, target='Chosen', features=None):
        """Fit choice models for each market."""
        df = self.preprocess_data(df)
        if features is None:
            features = [col for col in df.columns if col.startswith(('Price', 'CP', 'FC_', 'AH_', 'VAS_', 'Credit_', 'Brand_'))]
        
        # Check multicollinearity
        high_vif = self.check_multicollinearity(df, [f for f in features if f != 'Brand_Cargill'])
        features = [f for f in features if f not in high_vif]

        for market in self.markets:
            logging.info(f"Fitting model for {market}")
            market_df = df[df['Market'] == market].copy()
            
            # Ensure at least 3 alternatives per choice set
            choice_counts = market_df.groupby('obs_id').size()
            valid_obs = choice_counts[choice_counts >= 3].index
            market_df = market_df[market_df['obs_id'].isin(valid_obs)]
            logging.info(f"Choice set sizes:\n{market_df.groupby('obs_id').size().value_counts()}")

            # Create specification
            spec = OrderedDict()
            names = OrderedDict()
            for f in features:
                spec[f] = 'all_same'
                names[f] = f

            try:
                model = pl.create_choice_model(
                    data=market_df,
                    alt_id_col='alt_id',
                    obs_id_col='obs_id',
                    choice_col=target,
                    specification=spec,
                    model_type='MNL',
                    names=names
                )
                results = model.fit_mnl(method='bfgs', maxiter=2000)
                logging.info(f"{market} Log-likelihood: {results.log_likelihood}")
                logging.info(f"{market} Coefficients:\n{results.params}")
                self.models[market] = results
                self.coef_dfs[market] = pd.DataFrame({
                    'Coefficient': results.params,
                    'Std_Error': results.std_err,
                    'P_Value': results.p_values
                })

                # Check price coefficient
                if 'Price' in results.params and results.params['Price'] > 0:
                    logging.warning(f"Positive price coefficient ({results.params['Price']}) detected in {market}. Check Price data.")

                # Compute shares
                # Note: Brand column is gone, so we reconstruct it from dummy columns
                brand_columns = [col for col in df.columns if col.startswith('Brand_')]
                brand_names = [col.replace('Brand_', '') for col in brand_columns] + ['Cargill']  # Add reference brand
                market_df['Brand'] = pd.Series(np.zeros(len(market_df)), dtype=str)
                for i, row in market_df.iterrows():
                    for brand, col in zip(brand_names, brand_columns + ['Cargill']):
                        if col == 'Cargill' and all(row[brand_columns] == 0):
                            market_df.at[i, 'Brand'] = 'Cargill'
                        elif col != 'Cargill' and row[col] == 1:
                            market_df.at[i, 'Brand'] = brand

                self.brand_share[market] = market_df.groupby('Brand')[target].mean().reset_index()
                self.brand_share[market].columns = ['Brand', 'Actual_Brand_Share']
                self.value_share[market] = market_df.groupby('Brand').apply(
                    lambda x: (x['Price'] * x[target]).sum() / x['Price'].sum() if x['Price'].sum() > 0 else 0
                ).reset_index()
                self.value_share[market].columns = ['Brand', 'Actual_Value_Share']
            except Exception as e:
                logging.error(f"MNL failed for {market}: {str(e)}")
                self.models[market] = None
                self.coef_dfs[market] = pd.DataFrame()
                
                # Try binary logit (Cargill vs. others)
                logging.warning(f"Trying binary logit for {market}")
                binary_df = market_df.copy()
                binary_df[target] = (binary_df['Brand'] == 'Cargill').astype(int)
                try:
                    binary_model = pl.create_choice_model(
                        data=binary_df,
                        alt_id_col='alt_id',
                        obs_id_col='obs_id',
                        choice_col=target,
                        specification={'Price': 'all_same', 'CP': 'all_same'},
                        model_type='MNL',
                        names={'Price': 'Price', 'CP': 'CP'}
                    )
                    binary_results = binary_model.fit_mnl(method='bfgs', maxiter=2000)
                    logging.info(f"{market} Binary Logit Coefficients:\n{binary_results.params}")
                    self.models[market] = binary_results
                    self.coef_dfs[market] = pd.DataFrame({
                        'Coefficient': binary_results.params,
                        'Std_Error': binary_results.std_err,
                        'P_Value': binary_results.p_values
                    })
                except Exception as e:
                    logging.error(f"Binary logit failed for {market}: {str(e)}")
                    self.models[market] = None
                    self.coef_dfs[market] = pd.DataFrame()

    def add_model_summaries_to_workbook(self, workbook):
        """Write model results to Excel."""
        for market in self.markets:
            coef_df = self.coef_dfs.get(market, pd.DataFrame())
            logging.info(f"Writing {len(coef_df)} coefficients for {market}")
            
            if f"{market}_coefficients" in workbook.sheetnames:
                coef_sheet = workbook[f"{market}_coefficients"]
                for row in coef_sheet.iter_rows(min_row=1, max_row=coef_sheet.max_row, min_col=1, max_col=coef_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                coef_sheet = workbook.create_sheet(title=f"{market}_coefficients")
            
            if not coef_df.empty:
                coef_sheet['A1'] = 'Parameter'
                coef_sheet['B1'] = 'Coefficient'
                coef_sheet['C1'] = 'Std_Error'
                coef_sheet['D1'] = 'P_Value'
                for r, (index, row) in enumerate(coef_df.iterrows(), start=2):
                    coef_sheet[f'A{r}'] = index
                    coef_sheet[f'B{r}'] = row['Coefficient']
                    coef_sheet[f'C{r}'] = row['Std_Error']
                    coef_sheet[f'D{r}'] = row['P_Value']
            
            for col in ['A', 'B', 'C', 'D']:
                column_letter = col
                max_length = 0
                column = coef_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                coef_sheet.column_dimensions[column_letter].width = adjusted_width

            # Write brand shares
            if f"{market}_brand_share" in workbook.sheetnames:
                brand_sheet = workbook[f"{market}_brand_share"]
                for row in brand_sheet.iter_rows(min_row=1, max_row=brand_sheet.max_row, min_col=1, max_col=brand_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                brand_sheet = workbook.create_sheet(title=f"{market}_brand_share")
            
            brand_share_market = self.brand_share.get(market, pd.DataFrame(columns=['Brand', 'Actual_Brand_Share']))
            logging.info(f"Writing {len(brand_share_market)} brand share rows for {market}")
            brand_sheet['A1'] = 'Brand'
            brand_sheet['B1'] = 'Actual_Brand_Share'
            for r, (index, row) in enumerate(brand_share_market.iterrows(), start=2):
                brand_sheet[f'A{r}'] = row['Brand']
                brand_sheet[f'B{r}'] = row['Actual_Brand_Share']
            
            for col in ['A', 'B']:
                column_letter = col
                max_length = 0
                column = brand_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                brand_sheet.column_dimensions[column_letter].width = adjusted_width

            # Write value shares
            if f"{market}_value_share" in workbook.sheetnames:
                value_sheet = workbook[f"{market}_value_share"]
                for row in value_sheet.iter_rows(min_row=1, max_row=value_sheet.max_row, min_col=1, max_col=value_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                value_sheet = workbook.create_sheet(title=f"{market}_value_share")
            
            value_share_market = self.value_share.get(market, pd.DataFrame(columns=['Brand', 'Actual_Value_Share']))
            logging.info(f"Writing {len(value_share_market)} value share rows for {market}")
            value_sheet['A1'] = 'Brand'
            value_sheet['B1'] = 'Actual_Value_Share'
            for r, (index, row) in enumerate(value_share_market.iterrows(), start=2):
                value_sheet[f'A{r}'] = row['Brand']
                value_sheet[f'B{r}'] = row['Actual_Value_Share']
            
            for col in ['A', 'B']:
                column_letter = col
                max_length = 0
                column = value_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                value_sheet.column_dimensions[column_letter].width = adjusted_width

            # Write price sensitivity (placeholder)
            if f"{market}_price_sensitivity" in workbook.sheetnames:
                sensitivity_sheet = workbook[f"{market}_price_sensitivity"]
                for row in sensitivity_sheet.iter_rows(min_row=1, max_row=sensitivity_sheet.max_row, min_col=1, max_col=sensitivity_sheet.max_column):
                    for cell in row:
                        cell.value = None
            else:
                sensitivity_sheet = workbook.create_sheet(title=f"{market}_price_sensitivity")
            sensitivity_df = self.price_sensitivity.get(market, pd.DataFrame(columns=['Price_Change_Percent', 'Brand_Share', 'Value_Share']))
            logging.info(f"Writing {len(sensitivity_df)} price sensitivity rows for {market}")
            sensitivity_sheet['A1'] = 'Price_Change_Percent'
            sensitivity_sheet['B1'] = 'Brand_Share'
            sensitivity_sheet['C1'] = 'Value_Share'
            for r, (index, row) in enumerate(sensitivity_df.iterrows(), start=2):
                sensitivity_sheet[f'A{r}'] = row['Price_Change_Percent']
                sensitivity_sheet[f'B{r}'] = row['Brand_Share']
                sensitivity_sheet[f'C{r}'] = row['Value_Share']
            sensitivity_sheet['A{0}'.format(len(sensitivity_df) + 4)] = 'Note: Line chart can be created in Excel using the data above.'
            
            for col in ['A', 'B', 'C']:
                column_letter = col
                max_length = 0
                column = sensitivity_sheet[column_letter]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sensitivity_sheet.column_dimensions[column_letter].width = adjusted_width

if __name__ == "__main__":
    df = pd.read_csv("cattle_feed_data_final.csv")
    workbook = openpyxl.Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    model = DiscreteChoiceModel(markets=['Maharashtra', 'Gujarat', 'Punjab', 'South'], use_pylogit=True)
    model.fit(df)
    model.add_model_summaries_to_workbook(workbook)
    workbook.save('cargill_findings_v15.xlsx')