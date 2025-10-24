import sys
import os
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QToolBar, QStatusBar, QTabWidget, QWidget,
    QVBoxLayout, QLabel, QPushButton, QRadioButton, QTextEdit, QGridLayout,
    QFileDialog, QLineEdit, QDialog, QTableView, QAbstractItemView, QMessageBox,
    QMenu, QInputDialog, QRadioButton, QButtonGroup
)
from PySide6.QtGui import QIcon, QAction, QPixmap
from PySide6.QtCore import Qt, QSize, Signal
import logging
from reg import Registry
from datetime import date, datetime
import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.line import LineProperties
from discrete_choice_analyzer import DiscreteChoiceAnalyzer
from datagrid import PandasModel
import multiprocessing

class QCLineEdit(QLineEdit):
    clicked = Signal()
    def mousePressEvent(self, event):
        self.clicked.emit()
        QLineEdit.mousePressEvent(self, event)

class DCMApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Kantar Discrete Choice Model Simulator")
        self.setGeometry(100, 100, 1200, 800)
        self.setMinimumSize(800, 600)
        self.setWindowIcon(QIcon("model.jpeg"))

        # Define DCM and BDCM descriptions
        self.dcm_text = (
            "<p style='font-size:14px;'>"
            "The <b>Discrete Choice Model</b> is a tool used to analyze how people make choices "
            "among a set of options, such as selecting a product based on its features and price.<br>"
            "It simulates decision-making by creating choice sets where individuals pick their preferred "
            "option from a group of profiles, each defined by attributes like price, performance,<br> or advanced "
            "features. The model estimates how much each attribute influences the choice, revealing what "
            "factors matter most to decision-makers.<br>"
            "It can predict how changes in price affect preferences and calculate the relative importance"
            "of each attribute, helping businesses understand customer priorities and optimize their offerings.<br>"
            "Results are visualized through charts showing preferences, feature importance, and price sensitivity."
            "</p>"
        )
        self.bdcm_text = (
            "<p style='font-size:14px;'>"
            "The <b>Bayesian Discrete Choice Model</b> is a sophisticated statistical framework that "
            "analyzes <i>consumer preferences</i> by modeling choices under uncertainty.<br>"
            "Leveraging Bayesian inference, it provides <font color='blue'>precise estimates</font> of "
            "utility parameters and <b>Willingness to Pay</b>, with <i>Highest Density Intervals</i> "
            "to quantify uncertainty, enabling robust decision-making for product pricing and features."
            "</p>"
        )

        # Initialize dashboard widget and layout
        self.dashboard_widget = QWidget()
        self.dashboard_layout = QGridLayout()
        self.dashboard_layout.setSpacing(10)
        self.dashboard_layout.setContentsMargins(10, 10, 10, 10)
        self.dashboard_widget.setLayout(self.dashboard_layout)

        # Configure logging
        log_file = "kdcm_logger.log"
        self.logger = logging.getLogger("kdcm_logger")
        self.logger.setLevel(logging.DEBUG)
        self.logger.propagate = False

        file_handler = logging.FileHandler(log_file, mode='a', delay=False)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(logging.Formatter(
            "%(asctime)s %(levelname)-8s [%(filename)s:%(lineno)d] %(message)s",
            datefmt="%d-%m-%Y %H:%M:%S"
        ))

        if not self.logger.handlers:
            self.logger.addHandler(file_handler)

        self.logger.info("A new session of KDCM Simulator started")
        for handler in self.logger.handlers:
            if isinstance(handler, logging.FileHandler):
                handler.flush()

        # Initialize UI components
        self.table = None
        self.model = None
        self.btn_simulate = None
        self.grid_layout = None
        self.radio_DCM = QRadioButton("Discrete Choice Model")
        self.radio_BDCM = QRadioButton("Bayesian Discrete Choice Model")
        self.prof_xlsx = QCLineEdit()
        self.text_model = QLabel()
        self.text_model.setText(self.dcm_text)
        self.text_model.setWordWrap(True)
        self.text_model.setStyleSheet("""
            QLabel {
                font-size: 14px;
                margin: 10px;
                background: transparent;
            }
        """)

        # Create the ribbon bar
        self.ribbon = QToolBar("Ribbon Bar")
        self.ribbon.setMovable(False)
        self.ribbon.setToolButtonStyle(Qt.ToolButtonTextUnderIcon)
        self.addToolBar(Qt.TopToolBarArea, self.ribbon)

        # Create the status bar
        self.statusBar = QStatusBar(self)
        self.statusBar.setObjectName("statusBar")
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage('Status:')
        self.statusBar.setStyleSheet("color: black;")
        self.source = False
        self._status = ""

        # Create the tab widget
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        # Define button and tab data
        self.tabs = [
            {"name": "Home", "icon": "home.jpeg", "tooltip": "Go to Home"},
            {"name": "Models", "icon": "model.jpeg", "tooltip": "View Models"},
            {"name": "Dashboard", "icon": "dashboard.jpeg", "tooltip": "Open Dashboard"},
            {"name": "Simulator", "icon": "simulatorx.jpeg", "tooltip": "Simulation - predict shares for new profiles"}
        ]

        # Create tabs and populate with views
        self.tab_indices = {}
        for i, tab in enumerate(self.tabs):
            view_method = getattr(self, f"{tab['name'].lower()}_view")
            self.tab_widget.addTab(view_method(), tab["name"])
            self.tab_indices[tab["name"]] = i
        self.tab_widget.tabBar().setVisible(False)

        # Add buttons to the ribbon
        for tab in self.tabs:
            action = QAction(QIcon(tab["icon"]), tab["name"], self)
            action.setToolTip(tab["tooltip"])
            self.ribbon.addAction(action)

        # Style the ribbon bar
        self.ribbon.setStyleSheet("""
            QToolBar {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #f0f0f0, stop:1 #d4d4d4);
                border-bottom: 1px solid #a0a0a0;
                spacing: 2px;
                padding: 2px;
            }
            QToolButton {
                font-size: 12px;
                padding: 2px;
                margin: 1px;
                min-width: 64px;
                min-height: 48px;
            }
            QToolButton:hover {
                background-color: #e0e0e0;
                border: 1px solid #a0a0a0;
                border-radius: 4px;
            }
        """)
        self.ribbon.setIconSize(QSize(48, 48))

        # Registration check and event setup
        if self.check_reg():
            self.update_status("Ready")
            self.set_events()
            self.logger.info("Registration check has passed")
        else:
            self.update_status(f"Registration failed. If the registration wasn't done yet, you need to send the registration file({self.reg_file}) to Kantar")
            self.logger.error("Registration check has failed")

        self.update_dashboard()
        self.model_for = "All"

    def check_reg(self):
        self.reg_file = f"{os.environ['COMPUTERNAME']}.bin"
        if os.path.isfile(self.reg_file):
            self.reg = Registry.restore(self.reg_file)
            self.logger.info(self.reg.d_mac)
            if self.reg.check():
                return True
            else:
                try:
                    self.reg = Registry(**{"expiry_date": str(date.today()), "expired": "yes"})
                    self.reg.update()
                except Exception as err:
                    self.logger.error(f"Couldn't create the reg file {self.reg_file}: {err}")
        else:
            self.reg = Registry(**{"expiry_date": str(date.today()), "expired": "no"})
        return False

    def update_status(self, message):
        self.statusBar.showMessage(f"Status: {message}")
        self.statusBar.repaint()

    @property
    def Status(self):
        return self._status

    @Status.setter
    def Status(self, current_status):
        self._status = current_status
        self.update_status(self._status)

    def set_events(self):
        self.logger.info("Setting events for ribbon & radio buttons")
        self.ribbon.clear()
        for tab in self.tabs:
            action = QAction(QIcon(tab["icon"]), tab["name"], self)
            action.setToolTip(tab["tooltip"])
            action.triggered.connect(lambda checked, name=tab["name"]: self.button_clicked(name))
            self.ribbon.addAction(action)

        self.radio_DCM.clicked.connect(self.set_model_text)
        self.radio_BDCM.clicked.connect(self.set_model_text)
        self.radio_DCM.toggled.connect(self.update_dashboard)
        self.prof_xlsx.clicked.connect(self.get_profiles_to_evaluate)

    def get_base_dir(self):
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        return base_dir

    def run_simulate(self):
        self.update_status("You are running simulation. Wait for a moment...")
        self.logger.info("Attempting to run the simulator...")
        data_file = self.prof_xlsx.text()
        self.logger.info(f"Reading data file: {data_file}")

        # Define group name mapping
        d_model_for = {
            "All": "All",
            "Panel Builders": "Panel builder",
            "Non-Panel Builders": "Others"
        }

        try:
            # Get base directory
            base_dir = self.get_base_dir()
            self.logger.info(f"Base directory: {base_dir}")

            # Verify file existence
            for file in ['profiles.xlsx', 'data_final.xlsx', 'groups_final.xlsx']:
                file_path = os.path.join(base_dir, file)
                if not os.path.exists(file_path):
                    self.logger.error(f"File not found: {file_path}")
                    self.update_status(f"Error: File {file_path} not found")
                    raise FileNotFoundError(f"File {file_path} not found")

            # Load data
            profiles = pd.read_excel(os.path.join(base_dir, 'profiles.xlsx'), index_col=0)
            choices = pd.read_excel(os.path.join(base_dir, 'data_final.xlsx'))
            groups = pd.read_excel(os.path.join(base_dir, 'groups_final.xlsx'))
            self.logger.info("Data files read successfully")

            # Convert data types
            choices['respondent_id'] = pd.to_numeric(choices['respondent_id'], errors='coerce')
            choices['choice_set'] = pd.to_numeric(choices['choice_set'], errors='coerce')
            choices['chosen_profile'] = pd.to_numeric(choices['chosen_profile'], errors='coerce')
            self.logger.info("Data type conversion completed")

            # Drop unnamed columns
            for df in [profiles, choices, groups]:
                unnamed_cols = [col for col in df.columns if 'Unnamed' in col]
                if unnamed_cols:
                    self.logger.warning(f"Dropping Unnamed columns from DataFrame: {unnamed_cols}")
                    df.drop(columns=unnamed_cols, inplace=True)

            # Validate data
            if choices[['respondent_id', 'choice_set', 'chosen_profile']].isna().any().any():
                self.logger.error("Missing values in choices DataFrame")
                raise ValueError("Missing values in choices DataFrame")
            if not all(choices['respondent_id'].isin(groups['respondent_id'])):
                missing_ids = choices[~choices['respondent_id'].isin(groups['respondent_id'])]['respondent_id'].unique()
                self.logger.error(f"respondent_id values in choices not found in groups: {missing_ids}")
                raise ValueError(f"respondent_id values in choices not found in groups: {missing_ids}")

            # Initialize analyzer
            analyzer = DiscreteChoiceAnalyzer(profiles, choices, groups)
            self.logger.info("Analyzer initialized")

            # Load price scenario
            if not os.path.exists(data_file):
                self.logger.error(f"Price scenario file not found: {data_file}")
                self.update_status(f"Error: Price scenario file {data_file} not found")
                raise FileNotFoundError(f"Price scenario file {data_file} not found")

            df = pd.read_excel(data_file)
            d_dict = {}
            for col in df.columns[1:]:
                d_dict[col] = df[col].to_dict()
            self.logger.info(f"Price scenarios prepared with columns: {list(d_dict.keys())}")

            try:
                group_name = d_model_for.get(self.model_for, None)
                if group_name is None:
                    self.logger.error(f"Invalid model_for value: {self.model_for}")
                    raise ValueError(f"Invalid group selected: {self.model_for}")

                if group_name == "All":
                    analyzer = analyzer.restore_model("schneider_choice_model")
                    self.logger.info("Model restored successfully")
                    shares = analyzer.evaluate_price_scenario_lp(d_dict, plot=False)
                    value_shares = analyzer.estimate_value_share(d_dict, plot=False)
                else:
                    choice_data = analyzer.prepare_data()
                    self.logger.info(f"Choice data prepared for {self.model_for}, rows: {len(choice_data)}")
                    if choice_data.empty:
                        self.logger.error(f"Choice data is empty for {self.model_for}")
                        raise ValueError(f"No valid choice data for {self.model_for}")
                    utilities = analyzer.fit_model(group=group_name, use_interactions=False)
                    self.logger.info(f"Model fitted for {self.model_for} (group: {group_name})")
                    importance = analyzer.calculate_feature_importance()
                    self.logger.info(f"Feature importance calculated for {self.model_for}")
                    shares = analyzer.evaluate_price_scenario_lp(d_dict, plot=True, file=f"profile_shares_{self.model_for.replace(' ', '_')}")
                    value_shares = analyzer.estimate_value_share(d_dict, plot=True, file=f"value_shares_{self.model_for.replace(' ', '_')}")

                self.logger.info(f"Shares predicted for {self.model_for}: {shares.head().to_string()}")
                self.logger.info(f"Value shares predicted for {self.model_for}: {value_shares.head().to_string()}")

                # Prepare output file
                shares['Profile'] = df['Profile']
                shares = shares[list(df.columns)]
                value_shares['Profile'] = df['Profile']
                value_shares = value_shares[list(df.columns)]
                out_file = f"simulated_shares_for_{self.model_for.replace(' ', '_')}_{os.path.basename(data_file)}".lower()

                # Save both profile shares and value shares to the same Excel file
                with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                    shares.to_excel(writer, sheet_name='Profile Share', index=False)
                    value_shares.to_excel(writer, sheet_name='Value Share', index=False)

                # Create charts for both profile shares and value shares
                self.create_profile_share_chart(out_file, data_sheet='Profile Share')
                self.create_value_share_chart(out_file, data_sheet='Value Share')

                msg = f"Simulated profile and value shares have been saved to {out_file}"
                self.logger.info(msg)
                self.update_status(msg)
            except Exception as e:
                self.logger.error(f"Error occurred in simulation for {self.model_for}: {str(e)}", exc_info=True)
                self.update_status(f"Simulation failed: {str(e)}")
                raise
        except Exception as err:
            self.logger.error(f"Error found in reading data files: {str(err)}", exc_info=True)
            self.update_status(f"Error reading data files: {str(err)}")
            raise

    def set_model_text(self):
        self.logger.info(f"set_model_text called, DCM checked: {self.radio_DCM.isChecked()}")
        if self.radio_DCM.isChecked():
            self.text_model.setText(self.dcm_text)
            self.update_status(f"{self.radio_DCM.text()} has been selected")
        elif self.radio_BDCM.isChecked():
            self.text_model.setText(self.bdcm_text)
            self.update_status(f"{self.radio_BDCM.text()} has been selected")

    def button_clicked(self, tab_name):
        self.update_status(f"Moved to {tab_name}")
        self.tab_widget.setCurrentIndex(self.tab_indices[tab_name])

    def home_view(self):
        widget = QWidget()
        layout = QVBoxLayout()
        lab_title = QLabel()
        lab_title.setText("Welcome to <b>Kantar Discrete Choice Modeling</b>")
        lab_title.setStyleSheet("QLabel { font-size: 20px; }")
        layout.addWidget(lab_title)
        description = (
            "<p style='font-size:14px;'>"
            "This application delivers a <b>robust discrete choice modeling framework</b> designed to evaluate "
            "<i>consumer preferences</i> for circuit breaker products, incorporating diverse features and price points to <b>Schneider</b>.<br>"
            "The integrated <b>Simulator</b>, equipped with a trained model, enables "
            "<font color='blue'>precise estimates of market shares</font> for product profiles based on new pricing scenarios.<br>"
            "The application has meticulously been designed to:"
            "<ul>"
            "<li>Give advanced insights from a <i>Bayesian Discrete Choice Model</i></li>"
            "<li>Include distribution of model parameters and <font color='darkred'>Willingness to Pay (WTP) estimates</font></li>"
            "<li>Give <font color='darkred'>Feature Utilities with 95% Credible Intervals</font> (High Density Intervals)</li>"
            "<li>Provide a deeper understanding of underlying model relationships</li>"            
            "</ul>"
            "</p>"
        )
        lab_desc = QLabel()
        lab_desc.setText(description)
        lab_desc.setStyleSheet("QLabel { font-size: 14px; margin: 10px; }")
        lab_desc.setWordWrap(True)
        layout.addWidget(lab_desc)
        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def models_view(self):
        self.logger.info("Creating models_view")
        widget = QWidget()
        layout = QVBoxLayout()
        lab_title = QLabel()
        lab_title.setText("<b>Kantar Discrete Choice Models</b>")
        lab_title.setStyleSheet("QLabel { font-size: 20px; margin: 5px; }")
        layout.addWidget(lab_title)

        self.radio_DCM.setStyleSheet("QRadioButton { font-size: 14px; color: black; margin-left: 15px; }")
        self.radio_BDCM.setStyleSheet("QRadioButton { font-size: 14px; color: black; margin-left: 15px; }")
        self.radio_DCM.setChecked(True)
        self.text_model.setText(self.dcm_text)

        layout.addWidget(self.radio_DCM)
        layout.addWidget(self.radio_BDCM)
        layout.addWidget(self.text_model)
        layout.addStretch()
        widget.setLayout(layout)
        return widget

    def clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()
            else:
                nested_layout = item.layout()
                if nested_layout:
                    self.clear_layout(nested_layout)

    def update_dashboard(self):
        self.clear_layout(self.dashboard_layout)
        if self.radio_DCM.isChecked():
            self.logger.info("DCM checked")
            images = ['utilities.png', 'feature_importance_dcm.png', 'price_elasticity.png', 'profile_shares_line.png', 'value_shares_line.png']
            d_int = {
                "utilities.png": "Each bar represents the coefficient of the feature in the logit model, indicating its impact on choice probability.\n"
                                 "Positive coefficient increases the likelihood of the profile being chosen, negative coefficient decreases it.",
                "feature_importance_dcm.png": "Each value represents the feature’s average contribution to the model’s predictions, relative to other features.\n"
                                             "For example, Price: 0.05 means Price contributes 5% of the total impact on choice predictions.",
                "price_elasticity.png": "Price Elasticity measures the percentage change in a profile’s choice probability for a 1% change in its price.\n"
                                       "For example, an elasticity of -2 means a 1% price increase leads to a 2% decrease in choice probability.",
                "profile_shares_line.png": "Scenario Comparison:\n"
                                          "- Baseline: Reflects the current market shares based on original prices.\n"
                                          "            Profiles with higher shares are currently preferred.\n"
                                          "- 10% Increase: If a profile’s share decreases compared to Baseline, it’s price-sensitive.\n"
                                          "            Larger drops indicate higher sensitivity.\n"
                                          "- 10% Decrease: If a profile’s share increases, a price reduction boosts its appeal.\n"
                                          "            Larger increases suggest higher price elasticity.\n"
                                          "- Custom: Since all profiles have the same price (mean price), differences in shares\n"
                                          "            reflect preferences for non-price attributes (`SP`, `AF`). Profiles with higher shares in\n"
                                          "            this scenario are preferred for their features, not price.",
                "value_shares_line.png": "Value Share Comparison:\n"
                                        "- Baseline: Shows revenue shares based on original prices and choice probabilities.\n"
                                        "- 10% Increase: Higher prices may increase value share if demand is inelastic.\n"
                                        "- 10% Decrease: Lower prices may reduce value share unless demand is highly elastic.\n"
                                        "- Custom: Equal pricing highlights the impact of non-price attributes on revenue."
            }
        else:
            self.logger.info("BDCM checked")
            images = ['utilities_with_uncertainty_bdcm.png', 'wtp_analysis_enhanced_bdcm.png']
            d_int = {
                "utilities_with_uncertainty_bdcm.png":
                    "Feature Utilities: Features with large positive utilities (e.g., Adv_Feat_ElecLife) are strong drivers of choice, while\n"
                    "                   those with near-zero or negative utilities have little or negative impact.\n"
                    "Uncertainty: Narrow HDIs indicate robust estimates, while wider intervals suggest uncertainty, possibly\n"
                    "                   due to variability in preferences or limited data.\n"
                    "Relative Importance: Features with utilities significantly different from zero and large in magnitude\n"
                    "                   are key differentiators in consumer choice.\n"
                    "Baseline Reference: The dropped baseline levels (sp_baseline, af_baseline) have an implicit utility of zero,\n"
                    "                   so other feature utilities are relative to these baselines.",
                "wtp_analysis_enhanced_bdcm.png":
                    "High WTP Features: Features with large positive mean WTP and narrow HDIs (e.g., Adv_Feat_ElecLife or Adv_Feat_Current) are highly valued by consumers,\n"
                    "                   indicating strong market potential. These features could justify premium pricing.\n"
                    "Low or Negative WTP: Features with WTP near zero or negative (with HDIs crossing zero) are less valued or potentially\n"
                    "                   detrimental compared to the baseline, suggesting limited consumer interest.\n"
                    "Uncertainty: Wide HDIs suggest uncertainty in consumer preferences, possibly due to heterogeneous preferences or insufficient data.\n"
                    "                   Narrow HDIs indicate more reliable estimates.\n"
                    "Practical Implications: Features with WTP exceeding the median price are likely key drivers of purchase decisions, while those below may not\n"
                    "                   justify their cost of implementation."
            }

        base_dir = self.get_base_dir()
        image_dir = os.path.join(base_dir, 'images')

        self.dashboard_layout.setSpacing(0)
        self.dashboard_layout.setContentsMargins(0, 0, 0, 0)

        for i, img in enumerate(images):
            image_label = QLabel()
            image_path = os.path.join(image_dir, img)
            self.logger.info(f"Attempting to load image: {image_path}")
            pixmap = QPixmap(image_path)
            if pixmap.isNull():
                self.logger.warning(f"Failed to load image: {image_path}")
                placeholder_path = os.path.join(image_dir, 'placeholder.png')
                pixmap = QPixmap(placeholder_path)
                if pixmap.isNull():
                    self.logger.warning(f"Failed to load placeholder: {placeholder_path}")
                    image_label.setText(f"Image not found:\n{img}")
                    image_label.setStyleSheet("QLabel { color: red; font-size: 12px; }")
                    image_label.setAlignment(Qt.AlignCenter)
                    image_label.setMinimumSize(QSize(360, 240))
                    self.dashboard_layout.addWidget(image_label, i // 2, i % 2)
                    continue
            target_size = QSize(360, 240)
            pixmap = pixmap.scaled(target_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            image_label.setPixmap(pixmap)
            image_label.setAlignment(Qt.AlignCenter)
            image_label.setMinimumSize(target_size)
            image_label.setMaximumSize(target_size)
            image_label.setScaledContents(False)
            image_label.setToolTip(d_int[img] if img in d_int else "Interpretation not found")
            image_label.setContentsMargins(0, 0, 0, 0)
            image_label.setObjectName(f"image_{i}")
            image_label.mousePressEvent = lambda event, path=image_path: self.show_full_image(path)
            self.dashboard_layout.addWidget(image_label, i // 2, i % 2)

    def show_full_image(self, image_path):
        dialog = QDialog(self)
        dialog.setWindowTitle(self.windowTitle())
        dialog.setMinimumSize(800, 600)
        layout = QVBoxLayout()
        image_label = QLabel()
        pixmap = QPixmap(image_path)
        if not pixmap.isNull():
            pixmap = pixmap.scaled(800, 600, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            image_label.setPixmap(pixmap)
            image_label.setAlignment(Qt.AlignCenter)
        else:
            image_label.setText("Unable to load image")
            image_label.setStyleSheet("QLabel { color: red; font-size: 14px; }")
        layout.addWidget(image_label)
        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.close)
        layout.addWidget(close_button)
        dialog.setLayout(layout)
        dialog.exec()

    def dashboard_view(self):
        return self.dashboard_widget

    def simulator_view(self):
        self.grid_layout = QGridLayout()
        self.grid_layout.setContentsMargins(0, 0, 10, 0)
        self.grid_layout.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        widget = QWidget()
        lab_title = QLabel("Welcome to Kantar Discrete Choice Model Simulator")
        lab_title.setStyleSheet("QLabel { font-size: 20px; }")
        self.grid_layout.addWidget(lab_title, 0, 0, Qt.AlignTop | Qt.AlignLeft)
        lab_desc = QLabel("The simulator helps you predict profile shares for new prices")
        lab_desc.setStyleSheet("QLabel { font-size: 14px; }")
        self.grid_layout.addWidget(lab_desc, 1, 0, Qt.AlignTop | Qt.AlignLeft)

        lab_prof = QLabel("Profile (*.xlsx)")
        lab_prof.setStyleSheet("font-size: 10pt; color: black; margin: 5px 0px; width: 200px;")
        self.prof_xlsx.setText("Select profile file")
        self.prof_xlsx.setStyleSheet("font-size: 7pt; background-color: #bac8e0; color: black; width: 400px; height: 35px;")
        self.btn_simulate = QPushButton("Simulate")
        self.btn_simulate.setStyleSheet("font-size: 10pt; height: 30px; width: 150px; background-color: black; color: white; margin: 0px;")
        self.btn_simulate.setEnabled(False)
        self.btn_simulate.clicked.connect(self.run_simulate)
        self.grid_layout.addWidget(lab_prof, 2, 0, Qt.AlignTop | Qt.AlignRight)
        self.grid_layout.addWidget(self.prof_xlsx, 2, 1, 1, 3, Qt.AlignTop | Qt.AlignLeft)
        self.grid_layout.addWidget(self.btn_simulate, 3, 1, Qt.AlignTop | Qt.AlignLeft)

        # Create button group for radio buttons
        self.button_group_model = QButtonGroup(self)
        options = ['All', 'Panel Builders', 'Non-Panel Builders']
        for i, option in enumerate(options):
            radio_button = QRadioButton(option)
            self.button_group_model.addButton(radio_button)
            self.grid_layout.addWidget(radio_button, 3, 2+i, Qt.AlignTop | Qt.AlignLeft)

        self.button_group_model.buttons()[0].setChecked(True)
        self.button_group_model.buttonClicked.connect(self.on_radio_button_clicked)
        widget.setLayout(self.grid_layout)
        return widget

    def on_radio_button_clicked(self, button):
        self.model_for = button.text()

    def get_file(self, title, type):
        cwd = os.getcwd()
        openFileName = QFileDialog.getOpenFileName(self, title, cwd, ("XLSX (*.XLSX)"))
        if openFileName != ('', ''):
            file = openFileName[0]
            self.logger.info(f"{file} has been selected")
            return file
        else:
            return None

    def get_profiles_to_evaluate(self):
        file = self.get_file("Select a price profile file to evaluate on", type='xlsx')
        if file is not None:
            self.prof_xlsx.setText(file)
            self.btn_simulate.setEnabled(True)
            try:
                input_file = self.prof_xlsx.text()
                msg = f"{input_file} is selected"
                self.logger.info(msg)
                self.update_status(msg)
                self.get_dataview()
                self.add_actions_dataview()
            except Exception as err:
                msg = f"Problem in selecting profile file - {err}"
                self.logger.error(msg)
                self.update_status(msg)

    def get_dataview(self):
        file_name = self.prof_xlsx.text()
        if file_name:
            try:
                df = pd.read_excel(file_name)
                unnamed_cols = [col for col in df.columns if 'Unnamed' in col]
                if unnamed_cols:
                    self.logger.warning(f"Dropping Unnamed columns from choice_data: {unnamed_cols}")
                    df = df.drop(columns=unnamed_cols)
                if df.empty:
                    print("Warning: The selected file is empty.")
                    return

                self.model = PandasModel(df)
                if self.table is None:
                    self.table = QTableView()
                    self.table.setModel(self.model)
                    self.table.setSelectionMode(QAbstractItemView.SingleSelection)
                    self.table.setSelectionBehavior(QAbstractItemView.SelectColumns)
                    self.table.resizeColumnsToContents()
                    self.table.setEditTriggers(
                        QAbstractItemView.DoubleClicked |
                        QAbstractItemView.SelectedClicked |
                        QAbstractItemView.EditKeyPressed
                    )
                    self.table.verticalHeader().setVisible(False)
                    self.table.setContextMenuPolicy(Qt.CustomContextMenu)
                    self.table.customContextMenuRequested.connect(self.show_context_menu)
                    self.grid_layout.addWidget(self.table, 6, 1, 1, 3)
                else:
                    self.table.setModel(self.model)
                    self.table.resizeColumnsToContents()
                self.table.setVisible(True)
                self.btn_simulate.setEnabled(True)
            except Exception as e:
                print(f"Error loading file: {e}")
                self.prof_xlsx.setText("Select profile file")
                self.btn_simulate.setEnabled(False)

    def add_actions_dataview(self):
        add_row_button = QPushButton("Add Profile")
        delete_row_button = QPushButton("Delete Selected Profile")
        add_column_button = QPushButton("Add Price")
        delete_column_button = QPushButton("Delete Selected Price")
        copy_column_button = QPushButton("Copy Selected Price")
        save_button = QPushButton("Save Changes")
        saveas_button = QPushButton("Save Changes As")
        add_row_button.setStyleSheet("font-size: 10pt; height: 30px; width: 150px; background-color: black; color: white; margin: 0px;")
        delete_row_button.setStyleSheet(add_row_button.styleSheet())
        add_column_button.setStyleSheet(add_row_button.styleSheet())
        delete_column_button.setStyleSheet(add_row_button.styleSheet())
        copy_column_button.setStyleSheet(add_row_button.styleSheet())
        save_button.setStyleSheet(add_row_button.styleSheet())
        saveas_button.setStyleSheet(add_row_button.styleSheet())

        add_row_button.clicked.connect(self.add_row)
        delete_row_button.clicked.connect(self.delete_row)
        add_column_button.clicked.connect(self.add_column)
        delete_column_button.clicked.connect(self.delete_column)
        copy_column_button.clicked.connect(self.copy_column)
        save_button.clicked.connect(self.save_changes)
        saveas_button.clicked.connect(self.save_as)

        self.grid_layout.addWidget(add_row_button, 4, 1, Qt.AlignTop | Qt.AlignLeft)
        self.grid_layout.addWidget(delete_row_button, 4, 2, Qt.AlignTop | Qt.AlignLeft)
        self.grid_layout.addWidget(add_column_button, 4, 3, Qt.AlignTop | Qt.AlignLeft)
        self.grid_layout.addWidget(delete_column_button, 4, 4, Qt.AlignTop | Qt.AlignLeft)
        self.grid_layout.addWidget(copy_column_button, 5, 1, Qt.AlignTop | Qt.AlignLeft)
        self.grid_layout.addWidget(save_button, 5, 2, Qt.AlignTop | Qt.AlignLeft)
        self.grid_layout.addWidget(saveas_button, 5, 3, Qt.AlignTop | Qt.AlignLeft)
        empty_widget = QLabel("xyz")
        self.grid_layout.addWidget(empty_widget, 7, 0)
        empty_widget.setVisible(False)

    def show_context_menu(self, position):
        menu = QMenu()
        delete_row_action = menu.addAction("Delete Row")
        delete_row_action.triggered.connect(self.delete_row_context)
        menu.exec(self.table.mapToGlobal(position))

    def delete_row_context(self):
        current_index = self.table.indexAt(self.table.viewport().mapFromGlobal(QCursor.pos()))
        if not current_index.isValid():
            QMessageBox.warning(self, "Warning", "No cell selected")
            return
        row = current_index.row()
        reply = QMessageBox.question(
            self, "Confirm Delete", f"Delete row {row + 1}?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.model.delete_row(row)
            self.table.resizeColumnsToContents()

    def add_row(self):
        self.model.add_row()
        self.table.resizeColumnsToContents()

    def delete_row(self):
        current_index = self.table.currentIndex()
        if not current_index.isValid():
            QMessageBox.warning(self, "Warning", "No cell selected")
            return
        row = current_index.row()
        reply = QMessageBox.question(
            self, "Confirm Delete", f"Delete row {row + 1}?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.model.delete_row(row)
            self.table.resizeColumnsToContents()

    def add_column(self):
        column_name, ok = QInputDialog.getText(self, "Add Price", "Enter Price header:")
        if ok and column_name.strip():
            self.model.add_column(column_name)
            self.table.resizeColumnsToContents()

    def delete_column(self):
        selected = self.table.selectionModel().selectedColumns()
        if not selected:
            QMessageBox.warning(self, "Warning", "No Price selected")
            return
        column = selected[0].column()
        reply = QMessageBox.question(
            self, "Confirm Delete", f"Delete '{self.model._data.columns[column]}'?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.model.delete_column(column)
            self.table.resizeColumnsToContents()

    def copy_column(self):
        selected = self.table.selectionModel().selectedColumns()
        if not selected:
            QMessageBox.warning(self, "Warning", "No column selected")
            return
        column = selected[0].column()
        column_name, ok = QInputDialog.getText(
            self, "Copy Price",
            f"Enter new name for copied Price '{self.model._data.columns[column]}':",
            text=f"{self.model._data.columns[column]}_Copy"
        )
        if ok and column_name.strip():
            success, new_name = self.model.copy_column(column, column_name)
            if success:
                self.table.resizeColumnsToContents()
                QMessageBox.information(self, "Success", f"Price copied as '{new_name}'")
            else:
                QMessageBox.warning(self, "Error", new_name)

    def save_changes(self):
        updated_data = self.model.get_data()
        updated_data.to_excel("updated_profiles.xlsx", index=False)
        QMessageBox.information(self, "Success", "Changes saved")

    def save_as(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save the Price Profile As", "", "CSV Files (*.csv);;Excel Files (*.xlsx)"
        )
        if file_path:
            success, message = self.model.save_data(file_path)
            if success:
                QMessageBox.information(self, "Success", f"Data saved as '{file_path}'")
            else:
                QMessageBox.critical(self, "Error", f"Failed to save data: {message}")
        else:
            self.update_status("Saving the price profile as cancelled")

    def create_profile_share_chart(self, input_file="simulated_share.xlsx", data_sheet="Profile Share"):
        try:
            self.logger.info(f"Creating profile share chart for {input_file}")
            workbook = openpyxl.load_workbook(input_file)
            data_worksheet = workbook[data_sheet]

            if "Profile Share Plot" in workbook.sheetnames:
                workbook.remove(workbook["Profile Share Plot"])
            plot_worksheet = workbook.create_sheet("Profile Share Plot")

            headers = [cell.value for cell in data_worksheet[1] if cell.value]
            if "Profile" not in headers:
                raise ValueError("Expected 'Profile' column in the data")
            price_columns = [i + 1 for i, header in enumerate(headers) if header != "Profile"]

            categories = Reference(data_worksheet, min_col=1, min_row=2, max_row=data_worksheet.max_row)

            chart = LineChart()
            chart.title = "Simulated Profile Shares by Profile"
            chart.style = 10
            chart.x_axis.title = "Profile"
            chart.y_axis.title = "Profile Share (%)"

            for col in price_columns:
                data = Reference(data_worksheet, min_col=col, min_row=1, max_row=data_worksheet.max_row)
                chart.add_data(data, titles_from_data=True)
            for series in chart.series:
                series.graphicalProperties.line = LineProperties(w=0.75 * 12700)
            chart.set_categories(categories)

            plot_worksheet.add_chart(chart, "A1")
            workbook.save(input_file)
            self.logger.info(f"Profile share chart saved to {input_file}")
            for handler in self.logger.handlers:
                if isinstance(handler, logging.FileHandler):
                    handler.flush()

        except Exception as e:
            self.logger.error(f"Error creating profile share chart: {str(e)}")
            raise

    def create_value_share_chart(self, input_file="simulated_share.xlsx", data_sheet="Value Share"):
        try:
            self.logger.info(f"Creating value share chart for {input_file}")
            workbook = openpyxl.load_workbook(input_file)
            data_worksheet = workbook[data_sheet]

            if "Value Share Plot" in workbook.sheetnames:
                workbook.remove(workbook["Value Share Plot"])
            plot_worksheet = workbook.create_sheet("Value Share Plot")

            headers = [cell.value for cell in data_worksheet[1] if cell.value]
            if "Profile" not in headers:
                raise ValueError("Expected 'Profile' column in the data")
            price_columns = [i + 1 for i, header in enumerate(headers) if header != "Profile"]

            categories = Reference(data_worksheet, min_col=1, min_row=2, max_row=data_worksheet.max_row)

            chart = LineChart()
            chart.title = "Simulated Value Shares by Profile"
            chart.style = 10
            chart.x_axis.title = "Profile"
            chart.y_axis.title = "Value Share (%)"

            for col in price_columns:
                data = Reference(data_worksheet, min_col=col, min_row=1, max_row=data_worksheet.max_row)
                chart.add_data(data, titles_from_data=True)
            for series in chart.series:
                series.graphicalProperties.line = LineProperties(w=0.75 * 12700)
            chart.set_categories(categories)

            plot_worksheet.add_chart(chart, "A1")
            workbook.save(input_file)
            self.logger.info(f"Value share chart saved to {input_file}")
            for handler in self.logger.handlers:
                if isinstance(handler, logging.FileHandler):
                    handler.flush()

        except Exception as e:
            self.logger.error(f"Error creating value share chart: {str(e)}")
            raise

def main():
    multiprocessing.freeze_support()
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps)
    app = QApplication(sys.argv)
    window = DCMApp()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()